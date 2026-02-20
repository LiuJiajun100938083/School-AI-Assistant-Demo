"""创建学生批量上传Excel模板"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "學生名單"

    # 表头样式
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill('solid', fgColor='006633')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头
    headers = ['登入名稱', '密碼', '顯示名稱', '班級']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 示例数据样式
    example_fill = PatternFill('solid', fgColor='FFF2CC')
    example_font = Font(italic=True, color='666666')

    # 示例数据
    examples = [
        ('student001', 'Pass1234!', '張三', '1A'),
        ('student002', 'Pass1234!', '李四', '1A'),
        ('student003', 'Pass1234!', '王五', '1B'),
    ]

    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # 设置列宽
    col_widths = [20, 18, 20, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 添加说明工作表
    ws_help = wb.create_sheet("使用說明")
    instructions = [
        ("批量添加學生 - 使用說明", ""),
        ("", ""),
        ("欄位說明：", ""),
        ("登入名稱", "學生用於登入系統的帳號（必填，不可重複）"),
        ("密碼", "初始密碼，建議使用強密碼（必填，至少8字元）"),
        ("顯示名稱", "學生的真實姓名或暱稱（必填）"),
        ("班級", "學生所屬班級，如 1A, 2B 等（選填）"),
        ("", ""),
        ("注意事項：", ""),
        ("1.", "請刪除黃色示例行再填入真實數據"),
        ("2.", "登入名稱只能包含英文字母、數字和下劃線"),
        ("3.", "密碼必須包含大小寫字母和數字，至少8個字元"),
        ("4.", "每行一個學生，不要留空行"),
        ("5.", "請勿修改第一行的表頭"),
    ]

    for row_idx, (col1, col2) in enumerate(instructions, 1):
        ws_help.cell(row=row_idx, column=1, value=col1)
        ws_help.cell(row=row_idx, column=2, value=col2)
        if row_idx == 1:
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True, size=14, color='006633')
        elif col1 in ("欄位說明：", "注意事項："):
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True, size=11)

    ws_help.column_dimensions['A'].width = 15
    ws_help.column_dimensions['B'].width = 50

    output_path = '/sessions/inspiring-wonderful-ride/mnt/FastAPIProject1/web_static/templates/student_upload_template.xlsx'
    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✓ 模板已創建: {output_path}")
    return output_path

if __name__ == "__main__":
    create_template()
