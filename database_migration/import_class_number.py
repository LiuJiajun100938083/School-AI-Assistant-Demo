#!/usr/bin/env python3
"""
從 Excel 導入 class_number（班號）到 users 表

用法：
    python3 import_class_number.py <Excel路徑> [--db-user root] [--db-password ''] [--db-host localhost]

數據源：學號+班號對應.xlsx
欄位：ClassName, ClassNumber, UserLogin
匹配：UserLogin → users.username
"""

import argparse
import os
import sys

import openpyxl
import pymysql


def main():
    parser = argparse.ArgumentParser(description="從 Excel 導入 class_number 到 users 表")
    parser.add_argument("excel_path", help="Excel 文件路徑（學號+班號對應.xlsx）")
    parser.add_argument("--db-host", default="localhost", help="數據庫主機（默認 localhost）")
    parser.add_argument("--db-port", type=int, default=3306, help="數據庫端口（默認 3306）")
    parser.add_argument("--db-user", default="root", help="數據庫用戶（默認 root）")
    parser.add_argument("--db-password", default="", help="數據庫密碼（默認空）")
    parser.add_argument("--db-name", default="school_ai_assistant", help="數據庫名（默認 school_ai_assistant）")
    args = parser.parse_args()

    excel_path = os.path.abspath(args.excel_path)
    if not os.path.exists(excel_path):
        print(f"錯誤：找不到文件 {excel_path}")
        sys.exit(1)

    # 讀取 Excel
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb.active

    rows = []
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(header)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        class_name = str(row[col_map["ClassName"]]).strip() if row[col_map["ClassName"]] else ""
        class_number = row[col_map["ClassNumber"]]
        user_login = str(row[col_map["UserLogin"]]).strip() if row[col_map["UserLogin"]] else ""
        if user_login:
            rows.append((class_name, int(class_number) if class_number else None, user_login))

    wb.close()
    print(f"Excel 讀取完成：{len(rows)} 行數據")

    # 連接數據庫並更新
    conn = pymysql.connect(
        host=args.db_host,
        port=args.db_port,
        user=args.db_user,
        password=args.db_password,
        database=args.db_name,
        charset="utf8mb4",
    )
    cursor = conn.cursor()

    matched = 0
    not_found = []

    for class_name, class_number, username in rows:
        cursor.execute(
            "UPDATE users SET class_name = %s, class_number = %s WHERE username = %s",
            (class_name, class_number, username),
        )
        if cursor.rowcount > 0:
            matched += 1
        else:
            not_found.append(username)

    conn.commit()
    cursor.close()
    conn.close()

    print(f"\n匹配更新：{matched} 筆")
    print(f"未匹配（用戶不存在）：{len(not_found)} 筆")
    if not_found and len(not_found) <= 20:
        for u in not_found:
            print(f"  - {u}")
    elif not_found:
        print(f"  前 20 筆：{not_found[:20]}")
        print(f"  ...共 {len(not_found)} 筆")


if __name__ == "__main__":
    main()
