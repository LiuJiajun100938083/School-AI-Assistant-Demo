"""
考勤系統 — Excel 導出模塊

包含:
- build_morning_export_excel()   早讀點名導出
- build_detention_export_excel() 留堂導出
- build_detention_history_excel() 留堂歷史導出
- build_activity_export_excel()  課外活動導出
- calc_export_stats()            統計計算
"""

from typing import List, Optional, Tuple

from app.domains.attendance.constants import (
    ActivityCheckinStatus,
    ActivityCheckoutStatus,
    AttendanceStatus,
    DetentionReason,
    ExcelColors,
    SessionType,
)


def _excel_styles():
    """共用的 Excel 樣式"""
    from openpyxl.styles import Border, Font, PatternFill, Side

    return {
        "header_font": Font(bold=True, color=ExcelColors.WHITE),
        "header_fill": PatternFill(start_color=ExcelColors.HEADER_BLUE, end_color=ExcelColors.HEADER_BLUE, fill_type="solid"),
        "green_fill": PatternFill(start_color=ExcelColors.SUCCESS_GREEN, end_color=ExcelColors.SUCCESS_GREEN, fill_type="solid"),
        "orange_fill": PatternFill(start_color=ExcelColors.WARNING_ORANGE, end_color=ExcelColors.WARNING_ORANGE, fill_type="solid"),
        "red_fill": PatternFill(start_color=ExcelColors.ERROR_RED, end_color=ExcelColors.ERROR_RED, fill_type="solid"),
        "gray_fill": PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"),
        "homework_fill": PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),
        "morning_fill": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "both_fill": PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid"),
        "other_reason_fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }


def build_morning_export_excel(session, records):
    """生成早讀點名 Excel

    Returns
    -------
    (Workbook, filename: str)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "早读点名"

    ws["A1"] = "早读点名记录"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = f"日期: {session.get('session_date', '')}"
    ws["D2"] = f"应到时间: {session.get('target_time', '')}"

    headers = ["班级", "班号", "学号", "英文名", "中文名", "签到时间", "状态", "迟到(分钟)", "需补时(分钟)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = s["thin_border"]

    present_count = late_count = very_late_count = absent_count = 0
    for row_idx, record in enumerate(records, 5):
        status = record.get("status") or record.get("attendance_status") or AttendanceStatus.ABSENT
        if status == AttendanceStatus.PRESENT:
            present_count += 1
        elif status == AttendanceStatus.LATE:
            late_count += 1
        elif status == AttendanceStatus.VERY_LATE:
            very_late_count += 1
        else:
            absent_count += 1

        status_cn = AttendanceStatus.LABELS_ZH.get(status, "缺席")
        scan_time = record.get("scan_time")
        if scan_time and hasattr(scan_time, "strftime"):
            scan_time = scan_time.strftime("%H:%M:%S")
        elif not scan_time:
            scan_time = "-"

        row_data = [
            record.get("class_name"), record.get("class_number"), record.get("user_login"),
            record.get("english_name"), record.get("chinese_name"), scan_time, status_cn,
            record.get("late_minutes") or 0, record.get("makeup_minutes") or 0,
        ]
        fill_map = {
            AttendanceStatus.PRESENT: s["green_fill"],
            AttendanceStatus.LATE: s["orange_fill"],
            AttendanceStatus.VERY_LATE: s["red_fill"],
        }
        fill = fill_map.get(status, s["gray_fill"])
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            if status == AttendanceStatus.VERY_LATE:
                cell.font = Font(color=ExcelColors.WHITE)

    summary_row = len(records) + 6
    ws.cell(row=summary_row, column=1, value="统计").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"应到: {len(records)}")
    ws.cell(row=summary_row, column=3, value=f"准时: {present_count}").fill = s["green_fill"]
    ws.cell(row=summary_row, column=4, value=f"迟到: {late_count + very_late_count}").fill = s["orange_fill"]
    ws.cell(row=summary_row, column=5, value=f"缺席: {absent_count}").fill = s["gray_fill"]

    for col, w in enumerate([8, 6, 15, 20, 15, 12, 14, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    filename = f"早读点名_{session.get('session_date', '')}.xlsx"
    return wb, filename


def build_detention_export_excel(session, records):
    """生成留堂 Excel

    Returns
    -------
    (Workbook, filename: str)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "留堂记录"

    ws["A1"] = "留堂记录"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")
    ws["A2"] = f"日期: {session.get('session_date', '')}"

    headers = ["班级", "班号", "学号", "英文名", "中文名", "签到时间", "签退时间",
               "计划时长", "实际分钟", "实际节数", "状态", "原因"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = s["thin_border"]

    completed_count = incomplete_count = absent_count = 0
    for row_idx, record in enumerate(records, 5):
        status = record.get("status") or AttendanceStatus.ABSENT
        fill = None
        if status == AttendanceStatus.DETENTION_COMPLETED:
            planned_mins = record.get("planned_minutes")
            actual_mins = record.get("actual_minutes") or 0
            planned_p = record.get("planned_periods") or 0
            actual_p = record.get("actual_periods") or 0
            is_done = (actual_mins >= planned_mins) if planned_mins is not None else (actual_p >= planned_p)
            if is_done:
                status_cn, fill = "已完成", s["green_fill"]
                completed_count += 1
            else:
                status_cn, fill = "未完成", s["orange_fill"]
                incomplete_count += 1
        elif status == AttendanceStatus.DETENTION_ACTIVE:
            status_cn = "进行中"
        else:
            status_cn, fill = "未签到", s["gray_fill"]
            absent_count += 1

        scan_t = record.get("scan_time")
        scan_t = scan_t.strftime("%H:%M:%S") if scan_t and hasattr(scan_t, "strftime") else "-"
        out_t = record.get("checkout_time")
        out_t = out_t.strftime("%H:%M:%S") if out_t and hasattr(out_t, "strftime") else "-"

        reason = record.get("detention_reason") or ""
        reason_display = DetentionReason.LABELS_ZH.get(reason, reason or "未知")

        planned_mins_display = record.get("planned_minutes")
        planned_display = (f"{planned_mins_display}分钟" if planned_mins_display is not None
                           else f"{record.get('planned_periods') or 0}节")

        row_data = [
            record.get("class_name"), record.get("class_number"), record.get("user_login"),
            record.get("english_name"), record.get("chinese_name"), scan_t, out_t,
            planned_display, record.get("actual_minutes") or 0, record.get("actual_periods") or 0,
            status_cn, reason_display,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")
            if col <= 11 and fill:
                cell.fill = fill
            elif col == 12:
                if reason == DetentionReason.HOMEWORK:
                    cell.fill = s["homework_fill"]
                elif reason == DetentionReason.MORNING:
                    cell.fill = s["morning_fill"]
                elif reason == DetentionReason.BOTH:
                    cell.fill = s["both_fill"]
                elif reason:
                    cell.fill = s["other_reason_fill"]

    summary_row = len(records) + 6
    ws.cell(row=summary_row, column=1, value="统计").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"总人数: {len(records)}")
    ws.cell(row=summary_row, column=3, value=f"已完成: {completed_count}").fill = s["green_fill"]
    ws.cell(row=summary_row, column=4, value=f"未完成: {incomplete_count}").fill = s["orange_fill"]
    ws.cell(row=summary_row, column=5, value=f"未签到: {absent_count}").fill = s["gray_fill"]

    for col, w in enumerate([8, 6, 15, 20, 15, 12, 12, 10, 10, 10, 12, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    filename = f"留堂记录_{session.get('session_date', '')}.xlsx"
    return wb, filename


def build_detention_history_excel(history):
    """生成留堂歷史 Excel

    Returns
    -------
    Workbook
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "留堂历史"

    headers = ["日期", "班级", "班号", "学号", "英文名", "中文名", "时长(分钟)", "原因", "已完成", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = s["thin_border"]

    for row_idx, record in enumerate(history, 2):
        reason = record.get("reason") or ""
        reason_display = DetentionReason.LABELS_ZH.get(reason, reason or "未知")
        row_data = [
            str(record.get("detention_date", "")), record.get("class_name"), record.get("class_number"),
            record.get("user_login"), record.get("english_name"), record.get("chinese_name"),
            record.get("duration_minutes"), reason_display,
            "是" if record.get("completed") else "否", record.get("notes") or "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")
            if col == 8:
                if reason == DetentionReason.HOMEWORK:
                    cell.fill = s["homework_fill"]
                elif reason == DetentionReason.MORNING:
                    cell.fill = s["morning_fill"]
                elif reason == DetentionReason.BOTH:
                    cell.fill = s["both_fill"]
                elif reason:
                    cell.fill = s["other_reason_fill"]

    for col, w in enumerate([12, 8, 6, 15, 20, 15, 12, 30, 8, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return wb


def build_activity_export_excel(session, records):
    """生成課外活動 Excel

    Returns
    -------
    Workbook
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "课外活动点名"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{session.get('activity_name', '')} - {session.get('session_date', '')}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = f"应到时间: {session.get('start_time', '')} | 结束时间: {session.get('end_time', '')}"

    headers = ["班级", "学号", "姓名", "签到时间", "签到状态", "签退时间", "签退状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    checkin_colors = {
        ActivityCheckinStatus.ON_TIME: "34C759",
        ActivityCheckinStatus.LATE: "FF9500",
        ActivityCheckinStatus.NOT_ARRIVED: "FF3B30",
    }
    checkout_colors = {
        ActivityCheckoutStatus.NORMAL: "34C759",
        ActivityCheckoutStatus.EARLY: "FF9500",
        ActivityCheckoutStatus.STILL_HERE: "007AFF",
        ActivityCheckoutStatus.NOT_ARRIVED: "8E8E93",
    }
    checkin_text = {
        ActivityCheckinStatus.ON_TIME: "准时",
        ActivityCheckinStatus.LATE: "迟到",
        ActivityCheckinStatus.NOT_ARRIVED: "未到",
    }
    checkout_text = {
        ActivityCheckoutStatus.NORMAL: "正常",
        ActivityCheckoutStatus.EARLY: "早退",
        ActivityCheckoutStatus.STILL_HERE: "仍在",
        ActivityCheckoutStatus.NOT_ARRIVED: "-",
    }

    for row_idx, record in enumerate(records, 5):
        ws.cell(row=row_idx, column=1, value=f"{record.get('class_name', '')}-{record.get('class_number', '')}")
        ws.cell(row=row_idx, column=2, value=record.get("user_login"))
        ws.cell(row=row_idx, column=3, value=record.get("chinese_name"))

        ci_time = record.get("check_in_time")
        ws.cell(row=row_idx, column=4, value=ci_time.strftime("%H:%M:%S") if ci_time else "-")

        ci_status = record.get("check_in_status", ActivityCheckinStatus.NOT_ARRIVED)
        ci_cell = ws.cell(row=row_idx, column=5, value=checkin_text.get(ci_status, "-"))
        color = checkin_colors.get(ci_status, ExcelColors.WHITE)
        ci_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        co_time = record.get("check_out_time")
        ws.cell(row=row_idx, column=6, value=co_time.strftime("%H:%M:%S") if co_time else "-")

        co_status = record.get("check_out_status", ActivityCheckoutStatus.NOT_ARRIVED)
        co_cell = ws.cell(row=row_idx, column=7, value=checkout_text.get(co_status, "-"))
        color = checkout_colors.get(co_status, ExcelColors.WHITE)
        co_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.cell(row=row_idx, column=8, value=record.get("notes", ""))

    for letter, w in zip("ABCDEFGH", [12, 15, 15, 12, 10, 12, 10, 20]):
        ws.column_dimensions[letter].width = w

    return wb


def calc_export_stats(session_type, records):
    """計算導出統計數據"""
    student_count = len(records)
    present_count = late_count = absent_count = 0

    if session_type == SessionType.MORNING:
        for r in records:
            st = r.get("status") or r.get("attendance_status") or AttendanceStatus.ABSENT
            if st in (AttendanceStatus.PRESENT, AttendanceStatus.LATE, AttendanceStatus.VERY_LATE):
                present_count += 1
            if st in (AttendanceStatus.LATE, AttendanceStatus.VERY_LATE):
                late_count += 1
            if st == AttendanceStatus.ABSENT:
                absent_count += 1
    else:
        for r in records:
            st = r.get("status") or AttendanceStatus.ABSENT
            if st in (AttendanceStatus.DETENTION_COMPLETED, AttendanceStatus.DETENTION_ACTIVE):
                present_count += 1
            if st == AttendanceStatus.DETENTION_COMPLETED:
                pm = r.get("planned_minutes")
                am = r.get("actual_minutes") or 0
                pp = r.get("planned_periods") or 0
                ap = r.get("actual_periods") or 0
                if (pm is not None and am < pm) or (pm is None and ap < pp):
                    late_count += 1
            if st not in (AttendanceStatus.DETENTION_COMPLETED, AttendanceStatus.DETENTION_ACTIVE):
                absent_count += 1

    return {
        "student_count": student_count,
        "present_count": present_count,
        "late_count": late_count,
        "absent_count": absent_count,
    }
