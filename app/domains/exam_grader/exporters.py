"""
试卷批阅 — Excel 匯出
======================
匯出全班成績表和單個學生報告為 xlsx 格式。
"""

import io
import logging
from typing import Any, Dict, List

logger = logging.getLogger(__name__)


def export_class_report(
    exam: Dict[str, Any],
    papers: List[Dict[str, Any]],
    questions: List[Dict[str, Any]],
    per_question_stats: List[Dict[str, Any]],
    stats: Dict[str, Any],
) -> bytes:
    """匯出全班成績表 xlsx"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.warning("openpyxl 未安裝，無法匯出 xlsx")
        return b""

    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    title_font = Font(bold=True, size=14)

    total_max = sum(float(q.get("max_marks", 0)) for q in questions)

    # ── Sheet 1: 成績總覽 ──
    ws1 = wb.active
    ws1.title = "成績總覽"
    ws1.append([exam.get("title", "考試")])
    ws1["A1"].font = title_font
    ws1.append([])
    info = [
        ("科目", exam.get("subject", "")),
        ("班級", exam.get("class_name", "")),
        ("總分", total_max),
        ("考生人數", stats.get("total_students", len(papers))),
        ("平均分", stats.get("average_score")),
        ("最高分", stats.get("highest_score")),
        ("最低分", stats.get("lowest_score")),
        ("標準差", stats.get("std_deviation")),
    ]
    for label, val in info:
        ws1.append([label, val])
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: 學生成績 ──
    ws2 = wb.create_sheet("學生成績")
    headers = ["#", "姓名", "學號", "班別", "總分", "滿分", "百分比"]

    # 每題一列 header
    for q in questions:
        sec = q.get("section", "")
        num = q.get("question_number", "")
        headers.append(f"{sec}{num}")

    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for idx, p in enumerate(papers, 1):
        score = p.get("total_score") or 0
        pct = round(score / total_max * 100, 1) if total_max > 0 else 0
        row = [
            idx,
            p.get("student_name", ""),
            p.get("student_number", ""),
            p.get("class_name", ""),
            score,
            total_max,
            f"{pct}%",
        ]
        # 每題得分
        answers_map = p.get("_answers_map", {})
        for q in questions:
            qid = q["id"]
            ans = answers_map.get(qid)
            row.append(ans.get("score", 0) if ans else 0)
        ws2.append(row)

    # 列寬
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 8
    ws2.column_dimensions["F"].width = 8
    ws2.column_dimensions["G"].width = 10

    # ── Sheet 3: 題目分析 ──
    ws3 = wb.create_sheet("題目分析")
    ws3.append(["題號", "部分", "類型", "滿分", "平均得分", "正確率 (MC)"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for pqs in per_question_stats:
        q_type = pqs.get("question_type", "")
        cr = pqs.get("correct_rate")
        ws3.append([
            pqs.get("question_number", ""),
            pqs.get("section", ""),
            "選擇題" if q_type == "mc" else "簡答題",
            pqs.get("max_marks", 0),
            pqs.get("average_score", 0),
            f"{cr}%" if cr is not None else "-",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_student_report(
    exam: Dict[str, Any],
    paper: Dict[str, Any],
    answers: List[Dict[str, Any]],
) -> bytes:
    """匯出單個學生報告 xlsx"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.warning("openpyxl 未安裝，無法匯出 xlsx")
        return b""

    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    title_font = Font(bold=True, size=14)

    ws = wb.active
    ws.title = "成績報告"

    # 標題
    ws.append([exam.get("title", "考試")])
    ws["A1"].font = title_font
    ws.append([])

    # 學生資訊
    ws.append(["姓名", paper.get("student_name", "")])
    ws.append(["學號", paper.get("student_number", "")])
    ws.append(["班別", paper.get("class_name", "")])
    ws.append(["總分", paper.get("total_score", 0)])
    ws.append([])

    # 答題詳情
    detail_headers = ["題號", "部分", "類型", "題目", "學生答案", "參考答案", "得分", "滿分", "回饋"]
    ws.append(detail_headers)
    for cell in ws[8]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for a in answers:
        q_type = a.get("question_type", "")
        ws.append([
            a.get("question_number", ""),
            a.get("section", ""),
            "選擇題" if q_type == "mc" else "簡答題",
            (a.get("question_text") or a.get("question_content") or "")[:200],
            a.get("student_answer", ""),
            a.get("reference_answer") or a.get("correct_answer") or "",
            a.get("score", 0),
            a.get("max_marks") or a.get("max_score") or 0,
            a.get("feedback", ""),
        ])

    # 列寬
    for col, w in [("A", 6), ("B", 6), ("C", 8), ("D", 40),
                   ("E", 30), ("F", 30), ("G", 6), ("H", 6), ("I", 30)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
