"""
作業管理 — Excel 導出模塊

包含:
- build_grade_export_excel()  成績匯出
- build_plagiarism_export_excel()  抄袭檢測報告匯出
"""

import json
import math
from typing import List


# ================================================================
# 通用樣式
# ================================================================

def _styles_grade():
    """成績匯出樣式集"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    return {
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "title_font": Font(bold=True, size=14),
        "subtitle_font": Font(bold=True, size=11, color="4472C4"),
        "thin_border": thin,
        "center": Alignment(horizontal="center", vertical="center"),
        "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "gray_fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "green_fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "red_fill": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "yellow_fill": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        "light_blue_fill": PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid"),
    }


def _styles_plagiarism():
    """抄袭檢測匯出樣式集"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    return {
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "header_fill": PatternFill(start_color="D35400", end_color="D35400", fill_type="solid"),
        "title_font": Font(bold=True, size=14),
        "thin_border": thin,
        "center": Alignment(horizontal="center", vertical="center"),
        "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "red_fill": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "yellow_fill": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    }


# ================================================================
# 成績匯出
# ================================================================

def build_grade_export_excel(
    assignment: dict,
    submissions: List[dict],
    rubric_items: List[dict],
):
    """生成成績匯出 Excel (兩個 Sheet: 成績表 + 成績分析)

    Returns
    -------
    (Workbook, filename: str)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    s = _styles_grade()
    wb = Workbook()

    # ============================
    # Sheet 1: 成績表
    # ============================
    ws1 = wb.active
    ws1.title = "成績表"

    max_score = assignment.get("max_score") or 100
    ws1["A1"] = f"{assignment.get('title', '作業')} - 成績表"
    ws1["A1"].font = s["title_font"]
    ws1["A2"] = f"滿分: {max_score}"
    ws1["A2"].font = Font(size=11, color="666666")

    # Headers
    headers = ["學號", "姓名", "班級", "總分"]
    for ri in rubric_items:
        title = ri.get("title", "")
        mp = ri.get("max_points")
        headers.append(f"{title} ({mp})" if mp else title)
    headers.append("評語")
    headers.append("狀態")

    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["thin_border"]

    # rubric_item_id → index
    rubric_id_to_idx = {ri["id"]: idx for idx, ri in enumerate(rubric_items)}

    sorted_subs = sorted(submissions, key=lambda sub: sub.get("username") or "")

    graded_scores = []
    for row_idx, sub in enumerate(sorted_subs, header_row + 1):
        is_graded = sub.get("status") == "graded"
        score = sub.get("score")

        row_data = [
            sub.get("username") or "",
            sub.get("student_name") or "",
            sub.get("class_name") or "",
            float(score) if score is not None else "",
        ]

        score_map = {}
        for rs in sub.get("rubric_scores") or []:
            rid = rs.get("rubric_item_id")
            if rid in rubric_id_to_idx:
                pts = rs.get("points")
                score_map[rubric_id_to_idx[rid]] = float(pts) if pts is not None else ""

        for i in range(len(rubric_items)):
            row_data.append(score_map.get(i, ""))

        row_data.append(sub.get("feedback") or "")
        status_text = {"graded": "已批改", "submitted": "待批改", "returned": "已退回"}.get(
            sub.get("status", ""), sub.get("status", "")
        )
        row_data.append(status_text)

        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.border = s["thin_border"]
            cell.alignment = s["center"] if col != len(headers) - 1 else s["left_wrap"]
            if not is_graded:
                cell.fill = s["gray_fill"]

        if is_graded and score is not None:
            graded_scores.append(float(score))

    # Column widths
    col_widths = [12, 12, 10, 8] + [12] * len(rubric_items) + [30, 8]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    ws1.freeze_panes = f"A{header_row + 1}"

    # ============================
    # Sheet 2: 成績分析
    # ============================
    ws2 = wb.create_sheet(title="成績分析")
    ws2["A1"] = f"{assignment.get('title', '作業')} - 成績分析"
    ws2["A1"].font = s["title_font"]
    ws2.merge_cells("A1:D1")

    _build_stats_section(ws2, s, submissions, graded_scores, rubric_items, max_score)

    # Column widths
    for col_letter, w in zip("ABCD", [16, 12, 12, 12]):
        ws2.column_dimensions[col_letter].width = w

    title = assignment.get("title", "作業") or "作業"
    filename = f"{title}_成績.xlsx"
    return wb, filename


def _build_stats_section(ws, s, submissions, graded_scores, rubric_items, max_score):
    """填充成績分析 sheet"""
    from openpyxl.styles import Font
    row = 3
    ws.cell(row=row, column=1, value="基本統計").font = s["subtitle_font"]
    row += 1

    total_submitted = len(submissions)
    graded_count = len(graded_scores)

    if graded_scores:
        avg_score = sum(graded_scores) / len(graded_scores)
        sorted_scores = sorted(graded_scores)
        n = len(sorted_scores)
        median_score = (sorted_scores[n // 2] + sorted_scores[(n - 1) // 2]) / 2
        max_s = max(graded_scores)
        min_s = min(graded_scores)
        variance = sum((x - avg_score) ** 2 for x in graded_scores) / len(graded_scores)
        std_dev = math.sqrt(variance)
    else:
        avg_score = median_score = max_s = min_s = std_dev = 0

    stats = [
        ("提交人數", total_submitted),
        ("已批改數", graded_count),
        ("平均分", round(avg_score, 1)),
        ("中位數", round(median_score, 1)),
        ("最高分", round(max_s, 1) if graded_scores else "-"),
        ("最低分", round(min_s, 1) if graded_scores else "-"),
        ("標準差", round(std_dev, 1)),
        ("滿分", max_score),
    ]

    for label, val in stats:
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = Font(bold=True)
        cell_label.border = s["thin_border"]
        cell_label.fill = s["light_blue_fill"]
        cell_val = ws.cell(row=row, column=2, value=val)
        cell_val.border = s["thin_border"]
        cell_val.alignment = s["center"]
        row += 1

    # -- 分數段分佈 --
    row += 1
    ws.cell(row=row, column=1, value="分數段分佈").font = s["subtitle_font"]
    row += 1

    for col, h in enumerate(["分數段", "人數", "百分比"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["thin_border"]
    row += 1

    max_score_f = float(max_score) if max_score else 100
    ranges = [
        ("90-100%", 0.9 * max_score_f, max_score_f),
        ("80-89%", 0.8 * max_score_f, 0.9 * max_score_f),
        ("70-79%", 0.7 * max_score_f, 0.8 * max_score_f),
        ("60-69%", 0.6 * max_score_f, 0.7 * max_score_f),
        ("0-59%", 0, 0.6 * max_score_f),
    ]
    range_fills = [s["green_fill"], s["green_fill"], s["yellow_fill"], s["yellow_fill"], s["red_fill"]]

    for (label, low, high), fill in zip(ranges, range_fills):
        if label == "90-100%":
            count = sum(1 for sc in graded_scores if low <= sc <= high)
        else:
            count = sum(1 for sc in graded_scores if low <= sc < high)
        pct = f"{count / graded_count * 100:.1f}%" if graded_count else "0%"

        ws.cell(row=row, column=1, value=label).border = s["thin_border"]
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=2, value=count).border = s["thin_border"]
        ws.cell(row=row, column=2).alignment = s["center"]
        ws.cell(row=row, column=3, value=pct).border = s["thin_border"]
        ws.cell(row=row, column=3).alignment = s["center"]
        row += 1

    # -- 各評分項平均 --
    if rubric_items:
        row += 1
        ws.cell(row=row, column=1, value="各評分項統計").font = s["subtitle_font"]
        row += 1

        for col, h in enumerate(["評分項", "滿分", "平均分", "得分率"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = s["header_font"]
            cell.fill = s["header_fill"]
            cell.alignment = s["center"]
            cell.border = s["thin_border"]
        row += 1

        for ri in rubric_items:
            ri_max = float(ri.get("max_points") or 0)
            all_pts = []
            for sub in submissions:
                if sub.get("status") != "graded":
                    continue
                for rs in sub.get("rubric_scores") or []:
                    if rs.get("rubric_item_id") == ri["id"] and rs.get("points") is not None:
                        all_pts.append(float(rs["points"]))

            ri_avg = sum(all_pts) / len(all_pts) if all_pts else 0
            ri_rate = f"{ri_avg / ri_max * 100:.1f}%" if ri_max > 0 else "-"

            ws.cell(row=row, column=1, value=ri.get("title", "")).border = s["thin_border"]
            ws.cell(row=row, column=2, value=ri_max).border = s["thin_border"]
            ws.cell(row=row, column=2).alignment = s["center"]
            ws.cell(row=row, column=3, value=round(ri_avg, 1)).border = s["thin_border"]
            ws.cell(row=row, column=3).alignment = s["center"]
            cell_rate = ws.cell(row=row, column=4, value=ri_rate)
            cell_rate.border = s["thin_border"]
            cell_rate.alignment = s["center"]
            row += 1


# ================================================================
# 抄袭檢測報告匯出
# ================================================================

def build_plagiarism_export_excel(
    assignment: dict,
    report: dict,
    pairs: List[dict],
    clusters: List[dict],
    hub_students: List[dict],
):
    """生成抄袭檢測報告 Excel (兩個 Sheet: 配對明細 + 群組分析)

    Returns
    -------
    (Workbook, filename: str)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    s = _styles_plagiarism()
    wb = Workbook()

    assignment_title = assignment.get("title", "作業") if assignment else "作業"

    # ============================
    # Sheet 1: 配對明細
    # ============================
    ws1 = wb.active
    ws1.title = "配對明細"

    ws1["A1"] = f"{assignment_title} - 抄袭檢測報告"
    ws1["A1"].font = s["title_font"]
    ws1["A2"] = (
        f"閾值: {report.get('threshold', 60)}% · "
        f"總配對: {report.get('total_pairs', 0)} · "
        f"可疑: {report.get('flagged_pairs', 0)} · "
        f"檢測時間: {report.get('created_at', '-')}"
    )
    ws1["A2"].font = Font(size=11, color="666666")

    headers = ["學生A", "學生B", "綜合分數(%)", "結構分", "標識符分", "逐字分", "注釋分", "是否可疑", "信號", "AI 分析"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=h)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["thin_border"]

    sorted_pairs = sorted(pairs, key=lambda p: float(p.get("similarity_score", 0)), reverse=True)

    for row_idx, p in enumerate(sorted_pairs, header_row + 1):
        is_flagged = p.get("is_flagged", False)

        frags = p.get("matched_fragments") or []
        if isinstance(frags, str):
            try:
                frags = json.loads(frags)
            except (json.JSONDecodeError, TypeError):
                frags = []
        dim = None
        signals_text = ""
        for f in frags:
            if isinstance(f, dict) and f.get("type") == "dimension_breakdown":
                dim = f
                signals_text = ", ".join(f.get("signals", []))
                break

        row_data = [
            p.get("student_a_name", ""),
            p.get("student_b_name", ""),
            float(p.get("similarity_score", 0)),
            dim.get("structure_score", "") if dim else "",
            dim.get("identifier_score", "") if dim else "",
            dim.get("verbatim_score", "") if dim else "",
            dim.get("comment_score", "") if dim else "",
            "是" if is_flagged else "否",
            signals_text,
            (p.get("ai_analysis") or "")[:500],
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = s["thin_border"]
            cell.alignment = s["left_wrap"] if col_idx >= 9 else s["center"]

        if is_flagged:
            for col_idx in range(1, len(row_data) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = s["red_fill"]

    col_widths = [14, 14, 12, 10, 10, 10, 10, 10, 24, 40]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = w

    # ============================
    # Sheet 2: 群組分析
    # ============================
    ws2 = wb.create_sheet("群組分析")

    ws2["A1"] = "抄襲群組分析"
    ws2["A1"].font = s["title_font"]

    if clusters:
        row = 3
        g_headers = ["群組", "人數", "最高相似度(%)", "疑似源頭", "成員"]
        for col, h in enumerate(g_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = s["header_font"]
            cell.fill = s["header_fill"]
            cell.alignment = s["center"]
            cell.border = s["thin_border"]

        for c in clusters:
            row += 1
            source = c.get("source_student") or "-"
            members = ", ".join(m.get("name", "") for m in c.get("members", []))
            for col_idx, val in enumerate(
                [f"群組 {c.get('id', '')}", c.get("size", 0), c.get("max_score", 0), source, members], 1
            ):
                cell = ws2.cell(row=row, column=col_idx, value=val)
                cell.border = s["thin_border"]
                cell.alignment = s["left_wrap"] if col_idx == 5 else s["center"]

        for i, w in enumerate([10, 8, 16, 14, 40], 1):
            ws2.column_dimensions[chr(64 + i)].width = w
    else:
        ws2["A3"] = "未發現抄襲群組"

    # Hub students
    if hub_students:
        row = ws2.max_row + 3
        ws2.cell(row=row, column=1, value="疑似源頭學生").font = Font(bold=True, size=12)
        row += 1
        h_headers = ["學生", "關聯人數", "平均相似度(%)"]
        for col, h in enumerate(h_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = s["header_font"]
            cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            cell.alignment = s["center"]
            cell.border = s["thin_border"]

        for hs in hub_students:
            row += 1
            for col_idx, val in enumerate(
                [hs.get("name", ""), hs.get("degree", 0), hs.get("avg_score", 0)], 1
            ):
                cell = ws2.cell(row=row, column=col_idx, value=val)
                cell.border = s["thin_border"]
                cell.fill = s["red_fill"]
                cell.alignment = s["center"]

    filename = f"{assignment_title}_抄袭檢測報告.xlsx"
    return wb, filename
