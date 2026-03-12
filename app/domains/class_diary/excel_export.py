"""
課室日誌 — Excel 匯出模組

負責 XLSX 工作簿構建（單天 / 日期範圍），包含樣式、佈局、數據填充。
"""

import io
import json
import logging
from typing import Any, Callable, Dict, List, Optional

from app.domains.class_diary.constants import PERIOD_LABELS

logger = logging.getLogger(__name__)


# ====================================================================== #
#  格式化工具                                                               #
# ====================================================================== #

def _format_behavior_field(value: str) -> str:
    """Format JSON behavior field for display, with backward compatibility."""
    if not value or not value.strip():
        return ""
    try:
        data = json.loads(value)
        if isinstance(data, list):
            parts = []
            for item in data:
                reason = item.get("reason", "")
                students = item.get("students", [])
                if students:
                    joined = "、".join(students)
                    parts.append(f"{reason}: {joined}" if reason else joined)
            return "；".join(parts)
    except (json.JSONDecodeError, TypeError, AttributeError):
        pass
    return value


def _period_text(ps: int, pe: int) -> str:
    """節數文字"""
    if ps == pe and ps < len(PERIOD_LABELS):
        return PERIOD_LABELS[ps]
    return (
        f"{PERIOD_LABELS[min(ps, len(PERIOD_LABELS) - 1)]}"
        f"-{PERIOD_LABELS[min(pe, len(PERIOD_LABELS) - 1)]}"
    )


# ====================================================================== #
#  共用樣式                                                                 #
# ====================================================================== #

def _diary_excel_styles():
    """課室日誌 XLSX 共用樣式"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="C0C0C0")
    return {
        "title_font": Font(bold=True, size=16, name="Microsoft JhengHei"),
        "subtitle_font": Font(bold=True, size=12, color="4472C4", name="Microsoft JhengHei"),
        "header_font": Font(bold=True, color="FFFFFF", size=10, name="Microsoft JhengHei"),
        "header_fill": PatternFill(start_color="006633", end_color="006633", fill_type="solid"),
        "section_fill": PatternFill(start_color="E8F5EC", end_color="E8F5EC", fill_type="solid"),
        "alt_fill": PatternFill(start_color="F5FAF7", end_color="F5FAF7", fill_type="solid"),
        "stat_fill": PatternFill(start_color="D0E5D8", end_color="D0E5D8", fill_type="solid"),
        "good_fill": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "warn_fill": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "danger_fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center"),
        "wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "body_font": Font(size=10, name="Microsoft JhengHei"),
    }


def _apply_header_row(ws, row, headers, s):
    """在指定行寫入表頭"""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]


def _apply_data_cell(ws, row, col, value, s, center=True, alt=False):
    """寫入普通數據格"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = s["body_font"]
    c.border = s["border"]
    c.alignment = s["center"] if center else s["wrap"]
    if alt:
        c.fill = s["alt_fill"]
    return c


def _rating_fill(value, s):
    """根據評分返回填色"""
    try:
        v = float(value)
        if v <= 2.5:
            return s["danger_fill"]
        if v >= 4.0:
            return s["good_fill"]
    except (ValueError, TypeError):
        pass
    return None


def _set_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ====================================================================== #
#  單天 XLSX                                                               #
# ====================================================================== #

def build_single_day_xlsx(agg: Dict[str, Any], entry_date: str):
    """構建單天 XLSX workbook"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    s = _diary_excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "概覽"

    overview = agg["overview"]
    if not overview:
        ws["A1"] = f"課室日誌 — {entry_date}"
        ws["A1"].font = s["title_font"]
        ws["A3"] = "指定日期沒有課室日誌記錄"
        ws["A3"].font = s["body_font"]
        return wb

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"課室日誌 — {entry_date}"
    ws["A1"].font = s["title_font"]

    # Stats row
    stat_items = [
        ("總記錄", overview["total_entries"]),
        ("班級數", overview["total_classes"]),
        ("平均紀律", overview["avg_discipline"]),
        ("平均整潔", overview["avg_cleanliness"]),
        ("缺席人次", overview["total_absent"]),
        ("遲到人次", overview["total_late"]),
    ]
    for i, (label, val) in enumerate(stat_items):
        lc = ws.cell(row=3, column=i * 2 + 1, value=label)
        lc.font = Font(bold=True, size=10, name="Microsoft JhengHei")
        lc.fill = s["stat_fill"]
        lc.border = s["border"]
        vc = ws.cell(row=3, column=i * 2 + 2, value=val)
        vc.font = s["body_font"]
        vc.alignment = s["center"]
        vc.border = s["border"]

    # Class stats section
    row = 5
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "班級統計"
    ws[f"A{row}"].font = s["subtitle_font"]

    row = 6
    class_headers = ["班級", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次"]
    _apply_header_row(ws, row, class_headers, s)

    for cs in agg["class_stats"]:
        row += 1
        vals = [cs["class_code"], cs["entry_count"], cs["avg_discipline"],
                cs["avg_cleanliness"], cs["absent_count"], cs["late_count"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws, row, col, v, s)
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    # Detail section
    row += 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "詳細記錄"
    ws[f"A{row}"].font = s["subtitle_font"]

    row += 1
    detail_headers = ["班級", "節數", "科目", "紀律", "整潔", "缺席學生",
                      "遲到學生", "嘉許學生", "儀表問題", "課堂違規", "醫務室", "填寫人"]
    _apply_header_row(ws, row, detail_headers, s)
    freeze_row = row + 1

    idx = 0
    for e in agg["records"]:
        row += 1
        alt = idx % 2 == 1
        vals = [
            e.get("class_code", ""),
            _period_text(e.get("period_start", 0), e.get("period_end", 0)),
            e.get("subject", ""),
            e.get("discipline_rating", ""),
            e.get("cleanliness_rating", ""),
            e.get("absent_students", ""),
            e.get("late_students", ""),
            _format_behavior_field(e.get("commended_students", "")),
            _format_behavior_field(e.get("appearance_issues", "")),
            _format_behavior_field(e.get("rule_violations", "")),
            _format_behavior_field(e.get("medical_room_students", "")),
            e.get("submitted_by", ""),
        ]
        for col, v in enumerate(vals, 1):
            center = col <= 5 or col == 12
            c = _apply_data_cell(ws, row, col, v, s, center=center, alt=alt)
            if col in (4, 5):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f
        idx += 1

    _set_col_widths(ws, [10, 12, 12, 6, 6, 20, 20, 25, 25, 25, 20, 10])
    ws.freeze_panes = f"A{freeze_row}"

    return wb


# ====================================================================== #
#  日期範圍 XLSX（10 個工作表）                                              #
# ====================================================================== #

def build_date_range_xlsx(
    agg: Dict[str, Any], start_date: str, end_date: str
):
    """構建日期範圍 XLSX — 10 個工作表"""
    from collections import defaultdict as _dd

    from openpyxl import Workbook
    from openpyxl.styles import Font

    s = _diary_excel_styles()
    wb = Workbook()

    overview = agg["overview"]
    if not overview:
        ws = wb.active
        ws.title = "總覽"
        ws["A1"] = f"課室日誌 — {start_date} 至 {end_date}"
        ws["A1"].font = s["title_font"]
        ws["A3"] = "指定日期範圍內沒有課室日誌記錄"
        ws["A3"].font = s["body_font"]
        return wb

    # ==================== Sheet 1: 總覽 ====================
    ws = wb.active
    ws.title = "總覽"
    ws.merge_cells("A1:K1")
    ws["A1"] = f"課室日誌 — {start_date} 至 {end_date}"
    ws["A1"].font = s["title_font"]

    stat_items = [
        ("總記錄", overview["total_entries"]),
        ("日期數", overview["total_dates"]),
        ("班級數", overview["total_classes"]),
        ("平均紀律", overview["avg_discipline"]),
        ("平均整潔", overview["avg_cleanliness"]),
    ]
    for i, (label, val) in enumerate(stat_items):
        lc = ws.cell(row=3, column=i * 2 + 1, value=label)
        lc.font = Font(bold=True, size=10, name="Microsoft JhengHei")
        lc.fill = s["stat_fill"]
        lc.border = s["border"]
        vc = ws.cell(row=3, column=i * 2 + 2, value=val)
        vc.font = s["body_font"]
        vc.alignment = s["center"]
        vc.border = s["border"]

    # Daily summary table
    row = 5
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "每日匯總"
    ws[f"A{row}"].font = s["subtitle_font"]

    row = 6
    daily_headers = ["日期", "記錄數", "平均紀律", "平均整潔", "缺席", "遲到",
                     "違規", "表揚", "紀律變化", "7日均紀律", "7日均整潔"]
    _apply_header_row(ws, row, daily_headers, s)

    for ds in agg["daily_summary"]:
        row += 1
        vals = [ds["date"], ds["entry_count"], ds["avg_discipline"], ds["avg_cleanliness"],
                ds["absent_count"], ds["late_count"], ds["violation_count"], ds["praise_count"],
                ds["disc_diff"] if ds["disc_diff"] is not None else "",
                ds["ma7_discipline"], ds["ma7_cleanliness"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws, row, col, v, s)
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws, [12, 8, 8, 8, 6, 6, 6, 6, 8, 8, 8])
    ws.freeze_panes = "A7"

    # ==================== Sheet 2: 每日記錄 ====================
    ws2 = wb.create_sheet("每日記錄")
    row = 1
    detail_headers = ["日期", "班級", "節數", "科目", "紀律", "整潔", "缺席學生",
                      "遲到學生", "嘉許學生", "儀表問題", "課堂違規", "醫務室", "填寫人"]
    _apply_header_row(ws2, row, detail_headers, s)

    by_date = _dd(list)
    for e in agg.get("_raw_entries", []):
        by_date[str(e.get("entry_date", ""))].append(e)

    idx = 0
    for d in sorted(by_date):
        for e in by_date[d]:
            row += 1
            alt = idx % 2 == 1
            vals = [
                str(e.get("entry_date", "")),
                e.get("class_code", ""),
                _period_text(e.get("period_start", 0), e.get("period_end", 0)),
                e.get("subject", ""),
                e.get("discipline_rating", ""),
                e.get("cleanliness_rating", ""),
                e.get("absent_students", ""),
                e.get("late_students", ""),
                _format_behavior_field(e.get("commended_students", "")),
                _format_behavior_field(e.get("appearance_issues", "")),
                _format_behavior_field(e.get("rule_violations", "")),
                _format_behavior_field(e.get("medical_room_students", "")),
                e.get("submitted_by", ""),
            ]
            for col, v in enumerate(vals, 1):
                center = col <= 6 or col == 13
                c = _apply_data_cell(ws2, row, col, v, s, center=center, alt=alt)
                if col in (5, 6):
                    f = _rating_fill(v, s)
                    if f:
                        c.fill = f
            idx += 1

    _set_col_widths(ws2, [12, 10, 12, 12, 6, 6, 20, 20, 25, 25, 25, 20, 10])
    ws2.freeze_panes = "A2"

    # ==================== Sheet 3: 學生記錄 ====================
    ws3 = wb.create_sheet("學生記錄")
    stu_headers = ["學生", "班級", "缺席次數", "缺席日期", "遲到次數", "遲到日期",
                   "違規次數", "違規日期", "醫務室次數", "醫務室日期"]
    _apply_header_row(ws3, 1, stu_headers, s)

    row = 1
    for sr in agg["student_records"]:
        row += 1
        alt = (row % 2) == 0
        vals = [
            sr["name"], sr["class_code"],
            sr["absent_count"], ", ".join(sr["absent_dates"]),
            sr["late_count"], ", ".join(sr["late_dates"]),
            sr["violation_count"], ", ".join(sr["violation_dates"]),
            sr["medical_count"], ", ".join(sr["medical_dates"]),
        ]
        for col, v in enumerate(vals, 1):
            center = col in (1, 2, 3, 5, 7, 9)
            _apply_data_cell(ws3, row, col, v, s, center=center, alt=alt)

    _set_col_widths(ws3, [10, 8, 8, 25, 8, 25, 8, 25, 8, 25])
    ws3.freeze_panes = "A2"

    # ==================== Sheet 4: 學生風險名單 ====================
    ws4 = wb.create_sheet("學生風險名單")
    risk_headers = ["學生", "班級", "總記名", "違規", "缺席", "遲到", "醫務室",
                    "首次記錄", "最近記錄", "涉及天數", "風險標記"]
    _apply_header_row(ws4, 1, risk_headers, s)

    row = 1
    for rs in agg["risk_students"]:
        row += 1
        flags_text = "; ".join(
            f.replace("violation>=3", "違規>=3").replace("absent>=3", "缺席>=3").replace("7d>=3", "7天內>=3次")
            for f in rs["risk_flags"]
        )
        vals = [
            rs["name"], rs["class_code"], rs["total_incidents"],
            rs["violation_count"], rs["absent_count"], rs["late_count"], rs["medical_count"],
            rs["first_date"], rs["last_date"], rs["involved_days"], flags_text,
        ]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws4, row, col, v, s)
            c.fill = s["danger_fill"]

    if not agg["risk_students"]:
        ws4.cell(row=2, column=1, value="沒有高風險學生").font = s["body_font"]

    _set_col_widths(ws4, [10, 8, 8, 8, 8, 8, 8, 12, 12, 8, 25])
    ws4.freeze_panes = "A2"

    # ==================== Sheet 5: 班級統計 ====================
    ws5 = wb.create_sheet("班級統計")
    cs_data = agg["class_stats"]
    classes = cs_data.get("classes", [])
    dates = cs_data.get("dates", [])
    pivot = cs_data.get("data", {})

    # Header row 1: merged class names
    ws5.cell(row=1, column=1, value="日期").font = s["header_font"]
    ws5.cell(row=1, column=1).fill = s["header_fill"]
    ws5.cell(row=1, column=1).border = s["border"]
    for ci, cc in enumerate(classes):
        col_start = 2 + ci * 2
        col_end = col_start + 1
        ws5.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        c = ws5.cell(row=1, column=col_start, value=cc)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]
        ws5.cell(row=1, column=col_end).border = s["border"]

    # Header row 2: 紀律/整潔 sub-headers
    ws5.cell(row=2, column=1, value="").border = s["border"]
    for ci in range(len(classes)):
        col_d = 2 + ci * 2
        col_c = col_d + 1
        cd = ws5.cell(row=2, column=col_d, value="紀律")
        cd.font = s["header_font"]
        cd.fill = s["header_fill"]
        cd.alignment = s["center"]
        cd.border = s["border"]
        cc_cell = ws5.cell(row=2, column=col_c, value="整潔")
        cc_cell.font = s["header_font"]
        cc_cell.fill = s["header_fill"]
        cc_cell.alignment = s["center"]
        cc_cell.border = s["border"]

    # Data rows
    for ri, d in enumerate(dates):
        row = 3 + ri
        alt = ri % 2 == 1
        _apply_data_cell(ws5, row, 1, d, s, alt=alt)
        for ci, cc in enumerate(classes):
            col_d = 2 + ci * 2
            col_c = col_d + 1
            vals = pivot.get(d, {}).get(cc)
            if vals:
                c_d = _apply_data_cell(ws5, row, col_d, vals["avg_discipline"], s, alt=alt)
                f = _rating_fill(vals["avg_discipline"], s)
                if f:
                    c_d.fill = f
                c_c = _apply_data_cell(ws5, row, col_c, vals["avg_cleanliness"], s, alt=alt)
                f = _rating_fill(vals["avg_cleanliness"], s)
                if f:
                    c_c.fill = f
            else:
                _apply_data_cell(ws5, row, col_d, "", s, alt=alt)
                _apply_data_cell(ws5, row, col_c, "", s, alt=alt)

    widths = [12] + [8] * (len(classes) * 2)
    _set_col_widths(ws5, widths)
    ws5.freeze_panes = "B3"

    # ==================== Sheet 6: 節數分析 ====================
    ws6 = wb.create_sheet("節數分析")
    pa_headers = ["節數", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次",
                  "違規人次", "表揚人次", "醫務室人次", "違規率", "表揚率"]
    _apply_header_row(ws6, 1, pa_headers, s)

    row = 1
    for pa in agg["period_analysis"]:
        row += 1
        vals = [pa["period_label"], pa["entry_count"], pa["avg_discipline"], pa["avg_cleanliness"],
                pa["absent_count"], pa["late_count"], pa["violation_count"], pa["praise_count"],
                pa["medical_count"], pa["violation_rate"], pa["praise_rate"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws6, row, col, v, s, alt=(row % 2 == 0))
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws6, [10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8])

    # ==================== Sheet 7: 科目分析 ====================
    ws7 = wb.create_sheet("科目分析")
    sa_headers = ["科目", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次",
                  "違規人次", "表揚人次", "醫務室人次", "違規率", "表揚率"]
    _apply_header_row(ws7, 1, sa_headers, s)

    row = 1
    for sa in agg["subject_analysis"]:
        row += 1
        vals = [sa["subject"], sa["entry_count"], sa["avg_discipline"], sa["avg_cleanliness"],
                sa["absent_count"], sa["late_count"], sa["violation_count"], sa["praise_count"],
                sa["medical_count"], sa["violation_rate"], sa["praise_rate"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws7, row, col, v, s, alt=(row % 2 == 0))
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws7, [14, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8])

    # ==================== Sheet 8: 原因分析 ====================
    ws8 = wb.create_sheet("原因分析")
    ra_headers = ["原因", "類別", "總人次", "涉及學生數", "涉及班級數", "涉及日期數", "首次出現", "最近出現"]
    _apply_header_row(ws8, 1, ra_headers, s)

    row = 1
    for ra in agg["reason_analysis"]:
        row += 1
        vals = [ra["reason"], ra["category"], ra["total_count"], ra["student_count"],
                ra["class_count"], ra["date_count"], ra["first_date"], ra["last_date"]]
        for col, v in enumerate(vals, 1):
            _apply_data_cell(ws8, row, col, v, s, alt=(row % 2 == 0))

    _set_col_widths(ws8, [20, 10, 8, 10, 10, 10, 12, 12])

    # ==================== Sheet 9: 星期模式 ====================
    ws9 = wb.create_sheet("星期模式")
    wd_headers = ["星期", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次",
                  "違規人次", "表揚人次", "違規率", "表揚率"]
    _apply_header_row(ws9, 1, wd_headers, s)

    row = 1
    for wa in agg["weekday_analysis"]:
        row += 1
        vals = [wa["weekday_label"], wa["entry_count"], wa["avg_discipline"], wa["avg_cleanliness"],
                wa["absent_count"], wa["late_count"], wa["violation_count"], wa["praise_count"],
                wa["violation_rate"], wa["praise_rate"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws9, row, col, v, s, alt=(row % 2 == 0))
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws9, [10, 8, 8, 8, 8, 8, 8, 8, 8, 8])

    # ==================== Sheet 10: 填寫人分析 ====================
    ws10 = wb.create_sheet("填寫人分析")
    sub_headers = ["填寫人", "記錄數", "涉及班級數", "平均紀律", "平均整潔",
                   "缺席人次", "遲到人次", "違規人次", "表揚人次"]
    _apply_header_row(ws10, 1, sub_headers, s)

    row = 1
    for suba in agg["submitter_analysis"]:
        row += 1
        vals = [suba["submitter"], suba["entry_count"], suba["class_count"],
                suba["avg_discipline"], suba["avg_cleanliness"],
                suba["absent_count"], suba["late_count"],
                suba["violation_count"], suba["praise_count"]]
        for col, v in enumerate(vals, 1):
            _apply_data_cell(ws10, row, col, v, s, alt=(row % 2 == 0))

    _set_col_widths(ws10, [14, 8, 10, 8, 8, 8, 8, 8, 8])

    return wb


def workbook_to_bytes(wb) -> bytes:
    """將 workbook 輸出為 bytes"""
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
