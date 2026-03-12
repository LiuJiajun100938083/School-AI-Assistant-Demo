"""
共享 Excel 導出基礎設施
========================
提供通用的 Excel 樣式預設、行/列寫入、streaming response 輸出。

域內導出器（class_diary/excel_export.py, assignment/excel_export.py 等）
負責業務數據映射，本模塊負責通用能力。
"""

import io
from typing import List, Optional, Tuple
from urllib.parse import quote


def workbook_to_streaming_response(wb, filename: str):
    """
    將 openpyxl Workbook 轉為 FastAPI StreamingResponse。

    Parameters
    ----------
    wb : openpyxl.Workbook
    filename : str  (會自動 URL-encode 處理中文)

    Returns
    -------
    StreamingResponse
    """
    from fastapi.responses import StreamingResponse

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


def workbook_to_bytes(wb) -> bytes:
    """將 Workbook 序列化為 bytes（用於存儲或測試）"""
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def apply_header_row(ws, row: int, headers: List[str], styles: dict) -> None:
    """
    寫入標題行並套用樣式。

    Parameters
    ----------
    ws : Worksheet
    row : int  行號
    headers : list of str
    styles : dict  需包含 header_font, header_fill, center, thin_border
    """
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["thin_border"]


def set_column_widths(ws, widths: List[int], start_col: int = 1) -> None:
    """批量設置列寬"""
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(widths, start_col):
        ws.column_dimensions[get_column_letter(i)].width = w


def base_styles():
    """
    基礎樣式預設（藍色主題）。

    Returns
    -------
    dict  包含常用樣式對象:
        header_font, header_fill, title_font, thin_border,
        center, left_wrap, gray_fill, green_fill, red_fill, yellow_fill
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    return {
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "title_font": Font(bold=True, size=14),
        "thin_border": thin,
        "center": Alignment(horizontal="center", vertical="center"),
        "left_wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "gray_fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "green_fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "red_fill": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "yellow_fill": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
    }
