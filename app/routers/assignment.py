#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
作業管理路由
============
提供作業佈置、學生提交、文件上傳、評分批改的 API 端點。

老師端點 (/api/assignments/teacher):
    - CRUD 操作、發布、關閉、查看提交、批改、AI 批改

學生端點 (/api/assignments):
    - 查看作業、提交文件、查看成績

Swift 運行:
    - POST /api/assignments/run-swift
"""

import asyncio
import io
import logging
import math
import threading
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

from fastapi import APIRouter, Body, Depends, File, Form, Query, Request, UploadFile
from fastapi.responses import StreamingResponse

from app.core.dependencies import get_current_user, require_teacher
from app.core.responses import error_response, paginated_response, success_response
from app.domains.assignment.schemas import (
    CreateAssignmentRequest,
    GradeSubmissionRequest,
    RunSwiftRequest,
    UpdateAssignmentRequest,
)
from app.services.container import get_services

logger = logging.getLogger(__name__)
router = APIRouter(tags=["作業管理"])


# ================================================================
# 老師端點
# ================================================================

@router.post("/api/assignments/teacher")
async def create_assignment(
    request: CreateAssignmentRequest,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """創建作業 (草稿)"""
    username, role = teacher_info
    services = get_services()

    # 獲取教師 user id
    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0
    teacher_name = user.get("display_name", username) if user else username

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.create_assignment(
            teacher_id=teacher_id,
            teacher_name=teacher_name,
            title=request.title,
            description=request.description,
            target_type=request.target_type,
            target_value=request.target_value,
            deadline=request.deadline,
            max_files=request.max_files,
            allow_late=request.allow_late,
            rubric_type=request.rubric_type,
            rubric_config=request.rubric_config,
            rubric_items=[item.dict() for item in request.rubric_items],
        ),
    )
    return success_response(data=result, message="作業創建成功")


@router.get("/api/assignments/teacher")
async def list_teacher_assignments(
    status: str = Query("", description="狀態過濾: draft, published, closed"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """老師的作業列表"""
    username, _ = teacher_info
    services = get_services()

    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.list_teacher_assignments(
            teacher_id=teacher_id,
            status=status,
            page=page,
            page_size=page_size,
        ),
    )
    return paginated_response(
        data=result["items"],
        total=result["total"],
        page=result["page"],
        page_size=result["page_size"],
    )


@router.get("/api/assignments/teacher/targets")
async def get_targets(
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """獲取可選目標 (班級/學生列表)"""
    services = get_services()
    loop = asyncio.get_event_loop()
    targets = await loop.run_in_executor(
        None, services.assignment.get_available_targets
    )
    return success_response(data=targets)


@router.get("/api/assignments/teacher/plagiarism-presets")
async def get_plagiarism_presets(
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """返回檢測策略預設列表（前端用於顯示策略說明）"""
    from app.domains.assignment.plagiarism_service import DETECTION_PRESETS
    presets = {}
    for key, val in DETECTION_PRESETS.items():
        presets[key] = {
            "label": val["label"],
            "description": val["description"],
            "default_threshold": val["default_threshold"],
        }
    return success_response(data=presets)


@router.get("/api/assignments/teacher/{assignment_id}")
async def get_teacher_assignment_detail(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """作業詳情 (老師視角)"""
    services = get_services()
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None, services.assignment.get_assignment_detail, assignment_id,
    )
    return success_response(data=result)


@router.put("/api/assignments/teacher/{assignment_id}")
async def update_assignment(
    assignment_id: int,
    request: UpdateAssignmentRequest,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """更新作業（草稿和已發布均可）"""
    username, _ = teacher_info
    services = get_services()

    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    fields = request.dict(exclude_unset=True)
    if "rubric_items" in fields and fields["rubric_items"] is not None:
        fields["rubric_items"] = [item.dict() for item in request.rubric_items]
    if "rubric_scores" in fields:
        # rubric_scores belongs to grading, not update
        fields.pop("rubric_scores", None)

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.update_assignment(
            assignment_id, teacher_id, **fields,
        ),
    )
    return success_response(data=result, message="作業更新成功")


@router.post("/api/assignments/teacher/{assignment_id}/publish")
async def publish_assignment(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """發布作業"""
    username, _ = teacher_info
    services = get_services()

    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.publish_assignment(assignment_id, teacher_id),
    )
    return success_response(data=result, message="作業已發布")


@router.post("/api/assignments/teacher/{assignment_id}/close")
async def close_assignment(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """關閉作業"""
    username, _ = teacher_info
    services = get_services()

    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.close_assignment(assignment_id, teacher_id),
    )
    return success_response(data=result, message="作業已關閉")


@router.delete("/api/assignments/teacher/{assignment_id}")
async def delete_assignment(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """刪除作業 (軟刪除)"""
    username, _ = teacher_info
    services = get_services()

    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None,
        lambda: services.assignment.delete_assignment(assignment_id, teacher_id),
    )
    return success_response(message="作業已刪除")


@router.get("/api/assignments/teacher/{assignment_id}/submissions")
async def list_submissions(
    assignment_id: int,
    status: str = Query("", description="狀態過濾: submitted, graded"),
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """查看某作業的所有提交"""
    services = get_services()
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.list_submissions(assignment_id, status),
    )
    return success_response(data=result)


@router.get("/api/assignments/teacher/submissions/{submission_id}")
async def get_submission_detail(
    submission_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """查看單個提交詳情"""
    services = get_services()
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.get_submission_detail(submission_id),
    )
    return success_response(data=result)


@router.post("/api/assignments/teacher/submissions/{submission_id}/grade")
async def grade_submission(
    submission_id: int,
    request: GradeSubmissionRequest,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """批改提交"""
    username, _ = teacher_info
    services = get_services()

    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    scores_data = []
    for s in request.rubric_scores:
        sd = {"rubric_item_id": s.rubric_item_id}
        if s.points is not None:
            sd["points"] = s.points
        if s.selected_level is not None:
            sd["selected_level"] = s.selected_level
        scores_data.append(sd)

    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.grade_submission(
            submission_id=submission_id,
            teacher_id=teacher_id,
            rubric_scores=scores_data,
            feedback=request.feedback,
        ),
    )
    return success_response(data=result, message="批改完成")


@router.post("/api/assignments/teacher/submissions/{submission_id}/ai-grade")
async def ai_grade_submission(
    submission_id: int,
    req: Request,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """AI 自動批改（可附帶額外提示）"""
    extra_prompt = ""
    try:
        body = await req.json()
        extra_prompt = (body.get("extra_prompt") or "") if isinstance(body, dict) else ""
    except Exception:
        pass
    services = get_services()
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.ai_grade_submission(submission_id, extra_prompt=extra_prompt),
    )
    return success_response(data=result, message="AI 批改完成")


# ================================================================
# 批量 AI 批改（背景任務）
# ================================================================

_batch_jobs: Dict[int, dict] = {}  # assignment_id → job state


def _batch_grade_worker(assignment_id: int, submission_ids: List[int],
                         teacher_id: int, extra_prompt: str):
    """背景線程：逐份 AI 批改 + 自動保存"""
    job = _batch_jobs.get(assignment_id)
    if not job:
        return
    job["status"] = "running"

    services = get_services()

    for sub_id in submission_ids:
        # 檢查取消
        if job.get("cancelled"):
            job["status"] = "cancelled"
            return

        try:
            # Step 1: AI 批改
            result = services.assignment.ai_grade_submission(sub_id, extra_prompt=extra_prompt)
            if result.get("error"):
                job["fail"] += 1
                job["done"] += 1
                continue

            # Step 2: 構建分數並保存
            scores = []
            if result.get("selected_level") is not None:
                scores.append({
                    "rubric_item_id": 0,
                    "points": result.get("points", 0),
                    "selected_level": result.get("selected_level", ""),
                })
            else:
                for item in result.get("items", []):
                    entry: dict = {"rubric_item_id": item["rubric_item_id"]}
                    if item.get("points") is not None:
                        entry["points"] = item["points"]
                    if item.get("passed") is not None:
                        entry["points"] = 1 if item["passed"] else 0
                    if item.get("selected_level"):
                        entry["selected_level"] = item["selected_level"]
                    scores.append(entry)

            services.assignment.grade_submission(
                submission_id=sub_id,
                teacher_id=teacher_id,
                rubric_scores=scores,
                feedback=result.get("overall_feedback", "AI 自動批改"),
            )
            job["success"] += 1

        except Exception as e:
            logger.error("批量 AI 批改 submission #%d 失敗: %s", sub_id, e)
            job["fail"] += 1

        job["done"] += 1

    job["status"] = "done"


@router.post("/api/assignments/teacher/{assignment_id}/batch-ai-grade")
async def start_batch_ai_grade(
    assignment_id: int,
    req: Request,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """啟動批量 AI 批改（背景執行）"""
    # 防止重複啟動
    existing = _batch_jobs.get(assignment_id)
    if existing and existing.get("status") == "running":
        return success_response(data=existing, message="批改任務進行中")

    extra_prompt = ""
    mode = "remaining"
    try:
        body = await req.json()
        if isinstance(body, dict):
            extra_prompt = body.get("extra_prompt") or ""
            mode = body.get("mode") or "remaining"
    except Exception:
        pass

    services = get_services()
    username, _ = teacher_info
    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    # 根據模式查詢提交
    if mode == "all":
        # 全部重新批改：取所有已提交和已批改的
        all_subs = services.assignment.list_submissions(assignment_id)
        subs = [s for s in all_subs if s.get("status") in ("submitted", "graded")]
    else:
        # 批改剩餘：僅取未批改的
        subs = services.assignment.list_submissions(assignment_id, status="submitted")

    if not subs:
        return error_response("沒有可批改的提交")

    sub_ids = [s["id"] for s in subs]

    job = {
        "status": "running",
        "total": len(sub_ids),
        "done": 0,
        "success": 0,
        "fail": 0,
        "cancelled": False,
        "extra_prompt": extra_prompt,
    }
    _batch_jobs[assignment_id] = job

    # 啟動背景線程
    t = threading.Thread(
        target=_batch_grade_worker,
        args=(assignment_id, sub_ids, teacher_id, extra_prompt),
        daemon=True,
    )
    t.start()

    return success_response(data=job, message="批改任務已啟動")


@router.get("/api/assignments/teacher/{assignment_id}/batch-ai-grade/status")
async def get_batch_ai_status(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """查詢批量 AI 批改進度"""
    job = _batch_jobs.get(assignment_id)
    if not job:
        return success_response(data={"status": "idle"})
    return success_response(data={
        "status": job["status"],
        "total": job["total"],
        "done": job["done"],
        "success": job["success"],
        "fail": job["fail"],
    })


@router.post("/api/assignments/teacher/{assignment_id}/batch-ai-grade/cancel")
async def cancel_batch_ai_grade(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """取消批量 AI 批改"""
    job = _batch_jobs.get(assignment_id)
    if not job or job["status"] != "running":
        return error_response("沒有正在進行的批改任務")
    job["cancelled"] = True
    return success_response(message="已請求取消")


# ================================================================
# 作業附件
# ================================================================

@router.post("/api/assignments/teacher/{assignment_id}/attachments")
async def upload_attachments(
    assignment_id: int,
    files: List[UploadFile] = File(...),
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """上傳作業附件 (教師用，multipart)"""
    username, _ = teacher_info
    services = get_services()
    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    results = []
    for f in files:
        result = await services.assignment.upload_attachment(
            assignment_id=assignment_id,
            teacher_id=teacher_id,
            file=f,
        )
        results.append(result)

    return success_response(data=results, message=f"已上傳 {len(results)} 個附件")


@router.delete("/api/assignments/teacher/{assignment_id}/attachments/{file_id}")
async def delete_attachment(
    assignment_id: int,
    file_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """刪除作業附件"""
    username, _ = teacher_info
    services = get_services()
    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    loop = asyncio.get_event_loop()
    await loop.run_in_executor(
        None,
        lambda: services.assignment.delete_attachment(
            attachment_id=file_id,
            teacher_id=teacher_id,
        ),
    )
    return success_response(message="附件已刪除")


# ================================================================
# 學生端點
# ================================================================

@router.get("/api/assignments")
async def list_student_assignments(
    status: str = Query("", description="狀態過濾: not_submitted, submitted, graded"),
    current_user: dict = Depends(get_current_user),
):
    """學生的作業列表"""
    services = get_services()
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.list_student_assignments(
            username=current_user["username"],
            class_name=current_user.get("class_name", ""),
            status_filter=status,
        ),
    )
    return success_response(data=result)


@router.get("/api/assignments/{assignment_id}")
async def get_student_assignment_detail(
    assignment_id: int,
    current_user: dict = Depends(get_current_user),
):
    """學生的作業詳情 (含我的提交)"""
    services = get_services()
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.get_student_assignment_detail(
            assignment_id=assignment_id,
            username=current_user["username"],
        ),
    )
    return success_response(data=result)


@router.post("/api/assignments/{assignment_id}/submit")
async def submit_assignment(
    assignment_id: int,
    content: str = Form(default=""),
    files: List[UploadFile] = File(default=[]),
    current_user: dict = Depends(get_current_user),
):
    """學生提交作業 (multipart: 文字 + 文件)"""
    services = get_services()
    result = await services.assignment.submit_assignment(
        assignment_id=assignment_id,
        student=current_user,
        content=content,
        files=files if files else None,
    )
    return success_response(data=result, message="提交成功")


@router.post("/api/assignments/teacher/{assignment_id}/submit-for-student")
async def teacher_submit_for_student(
    assignment_id: int,
    student_username: str = Form(...),
    content: str = Form(default=""),
    files: List[UploadFile] = File(default=[]),
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """教師代替學生提交作業"""
    services = get_services()

    # 查找學生信息
    student_user = services.user.get_user(student_username)
    if not student_user:
        return error_response(f"找不到學生: {student_username}", status_code=404)

    student = {
        "id": student_user["id"],
        "username": student_user.get("username", student_username),
        "display_name": student_user.get("display_name") or student_user.get("name", student_username),
        "class_name": student_user.get("class_name", ""),
    }

    result = await services.assignment.submit_assignment(
        assignment_id=assignment_id,
        student=student,
        content=content or f"（由教師 {teacher_info[0]} 代為提交）",
        files=files if files else None,
    )
    return success_response(data=result, message="代提交成功")


# ================================================================
# Swift 運行
# ================================================================
# Excel 匯出
# ================================================================


def _build_grade_export_excel(
    assignment: dict,
    submissions: List[dict],
    rubric_items: List[dict],
):
    """生成成績匯出 Excel (兩個 Sheet: 成績表 + 成績分析)"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # -- 樣式 --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="4472C4")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    light_blue_fill = PatternFill(start_color="DBEEF4", end_color="DBEEF4", fill_type="solid")

    wb = Workbook()

    # ============================
    # Sheet 1: 成績表
    # ============================
    ws1 = wb.active
    ws1.title = "成績表"

    # Title row
    ws1["A1"] = f"{assignment.get('title', '作業')} - 成績表"
    ws1["A1"].font = title_font
    max_score = assignment.get("max_score") or 100
    ws1["A2"] = f"滿分: {max_score}"
    ws1["A2"].font = Font(size=11, color="666666")

    # Build headers
    headers = ["學號", "姓名", "班級", "總分"]
    for ri in rubric_items:
        title = ri.get("title", "")
        mp = ri.get("max_points")
        headers.append(f"{title} ({mp})" if mp else title)
    headers.append("評語")
    headers.append("狀態")

    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Build rubric_item_id → index mapping
    rubric_id_to_idx = {}
    for idx, ri in enumerate(rubric_items):
        rubric_id_to_idx[ri["id"]] = idx

    # Sort submissions by username (學號)
    sorted_subs = sorted(submissions, key=lambda s: s.get("username") or "")

    # Data rows
    graded_scores = []
    for row_idx, sub in enumerate(sorted_subs, header_row + 1):
        is_graded = sub.get("status") == "graded"
        score = sub.get("score")

        row_data = [
            sub.get("username") or "",
            sub.get("student_name") or "",
            sub.get("class_name") or "",
            float(score) if score is not None else "",
        ]

        # Rubric item scores
        score_map = {}
        for rs in sub.get("rubric_scores") or []:
            rid = rs.get("rubric_item_id")
            if rid in rubric_id_to_idx:
                pts = rs.get("points")
                score_map[rubric_id_to_idx[rid]] = float(pts) if pts is not None else ""

        for i in range(len(rubric_items)):
            row_data.append(score_map.get(i, ""))

        row_data.append(sub.get("feedback") or "")
        status_text = {"graded": "已批改", "submitted": "待批改", "returned": "已退回"}.get(
            sub.get("status", ""), sub.get("status", "")
        )
        row_data.append(status_text)

        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col != len(headers) - 1 else left_align

            if not is_graded:
                cell.fill = gray_fill

        if is_graded and score is not None:
            graded_scores.append(float(score))

    # Column widths
    col_widths = [12, 12, 10, 8]
    col_widths += [12] * len(rubric_items)
    col_widths += [30, 8]
    for col, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    # Freeze header row
    ws1.freeze_panes = f"A{header_row + 1}"

    # ============================
    # Sheet 2: 成績分析
    # ============================
    ws2 = wb.create_sheet(title="成績分析")

    ws2["A1"] = f"{assignment.get('title', '作業')} - 成績分析"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:D1")

    # -- 基本統計 --
    row = 3
    ws2.cell(row=row, column=1, value="基本統計").font = subtitle_font
    row += 1

    total_submitted = len(submissions)
    graded_count = len(graded_scores)

    if graded_scores:
        avg_score = sum(graded_scores) / len(graded_scores)
        sorted_scores = sorted(graded_scores)
        n = len(sorted_scores)
        median_score = (sorted_scores[n // 2] + sorted_scores[(n - 1) // 2]) / 2
        max_s = max(graded_scores)
        min_s = min(graded_scores)
        variance = sum((x - avg_score) ** 2 for x in graded_scores) / len(graded_scores)
        std_dev = math.sqrt(variance)
    else:
        avg_score = median_score = max_s = min_s = std_dev = 0

    stats = [
        ("提交人數", total_submitted),
        ("已批改數", graded_count),
        ("平均分", round(avg_score, 1)),
        ("中位數", round(median_score, 1)),
        ("最高分", round(max_s, 1) if graded_scores else "-"),
        ("最低分", round(min_s, 1) if graded_scores else "-"),
        ("標準差", round(std_dev, 1)),
        ("滿分", max_score),
    ]

    for label, val in stats:
        cell_label = ws2.cell(row=row, column=1, value=label)
        cell_label.font = Font(bold=True)
        cell_label.border = thin_border
        cell_label.fill = light_blue_fill
        cell_val = ws2.cell(row=row, column=2, value=val)
        cell_val.border = thin_border
        cell_val.alignment = center_align
        row += 1

    # -- 分數段分佈 --
    row += 1
    ws2.cell(row=row, column=1, value="分數段分佈").font = subtitle_font
    row += 1

    # Headers
    for col, h in enumerate(["分數段", "人數", "百分比"], 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    # Calculate distribution based on max_score percentage
    max_score_f = float(max_score) if max_score else 100
    ranges = [
        ("90-100%", 0.9 * max_score_f, max_score_f),
        ("80-89%", 0.8 * max_score_f, 0.9 * max_score_f),
        ("70-79%", 0.7 * max_score_f, 0.8 * max_score_f),
        ("60-69%", 0.6 * max_score_f, 0.7 * max_score_f),
        ("0-59%", 0, 0.6 * max_score_f),
    ]
    range_fills = [green_fill, green_fill, yellow_fill, yellow_fill, red_fill]

    for (label, low, high), fill in zip(ranges, range_fills):
        if label == "90-100%":
            count = sum(1 for s in graded_scores if low <= s <= high)
        else:
            count = sum(1 for s in graded_scores if low <= s < high)
        pct = f"{count / graded_count * 100:.1f}%" if graded_count else "0%"

        ws2.cell(row=row, column=1, value=label).border = thin_border
        ws2.cell(row=row, column=1).fill = fill
        ws2.cell(row=row, column=2, value=count).border = thin_border
        ws2.cell(row=row, column=2).alignment = center_align
        ws2.cell(row=row, column=3, value=pct).border = thin_border
        ws2.cell(row=row, column=3).alignment = center_align
        row += 1

    # -- 各評分項平均 --
    if rubric_items:
        row += 1
        ws2.cell(row=row, column=1, value="各評分項統計").font = subtitle_font
        row += 1

        for col, h in enumerate(["評分項", "滿分", "平均分", "得分率"], 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        for ri_idx, ri in enumerate(rubric_items):
            ri_max = float(ri.get("max_points") or 0)
            # Collect all scores for this rubric item
            all_pts = []
            for sub in submissions:
                if sub.get("status") != "graded":
                    continue
                for rs in sub.get("rubric_scores") or []:
                    if rs.get("rubric_item_id") == ri["id"] and rs.get("points") is not None:
                        all_pts.append(float(rs["points"]))

            ri_avg = sum(all_pts) / len(all_pts) if all_pts else 0
            ri_rate = f"{ri_avg / ri_max * 100:.1f}%" if ri_max > 0 else "-"

            ws2.cell(row=row, column=1, value=ri.get("title", "")).border = thin_border
            ws2.cell(row=row, column=2, value=ri_max).border = thin_border
            ws2.cell(row=row, column=2).alignment = center_align
            ws2.cell(row=row, column=3, value=round(ri_avg, 1)).border = thin_border
            ws2.cell(row=row, column=3).alignment = center_align
            cell_rate = ws2.cell(row=row, column=4, value=ri_rate)
            cell_rate.border = thin_border
            cell_rate.alignment = center_align
            row += 1

    # Column widths for Sheet 2
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12

    # -- Build filename --
    title = assignment.get("title", "作業") or "作業"
    filename = f"{title}_成績.xlsx"

    return wb, filename


@router.get("/api/assignments/teacher/{assignment_id}/export-excel")
async def export_grade_excel(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """匯出作業成績到 Excel"""
    services = get_services()

    loop = asyncio.get_event_loop()

    # 取得作業資訊（含 rubric_items）
    try:
        assignment = await loop.run_in_executor(
            None,
            lambda: services.assignment.get_assignment_detail(assignment_id),
        )
    except Exception:
        return error_response("作業不存在", status_code=404)

    rubric_items = assignment.get("rubric_items") or []

    # 取得所有提交（不篩選狀態）
    submissions = await loop.run_in_executor(
        None,
        lambda: services.assignment.list_submissions(assignment_id),
    )

    wb, filename = _build_grade_export_excel(assignment, submissions, rubric_items)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ================================================================

@router.post("/api/assignments/run-swift")
async def run_swift_code(
    request: RunSwiftRequest,
    current_user: dict = Depends(get_current_user),
):
    """運行 Swift 代碼"""
    services = get_services()
    loop = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None,
        lambda: services.assignment.run_swift_code(request.code),
    )
    return success_response(data=result)


# ================================================================
# 抄襲檢測
# ================================================================

_plagiarism_jobs: Dict[int, dict] = {}  # assignment_id → job state


def _plagiarism_worker(assignment_id: int, report_id: int) -> None:
    """背景線程：執行抄袭檢測，帶即時進度更新"""
    job = _plagiarism_jobs.get(assignment_id)
    if not job:
        return
    job["status"] = "running"

    def _on_progress(phase: str, done: int, total: int, detail: str = ""):
        """進度回調: 將進度寫入記憶體中的 job 字典，供 status API 讀取"""
        job["phase"] = phase
        job["phase_done"] = done
        job["phase_total"] = total
        job["detail"] = detail
        # 計算總體百分比（4 個階段各佔一定比例）
        phase_weights = {"extract": 5, "compare": 60, "ai": 30, "save": 5}
        weight = phase_weights.get(phase, 0)
        phase_pct = (done / max(total, 1)) * weight
        # 累計前面已完成階段的百分比
        phase_order = ["extract", "compare", "ai", "save"]
        completed_pct = sum(
            phase_weights[p] for p in phase_order
            if phase_order.index(p) < phase_order.index(phase)
        ) if phase in phase_order else 0
        job["progress"] = min(round(completed_pct + phase_pct), 100)

    try:
        services = get_services()
        services.plagiarism.run_check(report_id, progress_callback=_on_progress)
        # 完成後更新 job 狀態
        report = services.plagiarism.get_report_by_id(report_id)
        if report:
            job["status"] = report.get("status", "completed")
            job["total_pairs"] = report.get("total_pairs", 0)
            job["flagged_pairs"] = report.get("flagged_pairs", 0)
        else:
            job["status"] = "completed"
        job["progress"] = 100
    except Exception as e:
        logger.error("抄袭檢測失敗 (assignment #%d): %s", assignment_id, e)
        job["status"] = "failed"
        job["error"] = str(e)


@router.post("/api/assignments/teacher/{assignment_id}/plagiarism-check")
async def start_plagiarism_check(
    assignment_id: int,
    req: Request,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """啟動抄袭檢測（背景執行）"""
    # 防止重複啟動
    existing = _plagiarism_jobs.get(assignment_id)
    if existing and existing.get("status") == "running":
        return success_response(data=existing, message="檢測任務進行中")

    # 解析參數
    threshold = 60.0
    subject = ""
    detect_mode = "mixed"
    try:
        body = await req.json()
        if isinstance(body, dict):
            threshold = float(body.get("threshold", 60.0))
            subject = str(body.get("subject", ""))
            detect_mode = str(body.get("detect_mode", "mixed"))
    except Exception:
        pass

    services = get_services()
    username, _ = teacher_info
    user = services.user.get_user(username)
    teacher_id = user["id"] if user else 0

    loop = asyncio.get_event_loop()
    report = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.start_check(
            assignment_id=assignment_id,
            teacher_id=teacher_id,
            threshold=threshold,
            subject=subject,
            detect_mode=detect_mode,
        ),
    )

    report_id = report["report_id"]
    job: Dict[str, Any] = {
        "status": "running",
        "report_id": report_id,
        "total_pairs": 0,
        "flagged_pairs": 0,
    }
    _plagiarism_jobs[assignment_id] = job

    # 啟動背景線程
    t = threading.Thread(
        target=_plagiarism_worker,
        args=(assignment_id, report_id),
        daemon=True,
    )
    t.start()

    return success_response(data=job, message="抄袭檢測已啟動")


@router.get("/api/assignments/teacher/{assignment_id}/plagiarism-check/status")
async def get_plagiarism_status(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """查詢抄袭檢測進度"""
    # 先查記憶體中的任務狀態（包含即時進度）
    job = _plagiarism_jobs.get(assignment_id)
    if job and job.get("status") == "running":
        return success_response(data={
            "status": "running",
            "report_id": job.get("report_id"),
            "progress": job.get("progress", 0),
            "phase": job.get("phase", "extract"),
            "phase_done": job.get("phase_done", 0),
            "phase_total": job.get("phase_total", 0),
            "detail": job.get("detail", "啟動中..."),
            "total_pairs": job.get("total_pairs", 0),
            "flagged_pairs": job.get("flagged_pairs", 0),
        })

    # 查資料庫中的最新報告
    services = get_services()
    loop = asyncio.get_event_loop()
    report = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_report(assignment_id),
    )

    if not report:
        return success_response(data={"status": "idle"})

    return success_response(data={
        "status": report.get("status", "idle"),
        "report_id": report.get("id"),
        "total_pairs": report.get("total_pairs", 0),
        "flagged_pairs": report.get("flagged_pairs", 0),
        "created_at": str(report.get("created_at", "")),
        "completed_at": str(report.get("completed_at", "")) if report.get("completed_at") else None,
    })


@router.get("/api/assignments/teacher/{assignment_id}/plagiarism-report")
async def get_plagiarism_report(
    assignment_id: int,
    flagged_only: bool = Query(False, description="僅顯示可疑配對"),
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """取得抄袭檢測報告（含配對列表）"""
    services = get_services()
    loop = asyncio.get_event_loop()

    report = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_report(assignment_id),
    )
    if not report:
        return error_response("尚未執行過抄袭檢測", status_code=404)

    report_id = report["id"]
    pairs = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_pairs(report_id, flagged_only=flagged_only),
    )

    # 聚類分析
    clusters_data = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_clusters(report_id),
    )

    return success_response(data={
        "report": {
            "id": report["id"],
            "assignment_id": report["assignment_id"],
            "status": report["status"],
            "threshold": float(report.get("threshold", 60)),
            "total_pairs": report.get("total_pairs", 0),
            "flagged_pairs": report.get("flagged_pairs", 0),
            "created_at": str(report.get("created_at", "")),
            "completed_at": str(report.get("completed_at", "")) if report.get("completed_at") else None,
        },
        "pairs": pairs,
        "clusters": clusters_data.get("clusters", []),
        "hub_students": clusters_data.get("hub_students", []),
    })


@router.get("/api/assignments/teacher/plagiarism-pairs/{pair_id}")
async def get_plagiarism_pair_detail(
    pair_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """取得單個配對的詳細內容（含並排文本）"""
    services = get_services()
    loop = asyncio.get_event_loop()

    pair = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_pair_detail(pair_id),
    )
    if not pair:
        return error_response("配對不存在", status_code=404)

    return success_response(data=pair)


@router.post("/api/assignments/teacher/plagiarism-pairs/{pair_id}/ai-analyze")
async def ai_analyze_pair(
    pair_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """按需對單個配對進行 AI 深度分析（教師手動觸發）"""
    services = get_services()
    loop = asyncio.get_event_loop()

    result = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.ai_analyze_single_pair(pair_id),
    )

    return success_response(data={"ai_analysis": result})


@router.get("/api/assignments/teacher/{assignment_id}/plagiarism-report/export-excel")
async def export_plagiarism_excel(
    assignment_id: int,
    teacher_info: Tuple[str, str] = Depends(require_teacher),
):
    """匯出抄袭檢測報告到 Excel"""
    services = get_services()
    loop = asyncio.get_event_loop()

    report = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_report(assignment_id),
    )
    if not report:
        return error_response("尚未執行過抄袭檢測", status_code=404)

    report_id = report["id"]
    pairs = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_pairs(report_id),
    )
    clusters_data = await loop.run_in_executor(
        None,
        lambda: services.plagiarism.get_clusters(report_id),
    )

    assignment = await loop.run_in_executor(
        None,
        lambda: services.assignment.get_assignment_detail(assignment_id),
    )

    wb, filename = _build_plagiarism_export_excel(
        assignment, report, pairs,
        clusters_data.get("clusters", []),
        clusters_data.get("hub_students", []),
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def _build_plagiarism_export_excel(
    assignment: dict,
    report: dict,
    pairs: List[dict],
    clusters: List[dict],
    hub_students: List[dict],
):
    """生成抄袭檢測報告 Excel (三個 Sheet: 配對明細 + 群組分析 + 總覽)"""
    import json

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # -- 樣式 --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    wb = Workbook()

    # ============================
    # Sheet 1: 配對明細
    # ============================
    ws1 = wb.active
    ws1.title = "配對明細"

    assignment_title = assignment.get("title", "作業") if assignment else "作業"
    ws1["A1"] = f"{assignment_title} - 抄袭檢測報告"
    ws1["A1"].font = title_font
    ws1["A2"] = (
        f"閾值: {report.get('threshold', 60)}% · "
        f"總配對: {report.get('total_pairs', 0)} · "
        f"可疑: {report.get('flagged_pairs', 0)} · "
        f"檢測時間: {report.get('created_at', '-')}"
    )
    ws1["A2"].font = Font(size=11, color="666666")

    headers = ["學生A", "學生B", "綜合分數(%)", "結構分", "標識符分", "逐字分", "注釋分", "是否可疑", "信號", "AI 分析"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 按相似度倒序
    sorted_pairs = sorted(pairs, key=lambda p: float(p.get("similarity_score", 0)), reverse=True)

    for row_idx, p in enumerate(sorted_pairs, header_row + 1):
        is_flagged = p.get("is_flagged", False)

        # 解析維度分數
        frags = p.get("matched_fragments") or []
        if isinstance(frags, str):
            try:
                frags = json.loads(frags)
            except (json.JSONDecodeError, TypeError):
                frags = []
        dim = None
        signals_text = ""
        for f in frags:
            if isinstance(f, dict) and f.get("type") == "dimension_breakdown":
                dim = f
                signals_text = ", ".join(f.get("signals", []))
                break

        row_data = [
            p.get("student_a_name", ""),
            p.get("student_b_name", ""),
            float(p.get("similarity_score", 0)),
            dim.get("structure_score", "") if dim else "",
            dim.get("identifier_score", "") if dim else "",
            dim.get("verbatim_score", "") if dim else "",
            dim.get("comment_score", "") if dim else "",
            "是" if is_flagged else "否",
            signals_text,
            (p.get("ai_analysis") or "")[:500],
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = left_align if col_idx >= 9 else center_align

        # 可疑行標紅
        if is_flagged:
            for col_idx in range(1, len(row_data) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = red_fill

    # 列寬
    col_widths = [14, 14, 12, 10, 10, 10, 10, 10, 24, 40]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = w

    # ============================
    # Sheet 2: 群組分析
    # ============================
    ws2 = wb.create_sheet("群組分析")

    ws2["A1"] = "抄襲群組分析"
    ws2["A1"].font = title_font

    if clusters:
        row = 3
        g_headers = ["群組", "人數", "最高相似度(%)", "疑似源頭", "成員"]
        for col, h in enumerate(g_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for c in clusters:
            row += 1
            source = c.get("source_student") or "-"
            members = ", ".join(m.get("name", "") for m in c.get("members", []))
            for col_idx, val in enumerate(
                [f"群組 {c.get('id', '')}", c.get("size", 0), c.get("max_score", 0), source, members], 1
            ):
                cell = ws2.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = left_align if col_idx == 5 else center_align

        for i, w in enumerate([10, 8, 16, 14, 40], 1):
            ws2.column_dimensions[chr(64 + i)].width = w
    else:
        ws2["A3"] = "未發現抄襲群組"

    # ---- Hub 學生 ----
    if hub_students:
        row = ws2.max_row + 3
        ws2.cell(row=row, column=1, value="疑似源頭學生").font = Font(bold=True, size=12)
        row += 1
        h_headers = ["學生", "關聯人數", "平均相似度(%)"]
        for col, h in enumerate(h_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
            cell.alignment = center_align
            cell.border = thin_border

        for hs in hub_students:
            row += 1
            for col_idx, val in enumerate(
                [hs.get("name", ""), hs.get("degree", 0), hs.get("avg_score", 0)], 1
            ):
                cell = ws2.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.fill = red_fill
                cell.alignment = center_align

    filename = f"{assignment_title}_抄袭檢測報告.xlsx"
    return wb, filename
