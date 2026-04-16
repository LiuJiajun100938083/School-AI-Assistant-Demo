"""
用户管理路由 - UserRouter
==========================
处理所有用户管理相关的 HTTP 端点：
- GET  /api/profile                        - 当前用户资料
- GET  /api/user/info                      - 当前用户信息
- GET  /api/admin/users                    - 用户列表（管理员）
- POST /api/admin/users                    - 创建用户
- PUT  /api/admin/users/{username}         - 更新用户
- DELETE /api/admin/users/{username}       - 删除用户
- POST /api/admin/users/{username}/reset-password  - 重置密码
- POST /api/admin/users/{username}/status  - 切换状态
- POST /api/admin/users/batch             - 批量创建
- POST /api/admin/users/upload-excel      - Excel 导入
- GET  /api/admin/users/template          - 下载导入模板
"""

import io
import logging
from typing import Optional

from fastapi import APIRouter, Depends, Request, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse

from app.core.exceptions import AppException
from app.core.responses import error_response, success_response
from app.services import get_services

logger = logging.getLogger(__name__)

router = APIRouter(tags=["用户管理"])


# ====================================================================== #
#  当前用户端点                                                            #
# ====================================================================== #

@router.get("/api/profile")
async def get_profile(request: Request):
    """获取当前用户资料"""
    try:
        username, role = _verify_request(request)
        profile = get_services().user.get_user_profile(username)
        # 前端直接读取 userProfile.username / .display_name，返回扁平格式
        return profile
    except Exception as e:
        logger.error("获取资料失败: %s", e)
        from fastapi import HTTPException
        raise HTTPException(status_code=401, detail=str(e))


@router.get("/api/user/info")
async def get_user_info(request: Request):
    """获取当前用户详细信息"""
    try:
        username, role = _verify_request(request)
        user = get_services().user.get_user(username)
        user["role"] = role
        # 前端 auth.js getUserInfo() 期望 {success, data} 格式
        return {"success": True, "data": user}

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  管理员 - 用户查询                                                       #
# ====================================================================== #

@router.get("/api/admin/users")
async def get_users(request: Request):
    """获取用户列表（管理员）"""
    try:
        _verify_admin(request)
        role_filter = request.query_params.get("role")
        users = get_services().user.list_users(role=role_filter)
        # 前端期望 {users: [...]} 格式
        return {"users": users}

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("获取用户列表失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/admin/stats")
async def get_admin_stats(request: Request):
    """获取用户统计数据"""
    try:
        _verify_request(request)
        stats = get_services().user.get_user_stats()
        return success_response(stats)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  管理员 - 用户创建                                                       #
# ====================================================================== #

@router.post("/api/admin/users")
async def create_user(request: Request):
    """创建单个用户"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()

        class_number_raw = body.get("class_number")
        class_number = int(class_number_raw) if class_number_raw is not None else None

        user = get_services().user.create_user(
            username=body.get("username", "").strip(),
            password=body.get("password", ""),
            role=body.get("role", "student"),
            display_name=body.get("display_name", ""),
            class_name=body.get("class_name", ""),
            class_number=class_number,
            email=body.get("email", ""),
            english_name=body.get("english_name", ""),
            card_id=body.get("card_id") or None,
            created_by=admin_user,
        )
        return success_response(user, "用户创建成功")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("创建用户失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/admin/users/batch")
async def batch_create_users(request: Request):
    """批量创建用户"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        users_data = body.get("users", [])

        if not users_data:
            return error_response("VALIDATION_ERROR", "用户数据不能为空", status_code=400)

        result = get_services().user.batch_create_users(
            users_data=users_data,
            created_by=admin_user,
        )
        return success_response(result, "批量创建完成")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("批量创建失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/admin/users/upload-excel")
async def upload_excel_users(request: Request, file: UploadFile = File(...)):
    """Excel 批量导入用户"""
    try:
        admin_user, _ = _verify_admin(request)

        # 验证文件类型
        if not file.filename.endswith((".xlsx", ".xls")):
            return error_response("VALIDATION_ERROR", "请上传 Excel 文件 (.xlsx)", status_code=400)

        # 解析 Excel
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        import re as _re

        def _clean_header(h):
            """清理表头：去除 *、括号说明、首尾空格，转小写"""
            if not h:
                return ""
            s = str(h).strip().lower()
            s = s.replace("*", "")           # 去掉必填标记 *
            s = _re.sub(r'\(.*?\)', '', s)   # 去掉括号说明 (student/teacher)
            s = _re.sub(r'（.*?）', '', s)   # 去掉中文括号说明
            return s.strip()

        rows = []
        headers = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if idx == 0:
                headers = [_clean_header(h) for h in row]
                continue
            if not any(row):
                continue

            row_dict = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_dict[headers[col_idx]] = str(value).strip() if value else ""

            # 兼容不同表头名称（中英文）
            if "username" not in row_dict and "用户名" in row_dict:
                row_dict["username"] = row_dict["用户名"]
            if "username" not in row_dict and "用戶名" in row_dict:
                row_dict["username"] = row_dict["用戶名"]
            if "password" not in row_dict and "密码" in row_dict:
                row_dict["password"] = row_dict["密码"]
            if "password" not in row_dict and "密碼" in row_dict:
                row_dict["password"] = row_dict["密碼"]
            if "display_name" not in row_dict and "姓名" in row_dict:
                row_dict["display_name"] = row_dict["姓名"]
            if "display_name" not in row_dict and "顯示名稱" in row_dict:
                row_dict["display_name"] = row_dict["顯示名稱"]
            if "display_name" not in row_dict and "显示名称" in row_dict:
                row_dict["display_name"] = row_dict["显示名称"]
            if "class_name" not in row_dict and "班级" in row_dict:
                row_dict["class_name"] = row_dict["班级"]
            if "class_name" not in row_dict and "班級" in row_dict:
                row_dict["class_name"] = row_dict["班級"]
            if "role" not in row_dict and "角色" in row_dict:
                row_dict["role"] = row_dict["角色"]
            if "role" not in row_dict and "角色(student/teacher)" in row_dict:
                row_dict["role"] = row_dict["角色(student/teacher)"]
            # 英文名 / 卡號
            if "english_name" not in row_dict and "英文名" in row_dict:
                row_dict["english_name"] = row_dict["英文名"]
            if "english_name" not in row_dict and "英文名稱" in row_dict:
                row_dict["english_name"] = row_dict["英文名稱"]
            if "card_id" not in row_dict and "卡號" in row_dict:
                row_dict["card_id"] = row_dict["卡號"]
            if "card_id" not in row_dict and "卡号" in row_dict:
                row_dict["card_id"] = row_dict["卡号"]

            rows.append(row_dict)

        wb.close()

        if not rows:
            return error_response("VALIDATION_ERROR", "Excel 文件无有效数据", status_code=400)

        result = get_services().user.batch_import_from_excel(
            rows=rows, created_by=admin_user,
        )
        return success_response(result, f"导入完成：成功 {result['success_count']} 个")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("Excel 导入失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/admin/users/template")
async def download_user_template(request: Request):
    """下载用户导入 Excel 模板"""
    try:
        _verify_admin(request)

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "用户导入模板"

        # 表头
        headers = [
            "username", "password", "display_name", "english_name",
            "card_id", "class_name", "class_number", "role",
        ]
        chinese_headers = [
            "用户名*", "密码*", "显示名称*", "英文名",
            "卡號", "班级", "班號", "角色(student/teacher)",
        ]
        for col, (en, cn) in enumerate(zip(headers, chinese_headers), 1):
            ws.cell(row=1, column=col, value=cn)

        # 示例数据（含学生和教师示例）
        examples = [
            ["student001", "Pass123!", "张三", "ZHANG San", "CARD001", "1A", 1, "student"],
            ["student002", "Pass456!", "李四", "LI Si", "CARD002", "1B", 2, "student"],
            ["teacher001", "Pass789!", "王老师", "", "", "", "", "teacher"],
            ["teacher002", "Pass012!", "陈老师", "", "", "", "", "teacher"],
        ]
        for row_idx, example in enumerate(examples, 2):
            for col_idx, value in enumerate(example, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 调整列宽
        col_widths = [15, 12, 15, 18, 12, 10, 8, 20]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"},
        )

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  管理员 - 用户更新                                                       #
# ====================================================================== #

@router.put("/api/admin/users/{username}")
async def update_user(username: str, request: Request):
    """更新用户信息"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()

        # 如果提供了密码，单独处理密码更新
        new_password = body.pop("password", None)
        if new_password:
            client_ip = request.headers.get("x-real-ip", request.client.host if request.client else "")
            get_services().auth.reset_password(
                admin_username=admin_user,
                target_username=username,
                new_password=new_password,
                client_ip=client_ip,
            )

        # 更新其余用户信息（过滤掉 password 后的字段）
        remaining = {k: v for k, v in body.items() if k != "password"}
        if remaining:
            user = get_services().user.update_user(
                username=username,
                data=remaining,
                updated_by=admin_user,
            )
        else:
            user = get_services().user.get_user(username)

        msg = "更新成功（含密码）" if new_password else "更新成功"
        return success_response(user, msg)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("更新用户失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/admin/users/{username}/status")
async def toggle_user_status(username: str, request: Request):
    """切换用户启用/禁用状态"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        is_active = body.get("is_active", True)

        get_services().user.set_user_status(
            username=username,
            is_active=bool(is_active),
            operator=admin_user,
        )
        status_text = "启用" if is_active else "禁用"
        return success_response(None, f"用户已{status_text}")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/admin/users/{username}/reset-password")
async def reset_password(username: str, request: Request):
    """管理员重置用户密码"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        new_password = body.get("new_password", "")

        if not new_password:
            return error_response("VALIDATION_ERROR", "新密码不能为空", status_code=400)

        client_ip = request.headers.get("x-real-ip", request.client.host if request.client else "")
        get_services().auth.reset_password(
            admin_username=admin_user,
            target_username=username,
            new_password=new_password,
            client_ip=client_ip,
        )
        return success_response(None, "密码已重置")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("重置密码失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  管理员 - 用户删除                                                       #
# ====================================================================== #

@router.delete("/api/admin/users/{username}")
async def delete_user(username: str, request: Request):
    """删除用户（级联删除）"""
    try:
        admin_user, _ = _verify_admin(request)

        result = get_services().user.delete_user(
            username=username,
            operator=admin_user,
        )
        return success_response(result, "用户已删除")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("删除用户失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  管理员 - 封禁管理                                                       #
# ====================================================================== #

@router.get("/api/admin/blocked-accounts")
async def get_blocked_accounts(request: Request):
    """获取所有临时封锁的账户/IP（管理员）"""
    try:
        _verify_admin(request)
        blocked = get_services().auth.get_blocked_entries()
        return success_response(blocked)
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("获取封锁列表失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/admin/unblock")
async def unblock_account(request: Request):
    """手动解除临时封锁（管理员）"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        block_type = body.get("block_type", "")
        key = body.get("key", "")

        if block_type not in ("user", "ip", "ip_user") or not key:
            return error_response("VALIDATION_ERROR", "无效的参数", status_code=400)

        removed = get_services().auth.admin_unblock(block_type, key)

        # 获取管理员来源 IP
        client_ip = request.headers.get("x-forwarded-for", "").split(",")[0].strip() \
            or request.client.host if request.client else "unknown"

        if removed:
            logger.info(
                "管理员 %s 解除封锁: type=%s, key=%s, ip=%s",
                admin_user, block_type, key, client_ip,
            )
            return success_response(None, "已解除该条封锁")
        else:
            return error_response("NOT_FOUND", "该封锁已过期或不存在", status_code=404)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("解除封锁失败: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  管理员 - 系統日誌                                                       #
# ====================================================================== #

@router.get("/api/admin/system-logs")
async def get_system_logs(request: Request):
    """查詢系統日誌（管理員）

    Query params:
        log_type: "security_db" (DB 審計) | "app_file" (應用日誌文件)
        limit:    筆數上限 (default 100, max 500)
        search:   搜尋關鍵字 (用戶名 / IP / 事件)
        level:    日誌級別篩選 (INFO / WARNING / ERROR)
    """
    try:
        _verify_admin(request)

        log_type = request.query_params.get("log_type", "security_db")
        limit = min(int(request.query_params.get("limit", "100")), 500)
        search = request.query_params.get("search", "").strip()
        level = request.query_params.get("level", "").upper().strip()

        if log_type == "security_db":
            entries = _query_security_db_logs(limit=limit, search=search)
        else:
            entries = _query_app_file_logs(limit=limit, search=search, level=level)

        return success_response({"entries": entries, "total": len(entries)})

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.error("查詢系統日誌失敗: %s", e)
        return error_response("SERVER_ERROR", str(e), status_code=500)


def _query_security_db_logs(limit: int = 100, search: str = "") -> list:
    """從 security_audit_log 表查詢"""
    try:
        from app.core.audit import SecurityAuditLogger
        rows = SecurityAuditLogger.query(limit=limit)
        if search:
            kw = search.lower()
            rows = [r for r in rows if kw in str(r).lower()]
        # 格式化為前端需要的結構
        entries = []
        for r in rows:
            entries.append({
                "timestamp": str(r.get("created_at", "")),
                "level": "INFO",
                "event": r.get("action", ""),
                "username": r.get("actor", ""),
                "ip": r.get("ip_address", ""),
                "details": r.get("details_json") if isinstance(r.get("details_json"), dict) else {},
            })
        return entries
    except Exception as e:
        logger.warning("查詢 security_audit_log 失敗: %s", e)
        return []


def _query_app_file_logs(limit: int = 100, search: str = "", level: str = "") -> list:
    """讀取 logs/security_audit.log 文件尾部（JSON Lines 格式）"""
    import json as _json
    import os as _os

    log_path = _os.path.join("logs", "security_audit.log")
    if not _os.path.exists(log_path):
        return []

    entries = []
    try:
        # 讀取尾部（效率：只讀最後 200KB）
        file_size = _os.path.getsize(log_path)
        read_size = min(file_size, 200 * 1024)

        with open(log_path, "r", encoding="utf-8") as f:
            if file_size > read_size:
                f.seek(file_size - read_size)
                f.readline()  # 跳過可能截斷的第一行
            lines = f.readlines()

        # 反序（最新在前）
        for line in reversed(lines):
            if len(entries) >= limit:
                break
            line = line.strip()
            if not line:
                continue

            # 提取 JSON 部分（格式: "2025-xx-xx ... - INFO - {json}"）
            json_start = line.find("{")
            if json_start < 0:
                continue

            # 提取日誌級別
            line_level = "INFO"
            prefix = line[:json_start]
            for lv in ("ERROR", "WARNING", "INFO", "DEBUG"):
                if lv in prefix:
                    line_level = lv
                    break

            # 級別篩選
            if level and line_level != level:
                continue

            try:
                data = _json.loads(line[json_start:])
            except _json.JSONDecodeError:
                continue

            entry = {
                "timestamp": data.get("timestamp", ""),
                "level": line_level,
                "event": data.get("event", data.get("event_type", "")),
                "username": data.get("username", data.get("user_id", "")),
                "ip": data.get("ip", data.get("ip_address", "")),
                "details": {k: v for k, v in data.items()
                            if k not in ("timestamp", "event", "event_type", "username", "user_id", "ip", "ip_address")},
            }

            # 關鍵字搜索
            if search:
                kw = search.lower()
                haystack = f"{entry['event']} {entry['username']} {entry['ip']} {str(entry['details'])}".lower()
                if kw not in haystack:
                    continue

            entries.append(entry)

    except Exception as e:
        logger.warning("讀取應用日誌文件失敗: %s", e)

    return entries


# ====================================================================== #
#  辅助函数                                                               #
# ====================================================================== #

def _extract_token(request: Request) -> Optional[str]:
    """从请求头提取 JWT"""
    auth = request.headers.get("authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:]
    return request.query_params.get("token")


def _verify_request(request: Request):
    """验证请求并返回 (username, role)"""
    token = _extract_token(request)
    if not token:
        from app.core.exceptions import AuthenticationError
        raise AuthenticationError("未提供认证令牌")
    payload = get_services().auth.verify_token(token)
    return payload["username"], payload["role"]


def _verify_admin(request: Request):
    """验证管理员权限，返回 (username, role)"""
    username, role = _verify_request(request)
    if role != "admin":
        from app.core.exceptions import AuthorizationError
        raise AuthorizationError("需要管理员权限")
    return username, role
