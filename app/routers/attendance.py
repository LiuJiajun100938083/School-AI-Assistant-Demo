"""
考勤系统 API 路由
==================
提供早读、留堂、课外活动考勤的 RESTful 端点。

所有业务逻辑委托给 AttendanceService (app.domains.attendance.service)，
本模块只负责：
1. 解析 HTTP 请求参数
2. 调用 Service 方法
3. 返回标准化 JSON 响应
4. Excel 导出的文件生成（表现层逻辑）
"""

import csv
import io
import logging
import os
import uuid
from datetime import datetime, time, timedelta
from typing import Dict, List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel

from app.bridge import get_db
from app.core.dependencies import require_teacher_or_admin as verify_admin_or_teacher
from app.domains.attendance.schemas import (
    ActivityCheckoutRequest,
    ActivityGroupRequest,
    ActivityScanRequest,
    ActivitySessionRequest,
    AttendanceSession,
    CardScanRequest,
    DetentionCheckinRequest,
    DetentionManualCheckinRequest,
    FixedListRequest,
    ManualScanRequest,
    ModifyEndTimeRequest,
    ModifyPeriodsRequest,
)
from app.domains.attendance.constants import (
    AttendanceStatus,
    ActivityCheckinStatus,
    ActivityCheckoutStatus,
    DetentionReason,
    ExcelColors,
    SessionType,
)
from app.services.container import get_services

# Excel 导出
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/attendance", tags=["点名系统"])


# ==================================================================================
#                               辅助函数
# ==================================================================================

def _get_service():
    """获取 AttendanceService 实例"""
    return get_services().attendance


async def verify_token_from_query_or_header(
    token: str = Query(None),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer(auto_error=False)),
):
    """支持从 URL 参数或 Header 获取 token（用于 window.open 导出场景）"""
    from app.bridge import decode_jwt_token

    actual_token = token
    if not actual_token and credentials:
        actual_token = credentials.credentials

    if not actual_token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    payload = decode_jwt_token(actual_token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    username = payload.get("username")
    role = payload.get("role")
    if not username:
        raise HTTPException(status_code=401, detail="Invalid token payload")
    if role not in ("admin", "teacher"):
        raise HTTPException(status_code=403, detail="Admin or teacher privileges required")

    return username, role


def _extract_username(user_info) -> tuple:
    """从 user_info 提取 (username, display_name)"""
    if isinstance(user_info, dict):
        username = user_info.get("username") or user_info.get("user_login") or "unknown"
        display_name = user_info.get("display_name") or user_info.get("name") or username
        return username, display_name
    if isinstance(user_info, (list, tuple)) and len(user_info) >= 1:
        username = user_info[0] or "unknown"
        return username, username
    return "unknown", "unknown"


# ==================================================================================
#                               数据库初始化
# ==================================================================================

def init_attendance_tables():
    """初始化点名系统数据库表"""
    with get_db() as db:
        cursor = db.cursor()

        # attendance_students 已合併至 users 表（v3.0.55），不再建立

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_sessions (
                id INT AUTO_INCREMENT PRIMARY KEY,
                session_type ENUM('morning', 'detention') NOT NULL,
                session_date DATE NOT NULL,
                start_time TIME, end_time TIME,
                target_time TIME NOT NULL DEFAULT '07:30:00',
                late_threshold TIME NOT NULL DEFAULT '07:40:00',
                makeup_minutes INT DEFAULT 35,
                status ENUM('active', 'completed', 'cancelled') DEFAULT 'active',
                created_by VARCHAR(50), notes TEXT,
                open_mode BOOLEAN DEFAULT FALSE,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_date (session_date), INDEX idx_type (session_type), INDEX idx_status (status)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_session_students (
                id INT AUTO_INCREMENT PRIMARY KEY,
                session_id INT NOT NULL, user_login VARCHAR(50) NOT NULL,
                added_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (session_id) REFERENCES attendance_sessions(id) ON DELETE CASCADE,
                UNIQUE KEY unique_session_student (session_id, user_login),
                INDEX idx_session (session_id), INDEX idx_student (user_login)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_records (
                id INT AUTO_INCREMENT PRIMARY KEY,
                session_id INT NOT NULL, user_login VARCHAR(50) NOT NULL,
                card_id VARCHAR(50), scan_time DATETIME NOT NULL,
                checkout_time DATETIME,
                status ENUM('present','late','very_late','absent','detention_active','detention_completed') NOT NULL,
                late_minutes INT DEFAULT 0, makeup_minutes INT DEFAULT 0,
                is_registered BOOLEAN DEFAULT TRUE,
                planned_periods INT DEFAULT 0, planned_minutes INT DEFAULT NULL,
                planned_end_time DATETIME, actual_minutes INT DEFAULT 0, actual_periods INT DEFAULT 0,
                detention_reason VARCHAR(50),
                notes TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (session_id) REFERENCES attendance_sessions(id) ON DELETE CASCADE,
                UNIQUE KEY unique_session_record (session_id, user_login),
                INDEX idx_session (session_id), INDEX idx_student (user_login), INDEX idx_scan_time (scan_time)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        # 兼容旧数据库结构升级
        alter_sqls = [
            "ALTER TABLE attendance_sessions ADD COLUMN open_mode BOOLEAN DEFAULT FALSE AFTER notes",
            "ALTER TABLE attendance_records ADD COLUMN checkout_time DATETIME NULL AFTER scan_time",
            "ALTER TABLE attendance_records ADD COLUMN planned_periods INT DEFAULT 0 AFTER is_registered",
            "ALTER TABLE attendance_records ADD COLUMN planned_end_time DATETIME NULL AFTER planned_periods",
            "ALTER TABLE attendance_records ADD COLUMN actual_minutes INT DEFAULT 0 AFTER planned_end_time",
            "ALTER TABLE attendance_records ADD COLUMN actual_periods INT DEFAULT 0 AFTER actual_minutes",
            "ALTER TABLE attendance_records ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP",
            "ALTER TABLE attendance_records MODIFY COLUMN status ENUM('present','late','very_late','absent','detention_active','detention_completed') NOT NULL",
            "ALTER TABLE attendance_records ADD COLUMN detention_reason VARCHAR(50) NULL AFTER actual_periods",
            "ALTER TABLE attendance_records ADD COLUMN planned_minutes INT DEFAULT NULL AFTER planned_periods",
        ]
        for sql in alter_sqls:
            try:
                cursor.execute(sql)
            except Exception:
                pass

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_fixed_lists (
                id INT AUTO_INCREMENT PRIMARY KEY,
                list_name VARCHAR(100) NOT NULL, list_type ENUM('morning','detention') DEFAULT 'morning',
                created_by VARCHAR(50), is_default BOOLEAN DEFAULT FALSE,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY unique_list_name (list_name, list_type)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_fixed_list_students (
                id INT AUTO_INCREMENT PRIMARY KEY,
                list_id INT NOT NULL, user_login VARCHAR(50) NOT NULL,
                added_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (list_id) REFERENCES attendance_fixed_lists(id) ON DELETE CASCADE,
                UNIQUE KEY unique_list_student (list_id, user_login), INDEX idx_list (list_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS detention_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_login VARCHAR(50) NOT NULL, session_id INT,
                detention_date DATE NOT NULL, reason TEXT, duration_minutes INT DEFAULT 35,
                completed BOOLEAN DEFAULT FALSE, completed_at DATETIME,
                created_by VARCHAR(50), notes TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_user (user_login), INDEX idx_date (detention_date), INDEX idx_completed (completed)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance_exports (
                id INT AUTO_INCREMENT PRIMARY KEY,
                session_id INT NOT NULL, session_type ENUM('morning','detention') NOT NULL,
                session_date DATE NOT NULL,
                created_by VARCHAR(50) NOT NULL, created_by_name VARCHAR(100),
                file_name VARCHAR(255) NOT NULL, file_path VARCHAR(500) NOT NULL,
                file_size BIGINT DEFAULT 0,
                student_count INT DEFAULT 0, present_count INT DEFAULT 0,
                late_count INT DEFAULT 0, absent_count INT DEFAULT 0,
                notes TEXT, is_deleted BOOLEAN DEFAULT FALSE, deleted_at DATETIME NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_created_by (created_by), INDEX idx_session (session_id),
                INDEX idx_date (session_date), INDEX idx_type (session_type), INDEX idx_deleted (is_deleted)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS activity_groups (
                id INT AUTO_INCREMENT PRIMARY KEY,
                name VARCHAR(100) NOT NULL, created_by VARCHAR(50),
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY unique_group_name (name)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS activity_group_students (
                id INT AUTO_INCREMENT PRIMARY KEY,
                group_id INT NOT NULL, user_login VARCHAR(50) NOT NULL,
                added_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (group_id) REFERENCES activity_groups(id) ON DELETE CASCADE,
                UNIQUE KEY unique_group_student (group_id, user_login), INDEX idx_group (group_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS activity_sessions (
                id INT AUTO_INCREMENT PRIMARY KEY,
                session_date DATE NOT NULL, activity_name VARCHAR(200) NOT NULL,
                start_time TIME NOT NULL, end_time TIME NOT NULL,
                late_threshold INT DEFAULT 10, early_threshold INT DEFAULT 10,
                status ENUM('active','completed','cancelled') DEFAULT 'active',
                created_by VARCHAR(50), notes TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_date (session_date), INDEX idx_status (status)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS activity_session_students (
                id INT AUTO_INCREMENT PRIMARY KEY,
                session_id INT NOT NULL, user_login VARCHAR(50) NOT NULL,
                added_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (session_id) REFERENCES activity_sessions(id) ON DELETE CASCADE,
                UNIQUE KEY unique_session_student (session_id, user_login), INDEX idx_session (session_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS activity_records (
                id INT AUTO_INCREMENT PRIMARY KEY,
                session_id INT NOT NULL, user_login VARCHAR(50) NOT NULL,
                card_id VARCHAR(50),
                check_in_time DATETIME, check_in_status ENUM('on_time','late','not_arrived') DEFAULT 'not_arrived',
                check_out_time DATETIME, check_out_status ENUM('normal','early','not_arrived','still_here') DEFAULT 'not_arrived',
                late_minutes INT DEFAULT 0, early_minutes INT DEFAULT 0, notes TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (session_id) REFERENCES activity_sessions(id) ON DELETE CASCADE,
                UNIQUE KEY unique_session_record (session_id, user_login),
                INDEX idx_session (session_id), INDEX idx_student (user_login)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)

        db.commit()

        # --- 合併 attendance_students → users：為 users 表添加簽到欄位 ---
        alter_sqls = [
            "ALTER TABLE users ADD COLUMN english_name VARCHAR(100) DEFAULT '' COMMENT '英文名' AFTER display_name",
            "ALTER TABLE users ADD COLUMN card_id VARCHAR(50) DEFAULT NULL COMMENT '學生證 CardID' AFTER english_name",
            "ALTER TABLE users ADD INDEX idx_card_id (card_id)",
        ]
        for sql in alter_sqls:
            try:
                cursor.execute(sql)
                db.commit()
            except Exception:
                db.rollback()  # 欄位/索引已存在，忽略

        logger.info("点名系统数据库表初始化完成")


# ==================================================================================
#                               启动事件
# ==================================================================================

@router.on_event("startup")
async def startup_init_tables():
    """启动时初始化表"""
    try:
        init_attendance_tables()
    except Exception as e:
        logger.error(f"初始化点名表失败: {e}")


# ==================================================================================
#                               端点：初始化 & 学生管理
# ==================================================================================

@router.post("/init-tables")
async def api_init_tables(user_info=Depends(verify_admin_or_teacher)):
    """手动初始化数据库表"""
    try:
        init_attendance_tables()
        return {"success": True, "message": "数据库表初始化成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"初始化失败: {str(e)}")


@router.post("/upload-students")
async def upload_students_csv(
    file: UploadFile = File(...),
    user_info=Depends(verify_admin_or_teacher),
):
    """上传 CSV 或 Excel 导入学生数据"""
    filename = file.filename.lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="请上传 CSV 或 Excel (.xlsx) 文件")

    try:
        content = await file.read()
        students = _parse_student_file(filename, content)

        if not students:
            raise HTTPException(status_code=400, detail="文件中没有有效的学生数据")

        result = _get_service().import_students(students)
        return {
            "success": True,
            "message": f"成功导入 {result['imported_count']} 名学生",
            "count": result["imported_count"],
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"上传学生数据失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


@router.get("/students")
async def get_students(
    class_name: Optional[str] = None,
    search: Optional[str] = None,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取学生列表"""
    students = _get_service().list_students(class_name, search)
    return {"success": True, "students": students}


@router.get("/classes")
async def get_classes(user_info=Depends(verify_admin_or_teacher)):
    """获取所有班级列表"""
    classes = _get_service().list_classes()
    return {"success": True, "classes": classes}


# ==================================================================================
#                               端点：会话管理
# ==================================================================================

@router.post("/sessions")
async def create_session(
    session: AttendanceSession,
    user_info=Depends(verify_admin_or_teacher),
):
    """创建点名会话（早读或留堂）"""
    username = user_info[0] if isinstance(user_info, (list, tuple)) else user_info.get("username", "system")

    svc = _get_service()
    if session.session_type == SessionType.MORNING:
        result = svc.create_morning_session(
            student_logins=session.student_ids,
            open_mode=session.open_mode or False,
            created_by=username,
        )
    else:
        result = svc.create_detention_session(
            student_logins=session.student_ids,
            reason=session.notes or "",
            created_by=username,
        )

    type_text = "早读" if session.session_type == SessionType.MORNING else "留堂"
    if session.open_mode:
        msg = f"成功创建{type_text}点名会话（开放模式）"
    else:
        msg = f"成功创建{type_text}点名会话，共 {len(session.student_ids)} 名学生"

    return {
        "success": True,
        "session_id": result["session_id"],
        "open_mode": session.open_mode,
        "message": msg,
    }


@router.get("/sessions")
async def get_sessions(
    session_type: Optional[str] = None,
    date: Optional[str] = None,
    status: str = "active",
    user_info=Depends(verify_admin_or_teacher),
):
    """获取点名会话列表"""
    sessions = _get_service().get_sessions_filtered(session_type, date, status)
    return {"success": True, "sessions": sessions}


@router.get("/sessions/{session_id}")
async def get_session_detail(
    session_id: int,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取点名会话详情"""
    result = _get_service().get_session_detail(session_id)
    return {"success": True, **result}


@router.get("/detention/sessions/{session_id}")
async def get_detention_session_detail(
    session_id: int,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取留堂会话详情"""
    result = _get_service().get_detention_session_detail(session_id)
    return {"success": True, **result}


@router.put("/sessions/{session_id}/complete")
async def complete_session(
    session_id: int,
    user_info=Depends(verify_admin_or_teacher),
):
    """结束点名会话"""
    _get_service().complete_session(session_id)
    return {"success": True, "message": "点名会话已结束"}


# ==================================================================================
#                               端点：早读签到
# ==================================================================================

def _format_scan_response(result: dict, is_manual: bool = False) -> dict:
    """
    将 service.morning_scan() 的扁平返回值转换为前端期望的嵌套结构。

    前端 showScanNotification / showScanResult 期望:
        data.student  — 学生对象
        data.record   — {scan_time, status, late_minutes, makeup_minutes, ...}
    """
    # already_checked_in 走错误分支：前端用 data.message 显示提示
    status = result.get("status")
    if status == "already_checked_in":
        return {
            "success": False,
            "student": result.get("student", {}),
            "record": {},
            "message": result.get("message", "该学生已签到"),
            "is_manual": is_manual,
        }

    student = result.pop("student", {})
    record = {
        "scan_time": result.get("scan_time"),
        "status": result.get("status"),
        "late_minutes": result.get("late_minutes", 0),
        "makeup_minutes": result.get("makeup_minutes", 0),
        "is_registered": result.get("is_registered", True),
    }
    return {
        "success": True,
        "student": student,
        "record": record,
        "is_manual": is_manual,
        **result,  # 保留其余字段（status、message 等）以兼容旧逻辑
    }


@router.post("/scan")
async def scan_card(
    request: CardScanRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """拍卡签到"""
    result = _get_service().morning_scan(
        session_id=request.session_id,
        card_id=request.card_id,
    )
    return _format_scan_response(result)


@router.post("/manual-scan")
async def manual_scan(
    request: ManualScanRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """手动签到（学生忘带卡时使用）"""
    result = _get_service().morning_scan(
        session_id=request.session_id,
        user_login=request.user_login,
    )
    return _format_scan_response(result, is_manual=True)


# ==================================================================================
#                               端点：留堂管理
# ==================================================================================

def _format_detention_response(result: dict, is_manual: bool = False) -> dict:
    """
    将 service.detention_checkin / detention_checkout / detention_smart_scan
    的扁平返回值转换为前端期望的嵌套结构。

    前端 showScanNotification 留堂模式期望:
        data.student   — 学生对象
        data.record    — {scan_time, checkout_time, planned_periods,
                          planned_minutes, planned_end_time, actual_minutes,
                          actual_periods, status, ...}
        data.action    — "checkout" / "need_select_periods" / "already_completed"
    """
    # need_select_periods 走前端模态框，不经过 showScanResult
    action = result.get("action")
    if action == "need_select_periods":
        return {"success": True, **result}

    # already_completed / error 仍会经过 showScanResult → showScanNotification，
    # 前端统一访问 data.record / data.student，因此需要提供空对象避免 TypeError。
    # success=False 让前端走"错误提示"分支（显示 data.message），避免进入签到成功逻辑。
    if action in ("already_completed", "error"):
        return {
            "success": False,
            "record": {},
            "student": result.get("student", {}),
            "message": result.get("message", "该学生已完成留堂"),
            **result,
        }

    student = result.pop("student", {})

    # 生成 status_msg（前端 showScanResult 需要）
    is_completed = result.get("is_completed")
    if is_completed is True:
        status_msg = "✅ 留堂完成"
    elif is_completed is False:
        status_msg = "⚠️ 提前离开"
    else:
        status_msg = "📝 留堂签到"

    record = {
        "scan_time": result.get("scan_time"),
        "checkout_time": result.get("checkout_time"),
        "status": result.get("status"),
        "status_msg": status_msg,
        "planned_periods": result.get("planned_periods"),
        "planned_minutes": result.get("planned_minutes"),
        "planned_end_time": result.get("planned_end_time"),
        "duration_minutes": result.get("duration_minutes"),
        "actual_minutes": result.get("actual_minutes"),
        "actual_periods": result.get("actual_periods"),
        "is_completed": result.get("is_completed"),
    }
    return {
        "success": True,
        "student": student,
        "record": record,
        "is_manual": is_manual,
        **result,  # 保留 action 等额外字段
    }


@router.post("/detention/checkin")
async def detention_checkin(
    request: DetentionCheckinRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """留堂签到 - 拍卡并选择节数或分钟数"""
    result = _get_service().detention_checkin(
        session_id=request.session_id,
        card_id=request.card_id,
        planned_periods=request.planned_periods,
        planned_minutes=request.planned_minutes,
        detention_reason=request.detention_reason,
    )
    return _format_detention_response(result)


@router.post("/detention/manual-checkin")
async def detention_manual_checkin(
    request: DetentionManualCheckinRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """留堂手动签到"""
    result = _get_service().detention_checkin(
        session_id=request.session_id,
        user_login=request.user_login,
        planned_periods=request.planned_periods,
        planned_minutes=request.planned_minutes,
        detention_reason=request.detention_reason,
    )
    return _format_detention_response(result, is_manual=True)


@router.post("/detention/checkout")
async def detention_checkout(
    request: CardScanRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """留堂签退 - 拍卡"""
    result = _get_service().detention_checkout(
        session_id=request.session_id,
        card_id=request.card_id,
    )
    result["action"] = "checkout"
    return _format_detention_response(result)


@router.post("/detention/manual-checkout")
async def detention_manual_checkout(
    request: ManualScanRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """留堂手动签退"""
    result = _get_service().detention_checkout(
        session_id=request.session_id,
        user_login=request.user_login,
    )
    result["action"] = "checkout"
    return _format_detention_response(result, is_manual=True)


@router.post("/detention/scan")
async def detention_smart_scan(
    request: CardScanRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """留堂智能拍卡（自动判断签到/签退）"""
    result = _get_service().detention_smart_scan(
        session_id=request.session_id,
        card_id=request.card_id,
    )
    return _format_detention_response(result)


@router.put("/detention/modify-periods")
async def modify_detention_periods(
    request: ModifyPeriodsRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """修改留堂节数或分钟数"""
    result = _get_service().modify_detention_periods(
        session_id=request.session_id,
        user_login=request.user_login,
        new_periods=request.new_periods,
        new_minutes=request.new_minutes,
    )
    return {"success": True, "message": "已修改", **result}


@router.put("/detention/modify-end-time")
async def modify_detention_end_time(
    request: ModifyEndTimeRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """直接修改留堂结束时间"""
    result = _get_service().modify_detention_end_time(
        session_id=request.session_id,
        user_login=request.user_login,
        new_end_time=request.new_end_time,
    )
    return {"success": True, "message": f"已修改结束时间为 {request.new_end_time}", **result}


# ==================================================================================
#                               端点：固定名单
# ==================================================================================

@router.post("/fixed-lists")
async def create_fixed_list(
    request: FixedListRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """创建固定名单"""
    username = user_info[0] if isinstance(user_info, (list, tuple)) else "system"
    list_id = _get_service().create_fixed_list(
        list_name=request.list_name,
        list_type=request.list_type,
        student_logins=request.student_ids,
        created_by=username,
    )
    return {
        "success": True,
        "list_id": list_id,
        "message": f"成功保存名单 '{request.list_name}'，共 {len(request.student_ids)} 名学生",
    }


@router.get("/fixed-lists")
async def get_fixed_lists(
    list_type: Optional[str] = None,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取固定名单列表"""
    lists = _get_service().get_fixed_lists(list_type)
    return {"success": True, "lists": lists}


@router.get("/fixed-lists/{list_id}")
async def get_fixed_list_detail(
    list_id: int,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取固定名单详情"""
    detail = _get_service().get_fixed_list_detail(list_id)
    return {"success": True, "list": detail, "students": detail.get("students", [])}


@router.put("/fixed-lists/{list_id}")
async def update_fixed_list(
    list_id: int,
    request: FixedListRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """编辑固定名单"""
    _get_service().update_fixed_list(
        list_id=list_id,
        list_name=request.list_name,
        list_type=request.list_type,
        student_logins=request.student_ids,
    )
    return {
        "success": True,
        "list_id": list_id,
        "message": f"成功更新名单 '{request.list_name}'，共 {len(request.student_ids)} 名学生",
    }


@router.delete("/fixed-lists/{list_id}")
async def delete_fixed_list(
    list_id: int,
    user_info=Depends(verify_admin_or_teacher),
):
    """删除固定名单"""
    _get_service().delete_fixed_list(list_id)
    return {"success": True, "message": "名单已删除"}


# ==================================================================================
#                               端点：留堂历史
# ==================================================================================

@router.get("/detention-history")
async def get_detention_history(
    user_login: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    completed: Optional[bool] = None,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取留堂历史记录"""
    history = _get_service().get_detention_history_filtered(
        user_login=user_login,
        start_date=start_date,
        end_date=end_date,
        completed=completed,
    )
    return {"success": True, "history": history}


@router.get("/detention-summary/{user_login}")
async def get_detention_summary(
    user_login: str,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取学生留堂汇总"""
    summary = _get_service().get_detention_summary(user_login)
    return {"success": True, **summary}


# ==================================================================================
#                               端点：课外活动
# ==================================================================================

@router.get("/activity-groups")
async def get_activity_groups(user_info=Depends(verify_admin_or_teacher)):
    """获取所有固定组别"""
    groups = _get_service().list_activity_groups()
    return {"success": True, "groups": groups}


@router.get("/activity-groups/{group_id}")
async def get_activity_group(group_id: int, user_info=Depends(verify_admin_or_teacher)):
    """获取组别详情及学生列表"""
    detail = _get_service().get_activity_group_detail(group_id)
    return {"success": True, "group": detail, "students": detail.get("students", [])}


@router.post("/activity-groups")
async def create_activity_group(
    request: ActivityGroupRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """创建固定组别"""
    group_id = _get_service().create_activity_group(
        group_name=request.name,
        student_logins=request.student_ids,
        created_by=user_info[0],
    )
    return {"success": True, "group_id": group_id, "message": f"组别 '{request.name}' 创建成功"}


@router.put("/activity-groups/{group_id}")
async def update_activity_group(
    group_id: int,
    request: ActivityGroupRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """更新组别"""
    _get_service().update_activity_group(
        group_id=group_id,
        group_name=request.name,
        student_logins=request.student_ids,
    )
    return {"success": True, "message": "组别已更新"}


@router.delete("/activity-groups/{group_id}")
async def delete_activity_group(group_id: int, user_info=Depends(verify_admin_or_teacher)):
    """删除组别"""
    _get_service().delete_activity_group(group_id)
    return {"success": True, "message": "组别已删除"}


@router.post("/activity/sessions")
async def create_activity_session(
    request: ActivitySessionRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """创建课外活动会话"""
    username = user_info[0] if isinstance(user_info, (list, tuple)) else "system"
    if not request.student_ids:
        raise HTTPException(status_code=400, detail="请选择参加活动的学生")

    result = _get_service().create_activity_session(
        activity_name=request.activity_name,
        student_logins=request.student_ids,
        start_time=request.start_time,
        end_time=request.end_time,
        late_threshold=request.late_threshold,
        early_threshold=request.early_threshold,
        created_by=username,
    )
    return {
        "success": True,
        "session_id": result["session_id"],
        "message": f"课外活动 '{request.activity_name}' 点名已开始",
    }


@router.get("/activity/sessions/{session_id}")
async def get_activity_session(session_id: int, user_info=Depends(verify_admin_or_teacher)):
    """获取课外活动会话详情"""
    svc = _get_service()
    detail = svc._activity_session.get_session_detail(session_id)
    if not detail:
        raise HTTPException(status_code=404, detail="会话不存在")

    records = detail.get("records", [])
    stats = {
        "on_time": sum(1 for s in records if s.get("check_in_status") == ActivityCheckinStatus.ON_TIME),
        "late": sum(1 for s in records if s.get("check_in_status") == ActivityCheckinStatus.LATE),
        "absent": sum(1 for s in records if s.get("check_in_status") == ActivityCheckinStatus.NOT_ARRIVED),
        "normal_leave": sum(1 for s in records if s.get("check_out_status") == ActivityCheckoutStatus.NORMAL),
        "early_leave": sum(1 for s in records if s.get("check_out_status") == ActivityCheckoutStatus.EARLY),
        "still_here": sum(
            1 for s in records
            if s.get("check_in_status") != ActivityCheckinStatus.NOT_ARRIVED
            and s.get("check_out_status") in (ActivityCheckoutStatus.NOT_ARRIVED, ActivityCheckoutStatus.STILL_HERE)
        ),
    }
    return {"success": True, "session": detail, "students": records, "stats": stats}


def _format_activity_response(result: dict) -> dict:
    """
    将 service.activity_checkin / activity_checkout 的扁平返回值
    转换为前端期望的嵌套结构。

    前端 showScanNotification / showScanResult 期望:
        data.student  — 学生对象
        data.record   — {scan_time, checkout_time, ...}
        data.action   — "checkin" / "checkout"
        data.time     — 时间戳（兼容旧逻辑）
    """
    student = result.pop("student", {})
    record = {
        "scan_time": result.get("time"),
        "checkout_time": result.get("time"),
    }
    return {
        "success": True,
        "student": student,
        "record": record,
        **result,  # 保留 action, time, is_late, late_minutes, is_early, early_minutes
    }


@router.post("/activity/scan")
async def activity_scan(
    request: ActivityScanRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """课外活动拍卡签到"""
    result = _get_service().activity_checkin(
        session_id=request.session_id,
        card_id=request.card_id,
    )
    return _format_activity_response(result)


@router.post("/activity/checkout")
async def activity_checkout(
    request: ActivityCheckoutRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """课外活动手动签退"""
    result = _get_service().activity_checkout(
        session_id=request.session_id,
        user_login=request.user_login,
    )
    return _format_activity_response(result)


@router.post("/activity/sessions/{session_id}/end")
async def end_activity_session(session_id: int, user_info=Depends(verify_admin_or_teacher)):
    """结束课外活动会话"""
    _get_service().end_activity_session(session_id)
    return {"success": True, "message": "活动已结束"}


# ==================================================================================
#                               端点：Excel 导出
# ==================================================================================

EXPORTS_DIR = "attendance_exports"
try:
    if not os.path.exists(EXPORTS_DIR):
        os.makedirs(EXPORTS_DIR)
except Exception as _e:
    logger.error(f"创建导出目录失败: {_e}")


class ExportSaveRequest(BaseModel):
    """保存导出请求"""
    session_id: int
    notes: Optional[str] = None


@router.get("/export/{session_id}")
async def export_session_excel(
    session_id: int,
    user_info=Depends(verify_token_from_query_or_header),
):
    """导出早读点名记录到 Excel"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel导出功能不可用")

    svc = _get_service()
    session_detail = svc.get_session_detail(session_id)
    session = session_detail.get("session", session_detail)
    records = session_detail.get("records", [])

    wb, filename = _build_morning_export_excel(session, records)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/detention/export/{session_id}")
async def export_detention_excel(
    session_id: int,
    user_info=Depends(verify_token_from_query_or_header),
):
    """导出留堂记录到 Excel"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel导出功能不可用")

    svc = _get_service()
    detail = svc.get_detention_session_detail(session_id)
    session = detail.get("session", {})
    students = detail.get("students", [])

    wb, filename = _build_detention_export_excel(session, students)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/exports/save")
async def save_attendance_export(
    request: ExportSaveRequest,
    user_info=Depends(verify_admin_or_teacher),
):
    """保存点名记录到服务器"""
    username, display_name = _extract_username(user_info)
    svc = _get_service()

    session_detail = svc.get_session_detail(request.session_id)
    session = session_detail.get("session", session_detail)
    records = session_detail.get("records", [])
    session_type = session.get("session_type", SessionType.MORNING)
    session_date = session.get("session_date")

    # 生成 Excel
    if session_type == SessionType.MORNING:
        wb, _ = _build_morning_export_excel(session, records)
    else:
        detail = svc.get_detention_session_detail(request.session_id)
        wb, _ = _build_detention_export_excel(detail.get("session", {}), detail.get("students", []))
        records = detail.get("students", [])

    # 保存文件
    type_text = "早读" if session_type == SessionType.MORNING else "留堂"
    date_str = session_date.strftime("%Y-%m-%d") if hasattr(session_date, "strftime") else str(session_date)
    file_display_name = f"{type_text}点名_{date_str}_{display_name}.xlsx"
    file_uuid = str(uuid.uuid4())
    file_path = os.path.join(EXPORTS_DIR, f"{file_uuid}.xlsx")
    wb.save(file_path)
    file_size = os.path.getsize(file_path)

    # 统计
    stats = _calc_export_stats(session_type, records)
    export_id = svc.save_export(
        session_id=request.session_id,
        session_type=session_type,
        file_path=file_path,
        file_size=file_size,
        created_by=username,
        stats={
            "session_date": date_str,
            "created_by_name": display_name,
            "file_name": file_display_name,
            **stats,
        },
    )

    return {"success": True, "message": "点名记录已保存", "export_id": export_id, "file_name": file_display_name}


@router.get("/exports/list")
async def list_attendance_exports(
    page: int = 1,
    page_size: int = 20,
    session_type: Optional[str] = None,
    user_info=Depends(verify_admin_or_teacher),
):
    """获取点名导出记录列表"""
    username, _ = _extract_username(user_info)
    result = _get_service().list_exports(username, session_type, page, page_size)
    # paginate() 返回 {items, total, page, page_size}
    # 前端期望 {records, total, page, total_pages}
    records = result.get("items", [])
    for rec in records:
        fp = rec.pop("file_path", None)
        rec["file_available"] = bool(fp and os.path.exists(fp))
    total = result.get("total", 0)
    ps = result.get("page_size", page_size)
    total_pages = (total + ps - 1) // ps if ps > 0 else 0
    return {
        "success": True,
        "records": records,
        "total": total,
        "page": result.get("page", page),
        "total_pages": total_pages,
    }


@router.get("/exports/download/{export_id}")
async def download_attendance_export(
    export_id: int,
    user_info=Depends(verify_admin_or_teacher),
):
    """下载导出文件"""
    username, _ = _extract_username(user_info)
    try:
        export_info = _get_service().get_export_file(export_id, username)
    except Exception:
        logger.warning("Export %s: DB record not found for user=%s", export_id, username)
        raise HTTPException(status_code=404, detail="导出记录不存在或无权访问")

    file_path = export_info.get("file_path")
    file_name = export_info.get("file_name") or "attendance_export.xlsx"

    if not file_path or not os.path.exists(file_path):
        logger.warning("Export %s: file missing on disk, path=%s", export_id, file_path)
        raise HTTPException(status_code=404, detail="文件已过期，请重新导出")

    return FileResponse(
        path=file_path,
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.delete("/exports/{export_id}")
async def delete_attendance_export(
    export_id: int,
    user_info=Depends(verify_admin_or_teacher),
):
    """删除导出记录"""
    username, _ = _extract_username(user_info)
    _get_service().delete_export(export_id, username)
    return {"success": True, "message": "记录已删除"}


@router.get("/export-detention-history")
async def export_detention_history_excel(
    user_login: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user_info=Depends(verify_admin_or_teacher),
):
    """导出留堂历史到 Excel"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel导出功能不可用")

    history = _get_service().get_detention_history_filtered(
        user_login=user_login,
        start_date=start_date,
        end_date=end_date,
    )

    wb = _build_detention_history_excel(history)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"留堂历史记录_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/activity/export/{session_id}")
async def export_activity_excel(
    session_id: int,
    token: str = Query(None),
    credentials: HTTPAuthorizationCredentials = Depends(HTTPBearer(auto_error=False)),
):
    """导出课外活动 Excel"""
    await verify_token_from_query_or_header(token, credentials)

    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel导出功能不可用")

    svc = _get_service()
    detail = svc._activity_session.get_session_detail(session_id)
    if not detail:
        raise HTTPException(status_code=404, detail="会话不存在")

    records = detail.get("records", [])
    wb = _build_activity_export_excel(detail, records)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"课外活动_{detail.get('activity_name', '')}_{detail.get('session_date', '')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


# ==================================================================================
#                               Excel 生成辅助函数
# ==================================================================================

def _excel_styles():
    """共用的 Excel 样式"""
    return {
        "header_font": Font(bold=True, color=ExcelColors.WHITE),
        "header_fill": PatternFill(start_color=ExcelColors.HEADER_BLUE, end_color=ExcelColors.HEADER_BLUE, fill_type="solid"),
        "green_fill": PatternFill(start_color=ExcelColors.SUCCESS_GREEN, end_color=ExcelColors.SUCCESS_GREEN, fill_type="solid"),
        "orange_fill": PatternFill(start_color=ExcelColors.WARNING_ORANGE, end_color=ExcelColors.WARNING_ORANGE, fill_type="solid"),
        "red_fill": PatternFill(start_color=ExcelColors.ERROR_RED, end_color=ExcelColors.ERROR_RED, fill_type="solid"),
        "gray_fill": PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"),
        "homework_fill": PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid"),
        "morning_fill": PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
        "both_fill": PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid"),
        "other_reason_fill": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }


def _build_morning_export_excel(session, records):
    """生成早读点名 Excel"""
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "早读点名"

    ws["A1"] = "早读点名记录"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A2"] = f"日期: {session.get('session_date', '')}"
    ws["D2"] = f"应到时间: {session.get('target_time', '')}"

    headers = ["班级", "班号", "学号", "英文名", "中文名", "签到时间", "状态", "迟到(分钟)", "需补时(分钟)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = s["thin_border"]

    present_count = late_count = very_late_count = absent_count = 0
    for row_idx, record in enumerate(records, 5):
        status = record.get("status") or record.get("attendance_status") or AttendanceStatus.ABSENT
        if status == AttendanceStatus.PRESENT:
            present_count += 1
        elif status == AttendanceStatus.LATE:
            late_count += 1
        elif status == AttendanceStatus.VERY_LATE:
            very_late_count += 1
        else:
            absent_count += 1

        status_cn = AttendanceStatus.LABELS_ZH.get(status, "缺席")
        scan_time = record.get("scan_time")
        if scan_time and hasattr(scan_time, "strftime"):
            scan_time = scan_time.strftime("%H:%M:%S")
        elif not scan_time:
            scan_time = "-"

        row_data = [
            record.get("class_name"), record.get("class_number"), record.get("user_login"),
            record.get("english_name"), record.get("chinese_name"), scan_time, status_cn,
            record.get("late_minutes") or 0, record.get("makeup_minutes") or 0,
        ]
        fill_map = {AttendanceStatus.PRESENT: s["green_fill"], AttendanceStatus.LATE: s["orange_fill"], AttendanceStatus.VERY_LATE: s["red_fill"]}
        fill = fill_map.get(status, s["gray_fill"])
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            if status == AttendanceStatus.VERY_LATE:
                cell.font = Font(color=ExcelColors.WHITE)

    summary_row = len(records) + 6
    ws.cell(row=summary_row, column=1, value="统计").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"应到: {len(records)}")
    ws.cell(row=summary_row, column=3, value=f"准时: {present_count}").fill = s["green_fill"]
    ws.cell(row=summary_row, column=4, value=f"迟到: {late_count + very_late_count}").fill = s["orange_fill"]
    ws.cell(row=summary_row, column=5, value=f"缺席: {absent_count}").fill = s["gray_fill"]

    for col, w in enumerate([8, 6, 15, 20, 15, 12, 14, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    filename = f"早读点名_{session.get('session_date', '')}.xlsx"
    return wb, filename


def _build_detention_export_excel(session, records):
    """生成留堂 Excel"""
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "留堂记录"

    ws["A1"] = "留堂记录"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")
    ws["A2"] = f"日期: {session.get('session_date', '')}"

    headers = ["班级", "班号", "学号", "英文名", "中文名", "签到时间", "签退时间",
               "计划时长", "实际分钟", "实际节数", "状态", "原因"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = s["thin_border"]

    completed_count = incomplete_count = absent_count = 0
    for row_idx, record in enumerate(records, 5):
        status = record.get("status") or AttendanceStatus.ABSENT
        fill = None
        if status == AttendanceStatus.DETENTION_COMPLETED:
            planned_mins = record.get("planned_minutes")
            actual_mins = record.get("actual_minutes") or 0
            planned_p = record.get("planned_periods") or 0
            actual_p = record.get("actual_periods") or 0
            is_done = (actual_mins >= planned_mins) if planned_mins is not None else (actual_p >= planned_p)
            if is_done:
                status_cn, fill = "已完成", s["green_fill"]
                completed_count += 1
            else:
                status_cn, fill = "未完成", s["orange_fill"]
                incomplete_count += 1
        elif status == AttendanceStatus.DETENTION_ACTIVE:
            status_cn = "进行中"
        else:
            status_cn, fill = "未签到", s["gray_fill"]
            absent_count += 1

        scan_t = record.get("scan_time")
        scan_t = scan_t.strftime("%H:%M:%S") if scan_t and hasattr(scan_t, "strftime") else "-"
        out_t = record.get("checkout_time")
        out_t = out_t.strftime("%H:%M:%S") if out_t and hasattr(out_t, "strftime") else "-"

        reason = record.get("detention_reason") or ""
        reason_display = DetentionReason.LABELS_ZH.get(reason, reason or "未知")

        planned_mins_display = record.get("planned_minutes")
        planned_display = (f"{planned_mins_display}分钟" if planned_mins_display is not None
                           else f"{record.get('planned_periods') or 0}节")

        row_data = [
            record.get("class_name"), record.get("class_number"), record.get("user_login"),
            record.get("english_name"), record.get("chinese_name"), scan_t, out_t,
            planned_display, record.get("actual_minutes") or 0, record.get("actual_periods") or 0,
            status_cn, reason_display,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")
            if col <= 11 and fill:
                cell.fill = fill
            elif col == 12:
                if reason == DetentionReason.HOMEWORK:
                    cell.fill = s["homework_fill"]
                elif reason == DetentionReason.MORNING:
                    cell.fill = s["morning_fill"]
                elif reason == DetentionReason.BOTH:
                    cell.fill = s["both_fill"]
                elif reason:
                    cell.fill = s["other_reason_fill"]

    summary_row = len(records) + 6
    ws.cell(row=summary_row, column=1, value="统计").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"总人数: {len(records)}")
    ws.cell(row=summary_row, column=3, value=f"已完成: {completed_count}").fill = s["green_fill"]
    ws.cell(row=summary_row, column=4, value=f"未完成: {incomplete_count}").fill = s["orange_fill"]
    ws.cell(row=summary_row, column=5, value=f"未签到: {absent_count}").fill = s["gray_fill"]

    for col, w in enumerate([8, 6, 15, 20, 15, 12, 12, 10, 10, 10, 12, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    filename = f"留堂记录_{session.get('session_date', '')}.xlsx"
    return wb, filename


def _build_detention_history_excel(history):
    """生成留堂历史 Excel"""
    s = _excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "留堂历史"

    headers = ["日期", "班级", "班号", "学号", "英文名", "中文名", "时长(分钟)", "原因", "已完成", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = s["header_font"]
        cell.fill = s["header_fill"]
        cell.alignment = Alignment(horizontal="center")
        cell.border = s["thin_border"]

    for row_idx, record in enumerate(history, 2):
        reason = record.get("reason") or ""
        reason_display = DetentionReason.LABELS_ZH.get(reason, reason or "未知")
        row_data = [
            str(record.get("detention_date", "")), record.get("class_name"), record.get("class_number"),
            record.get("user_login"), record.get("english_name"), record.get("chinese_name"),
            record.get("duration_minutes"), reason_display,
            "是" if record.get("completed") else "否", record.get("notes") or "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = s["thin_border"]
            cell.alignment = Alignment(horizontal="center")
            if col == 8:
                if reason == DetentionReason.HOMEWORK:
                    cell.fill = s["homework_fill"]
                elif reason == DetentionReason.MORNING:
                    cell.fill = s["morning_fill"]
                elif reason == DetentionReason.BOTH:
                    cell.fill = s["both_fill"]
                elif reason:
                    cell.fill = s["other_reason_fill"]

    for col, w in enumerate([12, 8, 6, 15, 20, 15, 12, 30, 8, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    return wb


def _build_activity_export_excel(session, records):
    """生成课外活动 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "课外活动点名"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"{session.get('activity_name', '')} - {session.get('session_date', '')}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = f"应到时间: {session.get('start_time', '')} | 结束时间: {session.get('end_time', '')}"

    headers = ["班级", "学号", "姓名", "签到时间", "签到状态", "签退时间", "签退状态", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    checkin_colors = {ActivityCheckinStatus.ON_TIME: "34C759", ActivityCheckinStatus.LATE: "FF9500", ActivityCheckinStatus.NOT_ARRIVED: "FF3B30"}
    checkout_colors = {ActivityCheckoutStatus.NORMAL: "34C759", ActivityCheckoutStatus.EARLY: "FF9500", ActivityCheckoutStatus.STILL_HERE: "007AFF", ActivityCheckoutStatus.NOT_ARRIVED: "8E8E93"}
    checkin_text = {ActivityCheckinStatus.ON_TIME: "准时", ActivityCheckinStatus.LATE: "迟到", ActivityCheckinStatus.NOT_ARRIVED: "未到"}
    checkout_text = {ActivityCheckoutStatus.NORMAL: "正常", ActivityCheckoutStatus.EARLY: "早退", ActivityCheckoutStatus.STILL_HERE: "仍在", ActivityCheckoutStatus.NOT_ARRIVED: "-"}

    for row_idx, record in enumerate(records, 5):
        ws.cell(row=row_idx, column=1, value=f"{record.get('class_name', '')}-{record.get('class_number', '')}")
        ws.cell(row=row_idx, column=2, value=record.get("user_login"))
        ws.cell(row=row_idx, column=3, value=record.get("chinese_name"))

        ci_time = record.get("check_in_time")
        ws.cell(row=row_idx, column=4, value=ci_time.strftime("%H:%M:%S") if ci_time else "-")

        ci_status = record.get("check_in_status", ActivityCheckinStatus.NOT_ARRIVED)
        ci_cell = ws.cell(row=row_idx, column=5, value=checkin_text.get(ci_status, "-"))
        color = checkin_colors.get(ci_status, ExcelColors.WHITE)
        ci_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        co_time = record.get("check_out_time")
        ws.cell(row=row_idx, column=6, value=co_time.strftime("%H:%M:%S") if co_time else "-")

        co_status = record.get("check_out_status", ActivityCheckoutStatus.NOT_ARRIVED)
        co_cell = ws.cell(row=row_idx, column=7, value=checkout_text.get(co_status, "-"))
        color = checkout_colors.get(co_status, ExcelColors.WHITE)
        co_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.cell(row=row_idx, column=8, value=record.get("notes", ""))

    for letter, w in zip("ABCDEFGH", [12, 15, 15, 12, 10, 12, 10, 20]):
        ws.column_dimensions[letter].width = w

    return wb


def _calc_export_stats(session_type, records):
    """计算导出统计数据"""
    student_count = len(records)
    present_count = late_count = absent_count = 0

    if session_type == SessionType.MORNING:
        for r in records:
            st = r.get("status") or r.get("attendance_status") or AttendanceStatus.ABSENT
            if st in (AttendanceStatus.PRESENT, AttendanceStatus.LATE, AttendanceStatus.VERY_LATE):
                present_count += 1
            if st in (AttendanceStatus.LATE, AttendanceStatus.VERY_LATE):
                late_count += 1
            if st == AttendanceStatus.ABSENT:
                absent_count += 1
    else:
        for r in records:
            st = r.get("status") or AttendanceStatus.ABSENT
            if st in (AttendanceStatus.DETENTION_COMPLETED, AttendanceStatus.DETENTION_ACTIVE):
                present_count += 1
            if st == AttendanceStatus.DETENTION_COMPLETED:
                pm = r.get("planned_minutes")
                am = r.get("actual_minutes") or 0
                pp = r.get("planned_periods") or 0
                ap = r.get("actual_periods") or 0
                if (pm is not None and am < pm) or (pm is None and ap < pp):
                    late_count += 1
            if st not in (AttendanceStatus.DETENTION_COMPLETED, AttendanceStatus.DETENTION_ACTIVE):
                absent_count += 1

    return {
        "student_count": student_count,
        "present_count": present_count,
        "late_count": late_count,
        "absent_count": absent_count,
    }


# ==================================================================================
#                               文件解析辅助
# ==================================================================================

def _parse_student_file(filename: str, content: bytes) -> list:
    """解析 CSV 或 Excel 学生文件"""
    students = []

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            header_map = {}
            for i, h in enumerate(headers):
                if h:
                    h_lower = str(h).lower().strip()
                    if "classname" in h_lower or h == "班级":
                        header_map["class_name"] = i
                    elif "classnumber" in h_lower or h == "班号":
                        header_map["class_number"] = i
                    elif "userlogin" in h_lower or h == "学号":
                        header_map["user_login"] = i
                    elif "englishname" in h_lower or h == "英文名":
                        header_map["english_name"] = i
                    elif "chinesename" in h_lower or h == "中文名":
                        header_map["chinese_name"] = i
                    elif "cardid" in h_lower or h == "卡号":
                        header_map["card_id"] = i

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                user_login = str(row[header_map.get("user_login", 2)] or "").strip()
                if user_login:
                    cn_val = row[header_map.get("class_number", 1)] or 0
                    students.append({
                        "class_name": str(row[header_map.get("class_name", 0)] or "").strip(),
                        "class_number": int(cn_val) if str(cn_val).isdigit() else 0,
                        "user_login": user_login,
                        "english_name": str(row[header_map.get("english_name", 3)] or "").strip(),
                        "chinese_name": str(row[header_map.get("chinese_name", 4)] or "").strip(),
                        "card_id": str(row[header_map.get("card_id", 5)] or "").strip() or None,
                    })
        except ImportError:
            raise HTTPException(status_code=400, detail="服务器未安装 openpyxl")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Excel 解析失败: {str(e)}")
    else:
        # CSV
        text = None
        for encoding in ["utf-8", "utf-8-sig", "gbk", "gb2312", "big5", "cp1252"]:
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if text is None:
            raise HTTPException(status_code=400, detail="无法识别文件编码")

        first_line = text.split("\n")[0]
        delimiter = "\t" if "\t" in first_line else (";" if ";" in first_line else ",")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

        for row in reader:
            user_login = (row.get("UserLogin") or row.get("学号") or "").strip()
            if user_login:
                cn = (row.get("ClassNumber") or row.get("班号") or "0").strip()
                students.append({
                    "class_name": (row.get("ClassName") or row.get("班级") or "").strip(),
                    "class_number": int(cn) if cn.isdigit() else 0,
                    "user_login": user_login,
                    "english_name": (row.get("EnglishName") or row.get("英文名") or "").strip(),
                    "chinese_name": (row.get("ChineseName") or row.get("中文名") or "").strip(),
                    "card_id": (row.get("CardID") or row.get("卡号") or "").strip() or None,
                })

    return students


# ==================================================================================
#                               向后兼容
# ==================================================================================

def setup_attendance_router(app):
    """设置点名系统路由（兼容旧调用）"""
    app.include_router(router)
    logger.info("点名系统路由已加载")
