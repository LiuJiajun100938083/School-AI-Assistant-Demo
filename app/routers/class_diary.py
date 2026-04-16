"""
課室日誌路由
=============
教師掃碼課堂評級 + Review 查看 + 管理員 QR 碼 / Reviewer 管理。

教師端點（需教師/管理員登入）:
    POST  /api/class-diary/entries              提交評級
    GET   /api/class-diary/entries/by-class      查某班某日記錄（公開）
    PUT   /api/class-diary/entries/{id}          修改記錄

需登入端點:
    GET   /api/class-diary/review               Review 記錄
    GET   /api/class-diary/review/classes        班級列表
    GET   /api/class-diary/review/summary        匯總

管理員端點:
    GET   /api/class-diary/admin/qr/{code}       生成 QR 碼
    GET   /api/class-diary/admin/qr/batch        批量 QR ZIP
    GET   /api/class-diary/admin/reviewers       列出 reviewer
    POST  /api/class-diary/admin/reviewers       添加 reviewer
    DELETE /api/class-diary/admin/reviewers/{u}   移除 reviewer
    GET   /api/class-diary/admin/classes          班級列表
"""

import csv
import io
import json
import logging
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Query, Request
from fastapi.responses import Response, StreamingResponse

from app.core.exceptions import AppException, AuthorizationError
from app.core.responses import error_response, success_response
from app.domains.class_diary.schemas import (
    CreateEntryRequest,
    ReviewerRequest,
    UpdateEntryRequest,
)
from app.domains.class_diary.audit import ClassDiaryAuditLogger as _audit
from app.domains.class_diary.task_manager import task_manager as _tasks
from app.infrastructure.database.pool import get_database_pool

logger = logging.getLogger(__name__)

router = APIRouter(tags=["課室日誌"])


# ====================================================================== #
#  工具函數                                                                #
# ====================================================================== #

def _get_service():
    from app.services.container import get_services
    return get_services().class_diary


def _format_behavior_field(value: str) -> str:
    """Format JSON behavior field for display, with backward compatibility.

    New format:    [{"reason_code": "CHAT", "reason_text": "聊天", "students": [...]}]
    Legacy format: [{"reason": "聊天", "students": [...]}]
    Plain text:    comma/separator separated student names
    """
    if not value or not value.strip():
        return ""
    try:
        data = json.loads(value)
        if isinstance(data, list):
            parts = []
            for item in data:
                reason = item.get("reason_text") or item.get("reason", "")
                students = item.get("students", [])
                if students:
                    joined = "、".join(students)
                    parts.append(f"{reason}: {joined}" if reason else joined)
            return "；".join(parts)
    except (json.JSONDecodeError, TypeError, AttributeError):
        pass
    return value


def _verify_request(request: Request):
    """從 JWT 提取 (username, role)"""
    from app.services.container import get_services
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        token = request.query_params.get("token", "")
    if not token:
        raise AppException(code="UNAUTHORIZED", message="未登入", status_code=401)
    payload = get_services().auth.verify_token(token)
    return payload["username"], payload["role"]


def _verify_admin(request: Request):
    username, role = _verify_request(request)
    if role != "admin":
        raise AuthorizationError("需要管理員權限")
    return username, role


def _verify_reviewer(request: Request):
    """驗證 reviewer 或 admin 權限"""
    username, role = _verify_request(request)
    service = _get_service()
    if not service.check_review_access(username, role):
        from app.domains.class_diary.exceptions import ReviewAccessDeniedError
        raise ReviewAccessDeniedError()
    return username, role


def _get_base_url(request: Request) -> str:
    """從請求中提取站點根 URL"""
    scheme = request.headers.get("x-forwarded-proto", request.url.scheme)
    host = request.headers.get("x-forwarded-host", request.headers.get("host", "localhost"))
    return f"{scheme}://{host}"


# ====================================================================== #
#  數據庫初始化                                                            #
# ====================================================================== #

def init_class_diary_tables():
    """創建課室日誌相關數據表"""
    pool = get_database_pool()

    sqls = [
        """
        CREATE TABLE IF NOT EXISTS class_diary_entries (
            id                  INT AUTO_INCREMENT PRIMARY KEY,
            class_code          VARCHAR(20)  NOT NULL          COMMENT '班級代碼',
            entry_date          DATE         NOT NULL          COMMENT '上課日期',
            period_start        TINYINT      NOT NULL          COMMENT '起始節數 0=早會,1-9',
            period_end          TINYINT      NOT NULL          COMMENT '結束節數',
            subject             VARCHAR(100) NOT NULL          COMMENT '科目',
            absent_students     TEXT                           COMMENT '缺席學生',
            late_students       TEXT                           COMMENT '遲到學生',
            discipline_rating   TINYINT      NOT NULL DEFAULT 0 COMMENT '紀律 1-5',
            cleanliness_rating  TINYINT      NOT NULL DEFAULT 0 COMMENT '整潔 1-5',
            commended_students  TEXT                           COMMENT '嘉許學生',
            appearance_issues   TEXT                           COMMENT '儀表違規',
            rule_violations     TEXT                           COMMENT '課堂違規',
            signature           MEDIUMTEXT                     COMMENT '手寫簽名 base64',
            submitted_from      VARCHAR(255)                   COMMENT '提交來源 UA',
            created_at          TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at          TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            INDEX idx_class_date (class_code, entry_date),
            INDEX idx_date (entry_date)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """,
        """
        CREATE TABLE IF NOT EXISTS class_diary_reviewers (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            username    VARCHAR(100) NOT NULL UNIQUE     COMMENT '被授權用戶名',
            granted_by  VARCHAR(100) NOT NULL            COMMENT '授權管理員',
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """,
        """
        CREATE TABLE IF NOT EXISTS class_diary_daily_reports (
            id              INT AUTO_INCREMENT PRIMARY KEY,
            report_date     DATE NOT NULL UNIQUE          COMMENT '報告日期',
            report_text     MEDIUMTEXT NOT NULL            COMMENT 'AI 生成的報告文本',
            anomalies_json  MEDIUMTEXT                    COMMENT '異常記錄 JSON',
            status          VARCHAR(20) DEFAULT 'pending' COMMENT 'pending/generating/done/failed',
            created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """,
        """
        CREATE TABLE IF NOT EXISTS class_diary_report_recipients (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            username    VARCHAR(100) NOT NULL UNIQUE     COMMENT '報告接收人用戶名',
            granted_by  VARCHAR(100) NOT NULL            COMMENT '授權管理員',
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """,
    ]

    for sql in sqls:
        try:
            pool.execute_write(sql)
        except Exception as e:
            # 表已存在時忽略
            if "already exists" not in str(e).lower():
                logger.error("創建課室日誌表失敗: %s", e)

    # 安全新增 submitted_by 欄位（舊表可能沒有）
    try:
        pool.execute_write(
            "ALTER TABLE class_diary_entries "
            "ADD COLUMN submitted_by VARCHAR(100) DEFAULT NULL "
            "COMMENT '提交者用戶名' AFTER submitted_from"
        )
        logger.info("已新增 submitted_by 欄位")
    except Exception:
        pass  # 欄位已存在

    # 新增 UNIQUE 約束防重複節數（若不存在）
    try:
        pool.execute_write(
            "ALTER TABLE class_diary_entries "
            "ADD UNIQUE INDEX uq_class_date_period "
            "(class_code, entry_date, period_start, period_end)"
        )
        logger.info("已新增 uq_class_date_period UNIQUE 約束")
    except Exception:
        pass  # 約束已存在

    # 新增 medical_room_students 欄位（上次遷移可能未執行）
    try:
        pool.execute_write(
            "ALTER TABLE class_diary_entries "
            "ADD COLUMN medical_room_students TEXT DEFAULT NULL "
            "COMMENT '醫務室' AFTER rule_violations"
        )
        logger.info("已新增 medical_room_students 欄位")
    except Exception:
        pass  # 欄位已存在

    # 新增 summary_text 欄位（報告摘要版，供 report_recipient 查看）
    try:
        pool.execute_write(
            "ALTER TABLE class_diary_daily_reports "
            "ADD COLUMN summary_text MEDIUMTEXT DEFAULT NULL "
            "COMMENT '摘要版報告（不含學生姓名）' AFTER report_text"
        )
        logger.info("已新增 daily_reports.summary_text 欄位")
    except Exception:
        pass

    # 日期範圍 AI 報告表
    try:
        pool.execute_write("""
            CREATE TABLE IF NOT EXISTS class_diary_range_reports (
                id              INT AUTO_INCREMENT PRIMARY KEY,
                start_date      DATE NOT NULL              COMMENT '開始日期',
                end_date        DATE NOT NULL              COMMENT '結束日期',
                report_text     MEDIUMTEXT                 COMMENT '完整版報告',
                summary_text    MEDIUMTEXT                 COMMENT '摘要版報告',
                anomalies_json  MEDIUMTEXT                 COMMENT '異常記錄 JSON',
                status          VARCHAR(20) DEFAULT 'pending' COMMENT 'pending/generating/done/failed',
                requested_by    VARCHAR(100)               COMMENT '請求生成者',
                created_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE INDEX uq_date_range (start_date, end_date)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        logger.info("已創建 class_diary_range_reports 表")
    except Exception:
        pass

    # 權限模型表（role + scope）
    try:
        pool.execute_write("""
            CREATE TABLE IF NOT EXISTS class_diary_permissions (
                id          INT AUTO_INCREMENT PRIMARY KEY,
                username    VARCHAR(100) NOT NULL,
                role        VARCHAR(30)  NOT NULL  COMMENT 'reviewer|class_teacher|report_recipient',
                scope_json  TEXT                   COMMENT '{"classes":["S1A"],"grades":["S1"]}',
                granted_by  VARCHAR(100) NOT NULL,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE INDEX uq_user_role (username, role)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        logger.info("已創建 class_diary_permissions 表")
    except Exception:
        pass

    # 審計日誌表
    try:
        pool.execute_write("""
            CREATE TABLE IF NOT EXISTS class_diary_audit_log (
                id            BIGINT AUTO_INCREMENT PRIMARY KEY,
                action        VARCHAR(50)  NOT NULL   COMMENT 'CREATE|UPDATE|DELETE|GRANT_*|REVOKE_*|GENERATE_REPORT|EXPORT',
                target_type   VARCHAR(50)  NOT NULL   COMMENT 'entry|reviewer|recipient|daily_report|range_report|class',
                target_id     VARCHAR(100)            COMMENT '目標 ID',
                actor         VARCHAR(100) NOT NULL   COMMENT '操作人',
                old_value     MEDIUMTEXT              COMMENT '修改前 JSON',
                new_value     MEDIUMTEXT              COMMENT '修改後 JSON',
                metadata_json TEXT                    COMMENT '額外元數據',
                created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                INDEX idx_actor (actor),
                INDEX idx_target (target_type, target_id),
                INDEX idx_created (created_at)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """)
        logger.info("已創建 class_diary_audit_log 表")
    except Exception:
        pass

    # 報告表新增 findings_json 欄位（AI 證據鏈）
    for tbl in ("class_diary_daily_reports", "class_diary_range_reports"):
        try:
            pool.execute_write(
                f"ALTER TABLE {tbl} "
                "ADD COLUMN findings_json MEDIUMTEXT DEFAULT NULL "
                "COMMENT '結構化發現 JSON' AFTER anomalies_json"
            )
            logger.info("已新增 %s.findings_json 欄位", tbl)
        except Exception:
            pass  # 欄位已存在

    logger.info("課室日誌數據表初始化完成")


# ====================================================================== #
#  公開端點                                                                #
# ====================================================================== #

@router.get("/api/class-diary/teacher/recent-subjects")
async def get_recent_subjects(
    request: Request,
    limit: int = Query(10, ge=1, le=30),
):
    """獲取當前教師最近使用的科目（需教師/管理員登入）"""
    try:
        username, role = _verify_request(request)
        if role == "student":
            raise AuthorizationError("僅限教師使用")
        service = _get_service()
        subjects = service._entry_repo.get_recent_subjects(username, limit)
        return success_response(subjects)
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/reason-codes")
async def get_reason_codes():
    """返回行為原因代碼（公開，供前端動態載入）"""
    from app.domains.class_diary.constants import REASON_CODES
    return success_response(REASON_CODES)


@router.post("/api/class-diary/entries")
async def create_entry(request: Request):
    """提交課堂評級（需教師/管理員登入）"""
    try:
        username, role = _verify_request(request)
        if role == "student":
            raise AppException(
                code="FORBIDDEN",
                message="僅限教師或管理員提交評級",
                status_code=403,
            )

        body = await request.json()
        entry_data = CreateEntryRequest(**body)

        user_agent = request.headers.get("User-Agent", "")

        service = _get_service()
        result = service.create_entry(entry_data.model_dump(), user_agent, submitted_by=username)

        _audit.log("CREATE", "entry", username, target_id=result.get("id"),
                   new_value={"class_code": entry_data.class_code, "period": f"{entry_data.period_start}-{entry_data.period_end}", "subject": entry_data.subject})

        return success_response(result, "評級提交成功")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("提交評級失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/entries/by-class")
async def get_entries_by_class(
    class_code: str = Query(..., description="班級代碼"),
    entry_date: str = Query(..., description="日期 YYYY-MM-DD"),
):
    """查詢某班某日的評級記錄（公開，用於教師掃碼後查看已填記錄）"""
    try:
        service = _get_service()
        entries = service.get_entries_by_class_date(class_code, entry_date)

        # 公開端點不返回簽名數據（節省帶寬）
        for entry in entries:
            if "signature" in entry:
                entry["signature"] = bool(entry.get("signature"))

        return success_response(entries)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("查詢記錄失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/classes")
async def get_classes_public():
    """公開端點 — 取得班級列表（供移動端表單選擇班級）"""
    try:
        service = _get_service()
        classes = service.get_all_classes()
        return success_response(classes)
    except Exception as e:
        logger.exception("取得班級列表失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/students/{class_code}")
async def get_students_for_class(class_code: str):
    """公開端點 — 取得該班學生列表（供移動端表單選擇學生）"""
    try:
        service = _get_service()
        students = service.get_students_for_class(class_code)
        return success_response(students)
    except Exception as e:
        logger.exception("取得學生列表失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.put("/api/class-diary/entries/{entry_id}")
async def update_entry(entry_id: int, request: Request):
    """修改已提交的評級記錄（需教師/管理員登入）"""
    try:
        username, role = _verify_request(request)
        if role == "student":
            raise AppException(
                code="FORBIDDEN",
                message="僅限教師或管理員修改記錄",
                status_code=403,
            )

        service = _get_service()
        existing = service.get_entry(entry_id)

        # 權限檢查：admin 可改任何，教師只能改自己的
        if role != "admin":
            if not existing.get("submitted_by"):
                # 舊記錄（submitted_by IS NULL）只有 admin 可修改
                raise AuthorizationError("舊記錄僅限管理員修改")
            if existing["submitted_by"] != username:
                raise AuthorizationError("只能修改自己提交的記錄")

        body = await request.json()
        update_data = UpdateEntryRequest(**body)

        update_dict = update_data.model_dump(exclude_none=True)
        result = service.update_entry(entry_id, update_dict)

        _audit.log("UPDATE", "entry", username, target_id=entry_id,
                   old_value={"subject": existing.get("subject"), "discipline": existing.get("discipline_rating")},
                   new_value=update_dict)

        return success_response(result, "記錄更新成功")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("更新記錄失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.delete("/api/class-diary/entries/{entry_id}")
async def delete_entry_by_teacher(entry_id: int, request: Request):
    """刪除記錄（教師只能刪自己的，管理員可刪任何）"""
    try:
        username, role = _verify_request(request)
        if role == "student":
            raise AppException(
                code="FORBIDDEN",
                message="僅限教師或管理員刪除記錄",
                status_code=403,
            )

        service = _get_service()
        existing = service.get_entry(entry_id)

        # 權限檢查
        if role != "admin":
            if not existing.get("submitted_by"):
                raise AuthorizationError("舊記錄僅限管理員刪除")
            if existing["submitted_by"] != username:
                raise AuthorizationError("只能刪除自己提交的記錄")

        service.delete_entry(entry_id)
        _audit.log("DELETE", "entry", username, target_id=entry_id,
                   old_value={"class_code": existing.get("class_code"), "date": str(existing.get("entry_date")),
                              "period": f"{existing.get('period_start', 0)}-{existing.get('period_end', 0)}",
                              "subject": existing.get("subject")})
        return success_response(None, "記錄已刪除")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("刪除記錄失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  Review 端點（需登入 + reviewer 權限）                                    #
# ====================================================================== #

@router.get("/api/class-diary/review")
async def review_entries(
    request: Request,
    class_code: Optional[str] = Query(None, description="班級代碼"),
    entry_date: Optional[str] = Query(None, description="日期 YYYY-MM-DD"),
):
    """查看課堂評級記錄（需 reviewer 或 admin 權限）"""
    try:
        _verify_reviewer(request)

        service = _get_service()

        if not entry_date:
            entry_date = str(date.today())

        if class_code:
            entries = service.get_entries_by_class_date(class_code, entry_date)
        else:
            entries = service.get_entries_by_date(entry_date)

        return success_response(entries)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("Review 查詢失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/review/classes")
async def review_classes(
    request: Request,
    entry_date: Optional[str] = Query(None, description="日期 YYYY-MM-DD"),
):
    """獲取某日有記錄的班級列表"""
    try:
        _verify_reviewer(request)

        service = _get_service()
        if not entry_date:
            entry_date = str(date.today())

        classes = service.get_classes_with_entries(entry_date)
        return success_response(classes)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("查詢班級列表失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/review/summary")
async def review_summary(
    request: Request,
    class_code: str = Query(..., description="班級代碼"),
    entry_date: str = Query(..., description="日期 YYYY-MM-DD"),
):
    """某班某日的匯總數據"""
    try:
        _verify_reviewer(request)

        service = _get_service()
        summary = service.get_summary(class_code, entry_date)
        return success_response(summary)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("查詢匯總失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/review/check-access")
async def check_review_access(request: Request):
    """檢查當前用戶的 review 權限等級"""
    try:
        username, role = _verify_request(request)
        service = _get_service()
        tier_info = service.get_user_permission_tier(username, role)
        return success_response(tier_info)
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/review/submission-gaps")
async def get_submission_gaps(
    request: Request,
    entry_date: Optional[str] = Query(None, description="日期 YYYY-MM-DD"),
):
    """獲取某日各班提交缺口（reviewer/admin 限定）"""
    try:
        username, role = _verify_request(request)
        service = _get_service()
        tier = service.get_user_permission_tier(username, role)

        if tier["tier"] not in ("admin", "reviewer"):
            raise AuthorizationError("僅限 reviewer 或管理員查看提交狀況")

        if not entry_date:
            entry_date = str(date.today())

        result = service.get_submission_gaps(entry_date)
        return success_response(result)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("獲取提交缺口失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/review/dashboard-data")
async def get_dashboard_data(
    request: Request,
    mode: str = Query(..., description="single_day 或 date_range"),
    entry_date: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
):
    """統一數據 API — 按 tier 返回聚合數據"""
    try:
        username, role = _verify_request(request)
        service = _get_service()
        tier = service.get_user_permission_tier(username, role)

        if not tier["has_access"]:
            from app.domains.class_diary.exceptions import ReviewAccessDeniedError
            raise ReviewAccessDeniedError()

        # ---------- 取得 entries ----------
        if mode == "single_day":
            if not entry_date:
                entry_date = str(date.today())
            entries = service.get_entries_by_date(entry_date)
        elif mode == "date_range":
            if not start_date or not end_date:
                return error_response(
                    "INVALID_PARAMS", "date_range 模式需要 start_date 和 end_date",
                    status_code=400,
                )
            if start_date > end_date:
                return error_response(
                    "INVALID_PARAMS", "start_date 不能晚於 end_date",
                    status_code=400,
                )
            entries = service.get_all_entries_by_date_range(start_date, end_date)
        else:
            return error_response(
                "INVALID_PARAMS", "mode 必須為 single_day 或 date_range",
                status_code=400,
            )

        # ---------- 班主任：過濾 entries ----------
        if tier["tier"] == "class_teacher" and tier["own_classes"]:
            own_set = set(tier["own_classes"])
            entries = [e for e in entries if e.get("class_code") in own_set]

        # ---------- 聚合 ----------
        if mode == "single_day":
            agg = service.aggregate_single_day(entries)
        else:
            agg = service.aggregate_date_range(entries)

        # ---------- 構建穩定 response ----------
        overview = agg.get("overview", {})
        charts = {}
        tables = {}

        if tier["tier"] != "report_recipient":
            # 有圖表權限的 tier
            charts["class_stats"] = agg.get("class_stats", [])
            if mode == "date_range":
                # 日期序列化
                daily_summary = agg.get("daily_summary", [])
                for item in daily_summary:
                    if "date" in item and not isinstance(item["date"], str):
                        item["date"] = str(item["date"])
                charts["daily_summary"] = daily_summary
                charts["period_analysis"] = agg.get("period_analysis", [])
                charts["subject_analysis"] = agg.get("subject_analysis", [])
                charts["weekday_analysis"] = agg.get("weekday_analysis", [])
                charts["submitter_analysis"] = agg.get("submitter_analysis", [])
                charts["reason_analysis"] = agg.get("reason_analysis", [])

            # 有原始數據權限的 tier
            if tier["can_view_raw_data"]:
                tables["records"] = agg.get("records", [])
                if mode == "date_range":
                    tables["student_records"] = agg.get("student_records", [])
                    tables["risk_students"] = agg.get("risk_students", [])

        return success_response({
            "permission": tier,
            "mode": mode,
            "overview": overview,
            "charts": charts,
            "tables": tables,
        })

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("查詢儀表板數據失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  管理員端點                                                              #
# ====================================================================== #

@router.get("/api/class-diary/admin/classes")
async def admin_list_classes(request: Request):
    """列出所有班級（管理員用，含班主任資訊）"""
    try:
        _verify_admin(request)

        from app.infrastructure.database import get_database_pool
        pool = get_database_pool()

        classes = pool.execute(
            "SELECT class_code, class_name, grade, teacher_username, vice_teacher_username "
            "FROM classes ORDER BY class_code"
        )
        return success_response(classes or [])

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("查詢班級失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/admin/qr/batch")
async def batch_qr(request: Request):
    """批量生成所有班級 QR 碼 ZIP（必須在 {class_code} 路由之前，否則 'batch' 被當作 class_code）"""
    try:
        _verify_admin(request)

        service = _get_service()
        classes = service.get_all_classes()
        class_codes = [c["class_code"] for c in classes]

        if not class_codes:
            return error_response("NO_CLASSES", "沒有班級記錄", status_code=404)

        base_url = _get_base_url(request)
        zip_data = service.generate_qr_codes_zip(class_codes, base_url)

        return Response(
            content=zip_data,
            media_type="application/zip",
            headers={
                "Content-Disposition": 'attachment; filename="class_diary_qrcodes.zip"'
            },
        )

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("批量生成 QR 碼失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/admin/qr/{class_code}")
async def generate_qr(class_code: str, request: Request):
    """為指定班級生成 QR 碼 PNG"""
    try:
        _verify_admin(request)

        service = _get_service()
        base_url = _get_base_url(request)
        png_data = service.generate_qr_code(class_code, base_url)

        return Response(
            content=png_data,
            media_type="image/png",
            headers={"Content-Disposition": f'inline; filename="QR_{class_code}.png"'},
        )

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("生成 QR 碼失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ---------- Reviewer 管理 ---------- #

@router.get("/api/class-diary/admin/reviewers")
async def list_reviewers(request: Request):
    """列出所有 reviewer"""
    try:
        _verify_admin(request)
        service = _get_service()
        reviewers = service.get_all_reviewers()
        return success_response(reviewers)

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/class-diary/admin/reviewers")
async def add_reviewer(request: Request):
    """添加 reviewer"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        data = ReviewerRequest(**body)

        service = _get_service()
        added = service.add_reviewer(data.username, admin_user)

        if not added:
            return error_response("ALREADY_EXISTS", "該用戶已是 Reviewer", status_code=409)

        _audit.log("GRANT_REVIEWER", "reviewer", admin_user, target_id=data.username)
        return success_response({"username": data.username}, "Reviewer 添加成功")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.delete("/api/class-diary/admin/reviewers/{username}")
async def remove_reviewer(username: str, request: Request):
    """移除 reviewer"""
    try:
        admin_user, _ = _verify_admin(request)
        service = _get_service()
        removed = service.remove_reviewer(username)

        if not removed:
            return error_response("NOT_FOUND", "該用戶不是 Reviewer", status_code=404)

        _audit.log("REVOKE_REVIEWER", "reviewer", admin_user, target_id=username)
        return success_response(None, "Reviewer 已移除")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.delete("/api/class-diary/admin/entries/{entry_id}")
async def admin_delete_entry(entry_id: int, request: Request):
    """管理員刪除記錄"""
    try:
        admin_user, _ = _verify_admin(request)
        service = _get_service()
        existing = service.get_entry(entry_id)
        service.delete_entry(entry_id)
        _audit.log("DELETE", "entry", admin_user, target_id=entry_id,
                   old_value={"class_code": existing.get("class_code"), "date": str(existing.get("entry_date")),
                              "subject": existing.get("subject")})
        return success_response(None, "記錄已刪除")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ---------- 班級管理 ---------- #

@router.post("/api/class-diary/admin/classes")
async def create_class(request: Request):
    """創建班級"""
    try:
        _verify_admin(request)
        body = await request.json()

        class_code = (body.get("class_code") or "").strip()
        class_name = (body.get("class_name") or "").strip()
        grade = (body.get("grade") or "").strip()

        if not class_code:
            return error_response("VALIDATION_ERROR", "班級代碼不可為空", status_code=400)
        if not class_name:
            class_name = class_code  # 默認使用代碼作為名稱

        from app.infrastructure.database import get_database_pool
        pool = get_database_pool()

        # 檢查是否已存在
        existing = pool.execute(
            "SELECT id FROM classes WHERE class_code = %s", (class_code,)
        )
        if existing:
            return error_response("ALREADY_EXISTS", f"班級 {class_code} 已存在", status_code=409)

        pool.execute(
            "INSERT INTO classes (class_code, class_name, grade) VALUES (%s, %s, %s)",
            (class_code, class_name, grade),
        )

        return success_response(
            {"class_code": class_code, "class_name": class_name, "grade": grade},
            "班級創建成功",
        )

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("創建班級失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.put("/api/class-diary/admin/classes/{class_code}/teachers")
async def update_class_teachers(class_code: str, request: Request):
    """設定班級的班主任和副班主任"""
    try:
        _verify_admin(request)
        body = await request.json()

        teacher_username = (body.get("teacher_username") or "").strip() or None
        vice_teacher_username = (body.get("vice_teacher_username") or "").strip() or None

        from app.infrastructure.database import get_database_pool
        pool = get_database_pool()

        pool.execute(
            "UPDATE classes SET teacher_username = %s, vice_teacher_username = %s WHERE class_code = %s",
            (teacher_username, vice_teacher_username, class_code),
        )

        return success_response(
            {"class_code": class_code, "teacher_username": teacher_username, "vice_teacher_username": vice_teacher_username},
            "班主任設定已更新",
        )

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("更新班主任失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.delete("/api/class-diary/admin/classes/{class_code}")
async def delete_class(class_code: str, request: Request):
    """刪除班級"""
    try:
        _verify_admin(request)

        from app.infrastructure.database import get_database_pool
        pool = get_database_pool()

        result = pool.execute(
            "DELETE FROM classes WHERE class_code = %s", (class_code,)
        )

        return success_response(None, "班級已刪除")

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  每日報告 — 報告接收人管理                                                 #
# ====================================================================== #

@router.get("/api/class-diary/admin/report-recipients")
async def list_report_recipients(request: Request):
    """列出所有報告接收人"""
    try:
        _verify_admin(request)
        pool = get_database_pool()
        rows = pool.execute(
            "SELECT username, granted_by, created_at FROM class_diary_report_recipients ORDER BY created_at"
        )
        return success_response(rows or [])
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/class-diary/admin/report-recipients")
async def add_report_recipient(request: Request):
    """添加報告接收人"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        username = (body.get("username") or "").strip()
        if not username:
            return error_response("VALIDATION_ERROR", "用戶名不可為空", status_code=400)

        pool = get_database_pool()
        existing = pool.execute(
            "SELECT id FROM class_diary_report_recipients WHERE username = %s", (username,)
        )
        if existing:
            return error_response("ALREADY_EXISTS", "該用戶已是報告接收人", status_code=409)

        pool.execute_write(
            "INSERT INTO class_diary_report_recipients (username, granted_by) VALUES (%s, %s)",
            (username, admin_user),
        )
        _audit.log("GRANT_RECIPIENT", "recipient", admin_user, target_id=username)
        return success_response({"username": username}, "報告接收人添加成功")
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.delete("/api/class-diary/admin/report-recipients/{username}")
async def remove_report_recipient(username: str, request: Request):
    """移除報告接收人"""
    try:
        admin_user, _ = _verify_admin(request)
        pool = get_database_pool()
        pool.execute_write(
            "DELETE FROM class_diary_report_recipients WHERE username = %s", (username,)
        )
        _audit.log("REVOKE_RECIPIENT", "recipient", admin_user, target_id=username)
        return success_response(None, "報告接收人已移除")
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ---------- 審計日誌查詢 ---------- #

@router.get("/api/class-diary/admin/audit-log")
async def get_audit_log(
    request: Request,
    actor: Optional[str] = Query(None),
    target_type: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    """查詢審計日誌（僅管理員）"""
    try:
        _verify_admin(request)
        rows = _audit.query(actor=actor, target_type=target_type, action=action, limit=limit, offset=offset)
        return success_response(rows)
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ---------- 權限管理 (role + scope) ---------- #

@router.get("/api/class-diary/admin/permissions")
async def list_permissions(
    request: Request,
    username: Optional[str] = Query(None),
):
    """列出所有權限記錄（僅管理員）"""
    try:
        admin_user = _verify_admin(request)
        pool = get_database_pool()
        if username:
            rows = pool.execute_query(
                "SELECT id, username, role, scope_json, granted_by, created_at "
                "FROM class_diary_permissions WHERE username = %s ORDER BY created_at DESC",
                (username,),
            )
        else:
            rows = pool.execute_query(
                "SELECT id, username, role, scope_json, granted_by, created_at "
                "FROM class_diary_permissions ORDER BY username, role",
            )
        # 解析 scope_json
        for r in (rows or []):
            if r.get("scope_json"):
                try:
                    r["scope"] = json.loads(r["scope_json"])
                except (json.JSONDecodeError, TypeError):
                    r["scope"] = None
            else:
                r["scope"] = None
            if r.get("created_at"):
                r["created_at"] = str(r["created_at"])
        return success_response(rows or [])
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.post("/api/class-diary/admin/permissions")
async def grant_permission(request: Request):
    """授予權限（僅管理員）

    Body: {"username": "...", "role": "reviewer|class_teacher|report_recipient",
           "scope": {"classes": ["S1A"], "grades": ["S1"]}}
    scope 可省略表示無範圍限制。
    """
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        target = body.get("username", "").strip()
        role = body.get("role", "").strip()
        scope = body.get("scope")

        if not target:
            raise AppException(code="INVALID_INPUT", message="缺少 username", status_code=400)
        valid_roles = ("reviewer", "class_teacher", "report_recipient")
        if role not in valid_roles:
            raise AppException(
                code="INVALID_INPUT",
                message=f"role 必須是 {', '.join(valid_roles)} 之一",
                status_code=400,
            )

        scope_json = json.dumps(scope, ensure_ascii=False) if scope else None
        pool = get_database_pool()
        pool.execute_write(
            "INSERT INTO class_diary_permissions (username, role, scope_json, granted_by) "
            "VALUES (%s, %s, %s, %s) "
            "ON DUPLICATE KEY UPDATE scope_json = VALUES(scope_json), granted_by = VALUES(granted_by)",
            (target, role, scope_json, admin_user),
        )
        _audit.log(
            "GRANT_PERMISSION", "permission", admin_user,
            target_id=f"{target}:{role}",
            new_value=json.dumps({"role": role, "scope": scope}, ensure_ascii=False),
        )
        return success_response(None, f"已授予 {target} 角色 {role}")
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.delete("/api/class-diary/admin/permissions/{perm_id}")
async def revoke_permission(request: Request, perm_id: int):
    """撤銷權限（僅管理員）"""
    try:
        admin_user, _ = _verify_admin(request)
        pool = get_database_pool()
        # 先查出舊記錄供審計
        rows = pool.execute_query(
            "SELECT id, username, role, scope_json FROM class_diary_permissions WHERE id = %s",
            (perm_id,),
        )
        if not rows:
            raise AppException(code="NOT_FOUND", message="權限記錄不存在", status_code=404)
        old = rows[0]
        pool.execute_write(
            "DELETE FROM class_diary_permissions WHERE id = %s", (perm_id,),
        )
        _audit.log(
            "REVOKE_PERMISSION", "permission", admin_user,
            target_id=f"{old['username']}:{old['role']}",
            old_value=json.dumps(
                {"role": old["role"], "scope_json": old.get("scope_json")},
                ensure_ascii=False,
            ),
        )
        return success_response(None, f"已撤銷 {old['username']} 的 {old['role']} 權限")
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  每日報告 — 生成 & 查詢                                                  #
# ====================================================================== #

_REPORT_SYSTEM_PROMPT = """你是一位學校課室日誌分析助手。根據提供的課堂記錄數據，生成一份每日報告。

你必須生成兩個版本，用分隔線分開：

===完整版===
（以下為完整版內容，供管理層和 Reviewer 查看）
1. 先給出整體概況（共幾個班有記錄，平均紀律/整潔分數）
2. 逐班列出：班級名、記錄數、平均分、重點問題
3. 重點標出「異常」：紀律或整潔 ≤ 2 分的課堂、缺席人數多的課堂、有違規記錄的課堂
4. 可以包含具體學生姓名和班級問題細節
5. 最後給出簡短的總結建議

===摘要版===
（以下為摘要版內容，供報告接收人查看）
1. 只提供整體概況和趨勢描述
2. 絕對不能包含任何學生姓名
3. 不列出具體班級問題細節
4. 只描述整體表現、平均分數、主要趨勢和建議
5. 篇幅較短（3-5 段即可）

通用要求：
- 使用繁體中文，語氣正式但簡潔
- 不要使用 Markdown 標題語法，使用【】來標記章節

===FINDINGS_JSON===
最後，請輸出一段 JSON 陣列（不含其他文字），列出你的關鍵發現。格式：
[{"finding_type":"anomaly|trend|praise","severity":"high|medium|low","description":"簡短描述","evidence":[{"entry_id":<ID>,"date":"YYYY-MM-DD","class_code":"...","field":"discipline_rating|cleanliness_rating|absent_students|rule_violations|commended_students","value":"..."}],"recommendation":"建議"}]
如果沒有明顯發現，輸出空陣列 []。"""

_RANGE_REPORT_SYSTEM_PROMPT = """你是一位學校課室日誌分析助手。根據提供的多日匯總數據，生成一份日期範圍分析報告。

你必須生成兩個版本，用分隔線分開：

===完整版===
（以下為完整版內容，供管理層和 Reviewer 查看）
1. 整體概況：日期範圍、總記錄數、涉及班級數、平均紀律/整潔分數
2. 趨勢分析：紀律/整潔分數的變化趨勢
3. 重點班級：表現最好和最差的班級
4. 高風險學生：列出需要關注的學生及其記名詳情
5. 問題節數/科目：哪些時段或科目問題最多
6. 行為分析：最常見的違規和表揚原因
7. 總結建議：改進方向和建議

===摘要版===
（以下為摘要版內容，供報告接收人查看）
1. 只提供整體概況和趨勢描述
2. 絕對不能包含任何學生姓名
3. 不列出具體班級問題細節
4. 只描述整體表現趨勢、主要發現和建議
5. 篇幅較短（3-5 段即可）

通用要求：
- 使用繁體中文，語氣正式但簡潔
- 不要使用 Markdown 標題語法，使用【】來標記章節

===FINDINGS_JSON===
最後，請輸出一段 JSON 陣列（不含其他文字），列出你的關鍵發現。格式：
[{"finding_type":"anomaly|trend|praise","severity":"high|medium|low","description":"簡短描述","evidence":[{"date":"YYYY-MM-DD","class_code":"...","field":"discipline_rating|cleanliness_rating|absent_students|rule_violations|commended_students","value":"..."}],"recommendation":"建議"}]
如果沒有明顯發現，輸出空陣列 []。"""


def _detect_anomalies(entries: list) -> list:
    """規則檢測異常記錄"""
    anomalies = []
    for e in entries:
        reasons = []
        if e.get("discipline_rating") and int(e["discipline_rating"]) <= 2:
            reasons.append(f"紀律評分低 ({e['discipline_rating']}/5)")
        if e.get("cleanliness_rating") and int(e["cleanliness_rating"]) <= 2:
            reasons.append(f"整潔評分低 ({e['cleanliness_rating']}/5)")

        absent = (e.get("absent_students") or "").strip()
        if absent:
            count = len([s for s in absent.split(",") if s.strip()])
            if count > 3:
                reasons.append(f"缺席人數多 ({count}人)")

        violations_raw = (e.get("rule_violations") or "").strip()
        if violations_raw:
            reasons.append(f"課堂違規: {_format_behavior_field(violations_raw)[:80]}")
        appearance_raw = (e.get("appearance_issues") or "").strip()
        if appearance_raw:
            reasons.append(f"儀表問題: {_format_behavior_field(appearance_raw)[:80]}")
        medical_raw = (e.get("medical_room_students") or "").strip()
        if medical_raw:
            reasons.append(f"醫務室: {_format_behavior_field(medical_raw)[:80]}")

        if reasons:
            anomalies.append({
                "entry_id": e.get("id"),
                "class_code": e.get("class_code", ""),
                "period_start": e.get("period_start"),
                "period_end": e.get("period_end"),
                "subject": e.get("subject", ""),
                "reasons": reasons,
            })
    return anomalies


def _build_report_prompt(entries: list, anomalies: list, report_date: str) -> str:
    """把當天所有記錄組織成給 AI 的提問文本"""
    if not entries:
        return f"日期：{report_date}\n今天沒有任何課堂記錄。請簡單說明沒有數據。"

    # 按班級分組
    by_class = {}
    for e in entries:
        cc = e.get("class_code", "unknown")
        by_class.setdefault(cc, []).append(e)

    lines = [f"日期：{report_date}", f"共 {len(by_class)} 個班有記錄，合計 {len(entries)} 條。", ""]

    for cc, class_entries in sorted(by_class.items()):
        disc_avg = sum(int(e.get("discipline_rating") or 0) for e in class_entries) / len(class_entries)
        clean_avg = sum(int(e.get("cleanliness_rating") or 0) for e in class_entries) / len(class_entries)
        lines.append(f"【班級 {cc}】 {len(class_entries)} 條記錄, 平均紀律 {disc_avg:.1f}, 平均整潔 {clean_avg:.1f}")

        for e in class_entries:
            absent = (e.get("absent_students") or "").strip()
            late = (e.get("late_students") or "").strip()
            violations = _format_behavior_field((e.get("rule_violations") or "").strip())
            appearance = _format_behavior_field((e.get("appearance_issues") or "").strip())
            commended = _format_behavior_field((e.get("commended_students") or "").strip())
            medical = _format_behavior_field((e.get("medical_room_students") or "").strip())

            eid = e.get('id', '?')
            detail = f"  [ID:{eid}] 節{e.get('period_start', '?')}-{e.get('period_end', '?')} {e.get('subject', '')} 紀律{e.get('discipline_rating', '?')} 整潔{e.get('cleanliness_rating', '?')}"
            if absent:
                detail += f" 缺席:{absent}"
            if late:
                detail += f" 遲到:{late}"
            if violations:
                detail += f" 違規:{violations}"
            if appearance:
                detail += f" 儀表:{appearance}"
            if commended:
                detail += f" 嘉許:{commended}"
            if medical:
                detail += f" 醫務室:{medical}"
            lines.append(detail)
        lines.append("")

    if anomalies:
        lines.append(f"===== 異常記錄（共 {len(anomalies)} 條） =====")
        for a in anomalies:
            lines.append(f"  {a['class_code']} 節{a.get('period_start', '?')}-{a.get('period_end', '?')} {a['subject']}: {', '.join(a['reasons'])}")

    return "\n".join(lines)


def _split_report_versions(ai_output: str):
    """將 AI 輸出拆分為完整版、摘要版和 findings JSON"""
    import re as _re
    full_text = ai_output
    summary_text = ""
    findings_json = None

    # 先提取 findings JSON（在任何版本分割之前）
    if "===FINDINGS_JSON===" in ai_output:
        parts = ai_output.split("===FINDINGS_JSON===", 1)
        ai_output = parts[0].strip()
        raw_findings = parts[1].strip()
        try:
            # 嘗試解析 JSON — AI 可能前後帶有多餘文字
            match = _re.search(r'\[.*\]', raw_findings, _re.DOTALL)
            if match:
                findings_json = json.loads(match.group(0))
        except (json.JSONDecodeError, TypeError):
            findings_json = None
    else:
        # 回退：嘗試偵測文本末尾的 JSON 陣列（AI 未使用分隔符時）
        match = _re.search(r'\n\s*(\[\s*\{[\s\S]*\}[\s\n]*\])\s*$', ai_output)
        if match:
            try:
                parsed = json.loads(match.group(1))
                if isinstance(parsed, list) and parsed and isinstance(parsed[0], dict):
                    findings_json = parsed
                    ai_output = ai_output[:match.start()].strip()
            except (json.JSONDecodeError, TypeError):
                pass

    if "===摘要版===" in ai_output:
        parts = ai_output.split("===摘要版===", 1)
        full_text = parts[0].replace("===完整版===", "").strip()
        summary_text = parts[1].strip()
    elif "===完整版===" in ai_output:
        full_text = ai_output.replace("===完整版===", "").strip()
    else:
        full_text = ai_output.strip()

    return full_text, summary_text, findings_json


def _generate_report_sync(report_date: str):
    """同步版本的報告生成（供 run_in_executor 調用）"""
    pool = get_database_pool()

    pool.execute_write(
        "INSERT INTO class_diary_daily_reports (report_date, report_text, status) "
        "VALUES (%s, '', 'generating') "
        "ON DUPLICATE KEY UPDATE status = 'generating', updated_at = NOW()",
        (report_date,),
    )

    try:
        service = _get_service()
        entries = service.get_entries_by_date(report_date)
        anomalies = _detect_anomalies(entries)
        prompt_text = _build_report_prompt(entries, anomalies, report_date)

        findings = None
        if not entries:
            full_text = f"{report_date} 今天沒有任何課堂記錄。"
            summary_text = full_text
        else:
            from llm.services.qa_service import ask_ai_local
            ai_output, _ = ask_ai_local(
                question=prompt_text,
                subject="general",
                system_prompt=_REPORT_SYSTEM_PROMPT,
                task_type="summary",
            )
            full_text, summary_text, findings = _split_report_versions(ai_output)

        anomalies_json = json.dumps(anomalies, ensure_ascii=False)
        findings_str = json.dumps(findings, ensure_ascii=False) if findings else None
        pool.execute_write(
            "UPDATE class_diary_daily_reports "
            "SET report_text = %s, summary_text = %s, anomalies_json = %s, "
            "findings_json = %s, status = 'done', updated_at = NOW() "
            "WHERE report_date = %s",
            (full_text, summary_text, anomalies_json, findings_str, report_date),
        )
        logger.info("每日報告生成完成: %s (%d 條記錄, %d 條異常, %d findings)",
                     report_date, len(entries), len(anomalies), len(findings or []))
    except Exception as e:
        logger.exception("每日報告生成失敗: %s", report_date)
        pool.execute_write(
            "UPDATE class_diary_daily_reports SET status = 'failed', report_text = %s, updated_at = NOW() WHERE report_date = %s",
            (f"生成失敗: {str(e)[:500]}", report_date),
        )
        raise


@router.post("/api/class-diary/admin/generate-report")
async def admin_generate_report(request: Request):
    """管理員手動觸發報告生成（異步，立即返回 task_id）"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        report_date = (body.get("report_date") or str(date.today())).strip()

        def _daily_task():
            _generate_report_sync(report_date)

        task_id = _tasks.submit(
            task_type="daily_report",
            func=_daily_task,
            requested_by=admin_user,
            description=f"每日報告 {report_date}",
        )

        _audit.log("GENERATE_REPORT", "daily_report", admin_user, target_id=report_date)

        # 宠物金币：生成学生报告 +8
        try:
            from app.domains.pet.hooks import try_award_coins_by_username
            try_award_coins_by_username(admin_user, "generate_report", f"report_{report_date}", "teacher")
        except Exception:
            pass

        return success_response(
            {"report_date": report_date, "status": "generating", "task_id": task_id},
            "報告生成已啟動",
        )
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/review/daily-report")
async def get_daily_report(
    request: Request,
    entry_date: Optional[str] = Query(None, description="日期 YYYY-MM-DD"),
):
    """獲取每日 AI 報告（根據 tier 返回完整版或摘要版）"""
    try:
        username, role = _verify_request(request)
        service = _get_service()
        tier = service.get_user_permission_tier(username, role)

        if not tier["has_access"]:
            from app.domains.class_diary.exceptions import ReviewAccessDeniedError
            raise ReviewAccessDeniedError()

        if not entry_date:
            entry_date = str(date.today())

        pool = get_database_pool()

        rows = pool.execute(
            "SELECT report_text, summary_text, anomalies_json, findings_json, status "
            "FROM class_diary_daily_reports WHERE report_date = %s",
            (entry_date,),
        )
        report = rows[0] if rows else None

        is_reviewer = tier["tier"] in ("admin", "reviewer")
        is_recipient = tier["can_view_ai_report"]

        if report:
            anomalies = []
            if report.get("anomalies_json"):
                try:
                    anomalies = json.loads(report["anomalies_json"])
                except (json.JSONDecodeError, TypeError):
                    pass
            findings = []
            if report.get("findings_json"):
                try:
                    findings = json.loads(report["findings_json"])
                except (json.JSONDecodeError, TypeError):
                    pass

            # 根據 tier 返回對應版本
            if tier["tier"] == "report_recipient":
                text = report.get("summary_text") or report["report_text"]
            else:
                text = report["report_text"]

            return success_response({
                "report_text": text,
                "anomalies": anomalies if is_reviewer else [],
                "findings": findings if is_reviewer else [],
                "status": report["status"],
                "is_recipient": is_recipient,
                "is_reviewer": is_reviewer,
            })
        else:
            return success_response({
                "report_text": None,
                "anomalies": [],
                "findings": [],
                "status": "none",
                "is_recipient": is_recipient,
                "is_reviewer": is_reviewer,
            })

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("查詢每日報告失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  日期範圍 AI 報告                                                         #
# ====================================================================== #

def _build_range_report_prompt(agg: dict, start_date: str, end_date: str) -> str:
    """將日期範圍聚合數據組織成給 AI 的提問文本"""
    ov = agg.get("overview", {})
    lines = [
        f"日期範圍：{start_date} 至 {end_date}",
        f"總記錄：{ov.get('total_entries', 0)} 條，"
        f"涉及 {ov.get('total_classes', 0)} 個班級，"
        f"共 {ov.get('total_dates', 0)} 天",
        f"平均紀律：{ov.get('avg_discipline', '-')}，平均整潔：{ov.get('avg_cleanliness', '-')}",
        f"總缺席：{ov.get('total_absent', 0)} 人次，總遲到：{ov.get('total_late', 0)} 人次",
        "",
    ]

    # 每日趨勢
    daily = agg.get("daily_summary", [])
    if daily:
        lines.append("=== 每日趨勢 ===")
        for d in daily[-14:]:  # 最多顯示最近 14 天
            lines.append(
                f"  {d.get('date', '?')}: 記錄{d.get('entry_count', 0)} "
                f"紀律{d.get('avg_discipline', '-')} 整潔{d.get('avg_cleanliness', '-')} "
                f"缺席{d.get('absent_count', 0)} 遲到{d.get('late_count', 0)} "
                f"違規{d.get('violation_count', 0)} 表揚{d.get('praise_count', 0)}"
            )
        lines.append("")

    # 高風險學生
    risk = agg.get("risk_students", [])
    if risk:
        lines.append(f"=== 高風險學生（{len(risk)} 位）===")
        for s in risk[:10]:
            flags = ", ".join(s.get("risk_flags", []))
            lines.append(
                f"  {s.get('name', '?')} ({s.get('class_code', '?')}) "
                f"違規{s.get('violation_count', 0)} 缺席{s.get('absent_count', 0)} "
                f"遲到{s.get('late_count', 0)} 風險:{flags}"
            )
        lines.append("")

    # 原因分析 top 10
    reasons = agg.get("reason_analysis", [])
    if reasons:
        lines.append("=== 主要原因 (Top 10) ===")
        for r in reasons[:10]:
            lines.append(
                f"  {r.get('reason', '?')} ({r.get('category', '?')}) "
                f"{r.get('total_count', 0)}人次 涉及{r.get('student_count', 0)}學生"
            )
        lines.append("")

    return "\n".join(lines)


def _generate_range_report_sync(start_date: str, end_date: str, requested_by: str):
    """同步版本的範圍報告生成"""
    pool = get_database_pool()

    pool.execute_write(
        "INSERT INTO class_diary_range_reports "
        "(start_date, end_date, status, requested_by) "
        "VALUES (%s, %s, 'generating', %s) "
        "ON DUPLICATE KEY UPDATE status = 'generating', requested_by = %s, updated_at = NOW()",
        (start_date, end_date, requested_by, requested_by),
    )

    try:
        service = _get_service()
        entries = service.get_all_entries_by_date_range(start_date, end_date)
        agg = service.aggregate_date_range(entries)
        prompt_text = _build_range_report_prompt(agg, start_date, end_date)

        findings = None
        if not entries:
            full_text = f"{start_date} 至 {end_date} 沒有任何課堂記錄。"
            summary_text = full_text
        else:
            from llm.services.qa_service import ask_ai_local
            ai_output, _ = ask_ai_local(
                question=prompt_text,
                subject="general",
                system_prompt=_RANGE_REPORT_SYSTEM_PROMPT,
                task_type="summary",
            )
            full_text, summary_text, findings = _split_report_versions(ai_output)

        anomalies = _detect_anomalies(entries)
        anomalies_json = json.dumps(anomalies, ensure_ascii=False)
        findings_str = json.dumps(findings, ensure_ascii=False) if findings else None

        pool.execute_write(
            "UPDATE class_diary_range_reports "
            "SET report_text = %s, summary_text = %s, anomalies_json = %s, "
            "findings_json = %s, status = 'done', updated_at = NOW() "
            "WHERE start_date = %s AND end_date = %s",
            (full_text, summary_text, anomalies_json, findings_str, start_date, end_date),
        )
        logger.info("範圍報告生成完成: %s ~ %s (%d 條記錄, %d findings)",
                     start_date, end_date, len(entries), len(findings or []))
    except Exception as e:
        logger.exception("範圍報告生成失敗: %s ~ %s", start_date, end_date)
        pool.execute_write(
            "UPDATE class_diary_range_reports "
            "SET status = 'failed', report_text = %s, updated_at = NOW() "
            "WHERE start_date = %s AND end_date = %s",
            (f"生成失敗: {str(e)[:500]}", start_date, end_date),
        )
        raise


@router.post("/api/class-diary/admin/generate-range-report")
async def admin_generate_range_report(request: Request):
    """管理員觸發日期範圍報告生成"""
    try:
        admin_user, _ = _verify_admin(request)
        body = await request.json()
        start_date = (body.get("start_date") or "").strip()
        end_date = (body.get("end_date") or "").strip()

        if not start_date or not end_date:
            return error_response("VALIDATION_ERROR", "需要 start_date 和 end_date", status_code=400)
        if start_date > end_date:
            return error_response("VALIDATION_ERROR", "start_date 不能晚於 end_date", status_code=400)

        def _range_task():
            _generate_range_report_sync(start_date, end_date, admin_user)

        task_id = _tasks.submit(
            task_type="range_report",
            func=_range_task,
            requested_by=admin_user,
            description=f"範圍報告 {start_date}~{end_date}",
        )

        _audit.log("GENERATE_REPORT", "range_report", admin_user,
                   target_id=f"{start_date}~{end_date}")
        return success_response(
            {"start_date": start_date, "end_date": end_date, "status": "generating", "task_id": task_id},
            "範圍報告生成已啟動",
        )
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/review/range-report")
async def get_range_report(
    request: Request,
    start_date: str = Query(..., description="開始日期"),
    end_date: str = Query(..., description="結束日期"),
):
    """獲取日期範圍 AI 報告（根據 tier 返回完整版或摘要版）"""
    try:
        username, role = _verify_request(request)
        service = _get_service()
        tier = service.get_user_permission_tier(username, role)

        if not tier["has_access"]:
            from app.domains.class_diary.exceptions import ReviewAccessDeniedError
            raise ReviewAccessDeniedError()

        pool = get_database_pool()
        rows = pool.execute(
            "SELECT report_text, summary_text, anomalies_json, findings_json, status "
            "FROM class_diary_range_reports WHERE start_date = %s AND end_date = %s",
            (start_date, end_date),
        )
        report = rows[0] if rows else None

        is_reviewer = tier["tier"] in ("admin", "reviewer")

        if report:
            anomalies = []
            if report.get("anomalies_json"):
                try:
                    anomalies = json.loads(report["anomalies_json"])
                except (json.JSONDecodeError, TypeError):
                    pass
            findings = []
            if report.get("findings_json"):
                try:
                    findings = json.loads(report["findings_json"])
                except (json.JSONDecodeError, TypeError):
                    pass

            if tier["tier"] == "report_recipient":
                text = report.get("summary_text") or report["report_text"]
            else:
                text = report["report_text"]

            return success_response({
                "report_text": text,
                "anomalies": anomalies if is_reviewer else [],
                "findings": findings if is_reviewer else [],
                "status": report["status"],
                "is_reviewer": is_reviewer,
            })
        else:
            return success_response({
                "report_text": None,
                "anomalies": [],
                "findings": [],
                "status": "none",
                "is_reviewer": is_reviewer,
            })

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("查詢範圍報告失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  任務管理                                                                 #
# ====================================================================== #

@router.get("/api/class-diary/tasks/{task_id}")
async def get_task_status(request: Request, task_id: str):
    """查詢單個任務狀態"""
    try:
        _verify_request(request)
        task = _tasks.get_status(task_id)
        if not task:
            return error_response("NOT_FOUND", "任務不存在", status_code=404)
        return success_response(task)
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


@router.get("/api/class-diary/tasks")
async def list_tasks(request: Request, limit: int = Query(20, ge=1, le=100)):
    """列出任務（僅管理員）"""
    try:
        admin_user, _ = _verify_admin(request)
        tasks = _tasks.list_tasks(limit=limit)
        return success_response(tasks)
    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  匯出 XLSX                                                              #
# ====================================================================== #

_PERIOD_LABELS = ["早會", "第一節", "第二節", "第三節", "第四節",
                  "第五節", "第六節", "第七節", "第八節", "第九節"]


def _period_text(ps: int, pe: int) -> str:
    """節數文字"""
    if ps == pe and ps < len(_PERIOD_LABELS):
        return _PERIOD_LABELS[ps]
    return f"{_PERIOD_LABELS[min(ps, len(_PERIOD_LABELS)-1)]}-{_PERIOD_LABELS[min(pe, len(_PERIOD_LABELS)-1)]}"


def _diary_excel_styles():
    """課室日誌 XLSX 共用樣式"""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="C0C0C0")
    return {
        "title_font": Font(bold=True, size=16, name="Microsoft JhengHei"),
        "subtitle_font": Font(bold=True, size=12, color="4472C4", name="Microsoft JhengHei"),
        "header_font": Font(bold=True, color="FFFFFF", size=10, name="Microsoft JhengHei"),
        "header_fill": PatternFill(start_color="006633", end_color="006633", fill_type="solid"),
        "section_fill": PatternFill(start_color="E8F5EC", end_color="E8F5EC", fill_type="solid"),
        "alt_fill": PatternFill(start_color="F5FAF7", end_color="F5FAF7", fill_type="solid"),
        "stat_fill": PatternFill(start_color="D0E5D8", end_color="D0E5D8", fill_type="solid"),
        "good_fill": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "warn_fill": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "danger_fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center"),
        "wrap": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "body_font": Font(size=10, name="Microsoft JhengHei"),
    }


def _apply_header_row(ws, row, headers, s):
    """在指定行寫入表頭"""
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]


def _apply_data_cell(ws, row, col, value, s, center=True, alt=False):
    """寫入普通數據格"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = s["body_font"]
    c.border = s["border"]
    c.alignment = s["center"] if center else s["wrap"]
    if alt:
        c.fill = s["alt_fill"]
    return c


def _rating_fill(value, s):
    """根據評分返回填色"""
    try:
        v = float(value)
        if v <= 2.5:
            return s["danger_fill"]
        if v >= 4.0:
            return s["good_fill"]
    except (ValueError, TypeError):
        pass
    return None


def _set_col_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_single_day_xlsx(agg, entry_date):
    """構建單天 XLSX"""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    s = _diary_excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "概覽"

    overview = agg["overview"]
    if not overview:
        ws["A1"] = f"課室日誌 — {entry_date}"
        ws["A1"].font = s["title_font"]
        ws["A3"] = "指定日期沒有課室日誌記錄"
        ws["A3"].font = s["body_font"]
        return wb

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"課室日誌 — {entry_date}"
    ws["A1"].font = s["title_font"]

    # Stats row
    stat_items = [
        ("總記錄", overview["total_entries"]),
        ("班級數", overview["total_classes"]),
        ("平均紀律", overview["avg_discipline"]),
        ("平均整潔", overview["avg_cleanliness"]),
        ("缺席人次", overview["total_absent"]),
        ("遲到人次", overview["total_late"]),
    ]
    for i, (label, val) in enumerate(stat_items):
        lc = ws.cell(row=3, column=i * 2 + 1, value=label)
        lc.font = Font(bold=True, size=10, name="Microsoft JhengHei")
        lc.fill = s["stat_fill"]
        lc.border = s["border"]
        vc = ws.cell(row=3, column=i * 2 + 2, value=val)
        vc.font = s["body_font"]
        vc.alignment = s["center"]
        vc.border = s["border"]

    # Class stats section
    row = 5
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = "班級統計"
    ws[f"A{row}"].font = s["subtitle_font"]

    row = 6
    class_headers = ["班級", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次"]
    _apply_header_row(ws, row, class_headers, s)

    for cs in agg["class_stats"]:
        row += 1
        vals = [cs["class_code"], cs["entry_count"], cs["avg_discipline"],
                cs["avg_cleanliness"], cs["absent_count"], cs["late_count"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws, row, col, v, s)
            if col == 3:
                f = _rating_fill(v, s)
                if f:
                    c.fill = f
            if col == 4:
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    # Detail section
    row += 2
    ws.merge_cells(f"A{row}:L{row}")
    ws[f"A{row}"] = "詳細記錄"
    ws[f"A{row}"].font = s["subtitle_font"]

    row += 1
    detail_headers = ["班級", "節數", "科目", "紀律", "整潔", "缺席學生",
                      "遲到學生", "嘉許學生", "儀表問題", "課堂違規", "醫務室", "填寫人"]
    _apply_header_row(ws, row, detail_headers, s)
    freeze_row = row + 1

    idx = 0
    for e in agg["records"]:
        row += 1
        alt = idx % 2 == 1
        vals = [
            e.get("class_code", ""),
            _period_text(e.get("period_start", 0), e.get("period_end", 0)),
            e.get("subject", ""),
            e.get("discipline_rating", ""),
            e.get("cleanliness_rating", ""),
            e.get("absent_students", ""),
            e.get("late_students", ""),
            _format_behavior_field(e.get("commended_students", "")),
            _format_behavior_field(e.get("appearance_issues", "")),
            _format_behavior_field(e.get("rule_violations", "")),
            _format_behavior_field(e.get("medical_room_students", "")),
            e.get("submitted_by", ""),
        ]
        for col, v in enumerate(vals, 1):
            center = col <= 5 or col == 12
            c = _apply_data_cell(ws, row, col, v, s, center=center, alt=alt)
            if col in (4, 5):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f
        idx += 1

    _set_col_widths(ws, [10, 12, 12, 6, 6, 20, 20, 25, 25, 25, 20, 10])
    ws.freeze_panes = f"A{freeze_row}"

    return wb


@router.get("/api/class-diary/review/export")
async def export_entries_xlsx(
    request: Request,
    entry_date: Optional[str] = Query(None, description="日期 YYYY-MM-DD"),
):
    """匯出當天所有班級的記錄為 XLSX"""
    from openpyxl.styles import Font
    from urllib.parse import quote

    try:
        export_user, _ = _verify_reviewer(request)

        if not entry_date:
            entry_date = str(date.today())

        service = _get_service()
        entries = service.get_entries_by_date(entry_date)
        agg = service.aggregate_single_day(entries)

        wb = _build_single_day_xlsx(agg, entry_date)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        _audit.log("EXPORT", "entry", export_user, target_id=entry_date, metadata={"type": "single_day"})

        filename = f"課室日誌_{entry_date}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("匯出 XLSX 失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)


# ====================================================================== #
#  日期範圍匯出 XLSX（10 個工作表）                                          #
# ====================================================================== #

def _build_date_range_xlsx(agg, start_date, end_date):
    """構建日期範圍 XLSX — 10 個工作表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    s = _diary_excel_styles()
    wb = Workbook()

    overview = agg["overview"]
    if not overview:
        ws = wb.active
        ws.title = "總覽"
        ws["A1"] = f"課室日誌 — {start_date} 至 {end_date}"
        ws["A1"].font = s["title_font"]
        ws["A3"] = "指定日期範圍內沒有課室日誌記錄"
        ws["A3"].font = s["body_font"]
        return wb

    # ==================== Sheet 1: 總覽 ====================
    ws = wb.active
    ws.title = "總覽"
    ws.merge_cells("A1:K1")
    ws["A1"] = f"課室日誌 — {start_date} 至 {end_date}"
    ws["A1"].font = s["title_font"]

    stat_items = [
        ("總記錄", overview["total_entries"]),
        ("日期數", overview["total_dates"]),
        ("班級數", overview["total_classes"]),
        ("平均紀律", overview["avg_discipline"]),
        ("平均整潔", overview["avg_cleanliness"]),
    ]
    for i, (label, val) in enumerate(stat_items):
        lc = ws.cell(row=3, column=i * 2 + 1, value=label)
        lc.font = Font(bold=True, size=10, name="Microsoft JhengHei")
        lc.fill = s["stat_fill"]
        lc.border = s["border"]
        vc = ws.cell(row=3, column=i * 2 + 2, value=val)
        vc.font = s["body_font"]
        vc.alignment = s["center"]
        vc.border = s["border"]

    # Daily summary table
    row = 5
    ws.merge_cells(f"A{row}:K{row}")
    ws[f"A{row}"] = "每日匯總"
    ws[f"A{row}"].font = s["subtitle_font"]

    row = 6
    daily_headers = ["日期", "記錄數", "平均紀律", "平均整潔", "缺席", "遲到",
                     "違規", "表揚", "紀律變化", "7日均紀律", "7日均整潔"]
    _apply_header_row(ws, row, daily_headers, s)

    for ds in agg["daily_summary"]:
        row += 1
        vals = [ds["date"], ds["entry_count"], ds["avg_discipline"], ds["avg_cleanliness"],
                ds["absent_count"], ds["late_count"], ds["violation_count"], ds["praise_count"],
                ds["disc_diff"] if ds["disc_diff"] is not None else "",
                ds["ma7_discipline"], ds["ma7_cleanliness"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws, row, col, v, s)
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws, [12, 8, 8, 8, 6, 6, 6, 6, 8, 8, 8])
    ws.freeze_panes = "A7"

    # ==================== Sheet 2: 每日記錄 ====================
    ws2 = wb.create_sheet("每日記錄")
    row = 1
    detail_headers = ["日期", "班級", "節數", "科目", "紀律", "整潔", "缺席學生",
                      "遲到學生", "嘉許學生", "儀表問題", "課堂違規", "醫務室", "填寫人"]
    _apply_header_row(ws2, row, detail_headers, s)

    # Group entries by date for separator rows
    from collections import defaultdict as _dd
    by_date = _dd(list)
    for e in agg.get("_raw_entries", []):
        by_date[str(e.get("entry_date", ""))].append(e)

    idx = 0
    for d in sorted(by_date):
        for e in by_date[d]:
            row += 1
            alt = idx % 2 == 1
            vals = [
                str(e.get("entry_date", "")),
                e.get("class_code", ""),
                _period_text(e.get("period_start", 0), e.get("period_end", 0)),
                e.get("subject", ""),
                e.get("discipline_rating", ""),
                e.get("cleanliness_rating", ""),
                e.get("absent_students", ""),
                e.get("late_students", ""),
                _format_behavior_field(e.get("commended_students", "")),
                _format_behavior_field(e.get("appearance_issues", "")),
                _format_behavior_field(e.get("rule_violations", "")),
                _format_behavior_field(e.get("medical_room_students", "")),
                e.get("submitted_by", ""),
            ]
            for col, v in enumerate(vals, 1):
                center = col <= 6 or col == 13
                c = _apply_data_cell(ws2, row, col, v, s, center=center, alt=alt)
                if col in (5, 6):
                    f = _rating_fill(v, s)
                    if f:
                        c.fill = f
            idx += 1

    _set_col_widths(ws2, [12, 10, 12, 12, 6, 6, 20, 20, 25, 25, 25, 20, 10])
    ws2.freeze_panes = "A2"

    # ==================== Sheet 3: 學生記錄 ====================
    ws3 = wb.create_sheet("學生記錄")
    stu_headers = ["學生", "班級", "缺席次數", "缺席日期", "遲到次數", "遲到日期",
                   "違規次數", "違規日期", "醫務室次數", "醫務室日期"]
    _apply_header_row(ws3, 1, stu_headers, s)

    row = 1
    for sr in agg["student_records"]:
        row += 1
        alt = (row % 2) == 0
        vals = [
            sr["name"], sr["class_code"],
            sr["absent_count"], ", ".join(sr["absent_dates"]),
            sr["late_count"], ", ".join(sr["late_dates"]),
            sr["violation_count"], ", ".join(sr["violation_dates"]),
            sr["medical_count"], ", ".join(sr["medical_dates"]),
        ]
        for col, v in enumerate(vals, 1):
            center = col in (1, 2, 3, 5, 7, 9)
            _apply_data_cell(ws3, row, col, v, s, center=center, alt=alt)

    _set_col_widths(ws3, [10, 8, 8, 25, 8, 25, 8, 25, 8, 25])
    ws3.freeze_panes = "A2"

    # ==================== Sheet 4: 學生風險名單 ====================
    ws4 = wb.create_sheet("學生風險名單")
    risk_headers = ["學生", "班級", "總記名", "違規", "缺席", "遲到", "醫務室",
                    "首次記錄", "最近記錄", "涉及天數", "風險標記"]
    _apply_header_row(ws4, 1, risk_headers, s)

    row = 1
    for rs in agg["risk_students"]:
        row += 1
        flags_text = "; ".join(
            f.replace("violation>=3", "違規>=3").replace("absent>=3", "缺席>=3").replace("7d>=3", "7天內>=3次")
            for f in rs["risk_flags"]
        )
        vals = [
            rs["name"], rs["class_code"], rs["total_incidents"],
            rs["violation_count"], rs["absent_count"], rs["late_count"], rs["medical_count"],
            rs["first_date"], rs["last_date"], rs["involved_days"], flags_text,
        ]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws4, row, col, v, s)
            c.fill = s["danger_fill"]

    if not agg["risk_students"]:
        ws4.cell(row=2, column=1, value="沒有高風險學生").font = s["body_font"]

    _set_col_widths(ws4, [10, 8, 8, 8, 8, 8, 8, 12, 12, 8, 25])
    ws4.freeze_panes = "A2"

    # ==================== Sheet 5: 班級統計 ====================
    ws5 = wb.create_sheet("班級統計")
    cs_data = agg["class_stats"]
    classes = cs_data.get("classes", [])
    dates = cs_data.get("dates", [])
    pivot = cs_data.get("data", {})

    # Header row 1: merged class names
    ws5.cell(row=1, column=1, value="日期").font = s["header_font"]
    ws5.cell(row=1, column=1).fill = s["header_fill"]
    ws5.cell(row=1, column=1).border = s["border"]
    for ci, cc in enumerate(classes):
        col_start = 2 + ci * 2
        col_end = col_start + 1
        ws5.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        c = ws5.cell(row=1, column=col_start, value=cc)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]
        ws5.cell(row=1, column=col_end).border = s["border"]

    # Header row 2: 紀律/整潔 sub-headers
    ws5.cell(row=2, column=1, value="").border = s["border"]
    for ci in range(len(classes)):
        col_d = 2 + ci * 2
        col_c = col_d + 1
        cd = ws5.cell(row=2, column=col_d, value="紀律")
        cd.font = s["header_font"]
        cd.fill = s["header_fill"]
        cd.alignment = s["center"]
        cd.border = s["border"]
        cc_cell = ws5.cell(row=2, column=col_c, value="整潔")
        cc_cell.font = s["header_font"]
        cc_cell.fill = s["header_fill"]
        cc_cell.alignment = s["center"]
        cc_cell.border = s["border"]

    # Data rows
    for ri, d in enumerate(dates):
        row = 3 + ri
        alt = ri % 2 == 1
        _apply_data_cell(ws5, row, 1, d, s, alt=alt)
        for ci, cc in enumerate(classes):
            col_d = 2 + ci * 2
            col_c = col_d + 1
            vals = pivot.get(d, {}).get(cc)
            if vals:
                c_d = _apply_data_cell(ws5, row, col_d, vals["avg_discipline"], s, alt=alt)
                f = _rating_fill(vals["avg_discipline"], s)
                if f:
                    c_d.fill = f
                c_c = _apply_data_cell(ws5, row, col_c, vals["avg_cleanliness"], s, alt=alt)
                f = _rating_fill(vals["avg_cleanliness"], s)
                if f:
                    c_c.fill = f
            else:
                _apply_data_cell(ws5, row, col_d, "", s, alt=alt)
                _apply_data_cell(ws5, row, col_c, "", s, alt=alt)

    # Column widths
    widths = [12] + [8] * (len(classes) * 2)
    _set_col_widths(ws5, widths)
    ws5.freeze_panes = "B3"

    # ==================== Sheet 6: 節數分析 ====================
    ws6 = wb.create_sheet("節數分析")
    pa_headers = ["節數", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次",
                  "違規人次", "表揚人次", "醫務室人次", "違規率", "表揚率"]
    _apply_header_row(ws6, 1, pa_headers, s)

    row = 1
    for pa in agg["period_analysis"]:
        row += 1
        vals = [pa["period_label"], pa["entry_count"], pa["avg_discipline"], pa["avg_cleanliness"],
                pa["absent_count"], pa["late_count"], pa["violation_count"], pa["praise_count"],
                pa["medical_count"], pa["violation_rate"], pa["praise_rate"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws6, row, col, v, s, alt=(row % 2 == 0))
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws6, [10, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8])

    # ==================== Sheet 7: 科目分析 ====================
    ws7 = wb.create_sheet("科目分析")
    sa_headers = ["科目", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次",
                  "違規人次", "表揚人次", "醫務室人次", "違規率", "表揚率"]
    _apply_header_row(ws7, 1, sa_headers, s)

    row = 1
    for sa in agg["subject_analysis"]:
        row += 1
        vals = [sa["subject"], sa["entry_count"], sa["avg_discipline"], sa["avg_cleanliness"],
                sa["absent_count"], sa["late_count"], sa["violation_count"], sa["praise_count"],
                sa["medical_count"], sa["violation_rate"], sa["praise_rate"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws7, row, col, v, s, alt=(row % 2 == 0))
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws7, [14, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8])

    # ==================== Sheet 8: 原因分析 ====================
    ws8 = wb.create_sheet("原因分析")
    ra_headers = ["原因", "類別", "總人次", "涉及學生數", "涉及班級數", "涉及日期數", "首次出現", "最近出現"]
    _apply_header_row(ws8, 1, ra_headers, s)

    row = 1
    for ra in agg["reason_analysis"]:
        row += 1
        vals = [ra["reason"], ra["category"], ra["total_count"], ra["student_count"],
                ra["class_count"], ra["date_count"], ra["first_date"], ra["last_date"]]
        for col, v in enumerate(vals, 1):
            _apply_data_cell(ws8, row, col, v, s, alt=(row % 2 == 0))

    _set_col_widths(ws8, [20, 10, 8, 10, 10, 10, 12, 12])

    # ==================== Sheet 9: 星期模式 ====================
    ws9 = wb.create_sheet("星期模式")
    wd_headers = ["星期", "記錄數", "平均紀律", "平均整潔", "缺席人次", "遲到人次",
                  "違規人次", "表揚人次", "違規率", "表揚率"]
    _apply_header_row(ws9, 1, wd_headers, s)

    row = 1
    for wa in agg["weekday_analysis"]:
        row += 1
        vals = [wa["weekday_label"], wa["entry_count"], wa["avg_discipline"], wa["avg_cleanliness"],
                wa["absent_count"], wa["late_count"], wa["violation_count"], wa["praise_count"],
                wa["violation_rate"], wa["praise_rate"]]
        for col, v in enumerate(vals, 1):
            c = _apply_data_cell(ws9, row, col, v, s, alt=(row % 2 == 0))
            if col in (3, 4):
                f = _rating_fill(v, s)
                if f:
                    c.fill = f

    _set_col_widths(ws9, [10, 8, 8, 8, 8, 8, 8, 8, 8, 8])

    # ==================== Sheet 10: 填寫人分析 ====================
    ws10 = wb.create_sheet("填寫人分析")
    sub_headers = ["填寫人", "記錄數", "涉及班級數", "平均紀律", "平均整潔",
                   "缺席人次", "遲到人次", "違規人次", "表揚人次"]
    _apply_header_row(ws10, 1, sub_headers, s)

    row = 1
    for suba in agg["submitter_analysis"]:
        row += 1
        vals = [suba["submitter"], suba["entry_count"], suba["class_count"],
                suba["avg_discipline"], suba["avg_cleanliness"],
                suba["absent_count"], suba["late_count"],
                suba["violation_count"], suba["praise_count"]]
        for col, v in enumerate(vals, 1):
            _apply_data_cell(ws10, row, col, v, s, alt=(row % 2 == 0))

    _set_col_widths(ws10, [14, 8, 10, 8, 8, 8, 8, 8, 8])

    return wb


@router.get("/api/class-diary/review/export-range")
async def export_entries_range_xlsx(
    request: Request,
    start_date: str = Query(..., description="開始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="結束日期 YYYY-MM-DD"),
):
    """匯出日期範圍內所有班級的記錄為 XLSX（含學生統計等 10 個工作表）"""
    from urllib.parse import quote

    try:
        export_user, _ = _verify_reviewer(request)

        if start_date > end_date:
            return error_response("INVALID_DATE_RANGE", "開始日期不能晚於結束日期", status_code=400)

        service = _get_service()
        entries = service.get_all_entries_by_date_range(start_date, end_date)
        agg = service.aggregate_date_range(entries)
        # Attach raw entries for the detail sheet
        agg["_raw_entries"] = entries

        wb = _build_date_range_xlsx(agg, start_date, end_date)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        _audit.log("EXPORT", "entry", export_user, target_id=f"{start_date}~{end_date}", metadata={"type": "date_range"})

        filename = f"課室日誌_{start_date}_至_{end_date}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
        )

    except AppException as e:
        return error_response(e.code, e.message, status_code=e.status_code)
    except Exception as e:
        logger.exception("匯出日期範圍 XLSX 失敗")
        return error_response("SERVER_ERROR", str(e), status_code=500)
